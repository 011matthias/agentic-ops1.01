# /// script
# requires-python = ">=3.11"
# dependencies = ["openpyxl>=3.1"]
# ///
"""Safe append + status for the Brisken hourly-agreement hours tracker.

Backs the `/comd_brisken-hours` command. The agent decides WHICH rows to log
(scope-based session estimates grounded in git history + session checkpoints);
this tool does the mechanical, gotcha-safe write into
`workspace/hours-tracker.xlsx`.

The workbook has two engagement tabs, both billed at the B5 rate (EUR 14/hr):
  - "Timesheet"        table HoursLog    -> Brisken Expense Reconciliation (p1)
  - "Lead Generation"  table LeadGenLog  -> Brisken Lead Generation        (p2)

Why a tool and not inline openpyxl every time (the recurring gotchas this
encapsulates, from feedback_hours_tracker_format):
  - a programmatically-added row does NOT inherit the Hours/Earnings formulas
    (they are not Excel "calculated columns") -> they are set explicitly here;
  - a fresh cell defaults to number_format "General" -> the per-cell style is
    copied from the last existing data row so Earnings renders "28.00 EUR"
    not a bare "28", and Start/End/Hours keep their formats;
  - the Excel Table `ref` must be extended or the KPIs (structured refs
    LeadGenLog[...]) and the Billable dropdown miss the new rows;
  - concurrent same-day sessions shift row numbers -> the last data row is
    re-found at write time, never assumed.

Modes:
  --status                 print last-logged date/time + computed totals per tab
  --add  <rows.json>       append rows (idempotent), verify (no CSV mirror)
  --add  -                 read the rows JSON from stdin
  --export-csv             opt-in: write the gitignored CSV mirrors from the xlsx
  --dry-run                with --add: print what WOULD be written, do not save

rows.json schema (a JSON list):
  [{"tab": "Lead Generation", "date": "2026-06-21", "start": "12:30",
    "end": "14:00", "task": "expandable cards, bigger hero", "billable": "Yes"}]
  `billable` is optional (default "Yes"). `tab` accepts the full sheet name or
  the shorthands "lead"/"leadgen"/"p2" and "time"/"timesheet"/"recon"/"p1".

Exit codes: 0 ok | 2 workbook locked (close Excel, retry) | 1 other error.
"""

import argparse
import csv
import datetime as dt
import json
import re
import sys
from copy import copy
from pathlib import Path

from openpyxl import load_workbook

# Monthly workbooks live in this folder, one dated file per month
# (hours-tracker-YYYY-MM-<name>.xlsx). The tool auto-resolves the latest month,
# so a new month's rollover needs no code change.
FOLDER = Path("workspace/hours-tracker")
_DATED = re.compile(r"hours-tracker-(\d{4})-(\d{2})")


def current_xlsx() -> "Path | None":
    """Latest dated month workbook in FOLDER, keyed on the YYYY-MM in its name."""
    cands = []
    for p in FOLDER.glob("hours-tracker-*.xlsx"):
        m = _DATED.search(p.name)
        if m:
            cands.append((m.group(1) + m.group(2), p))
    return max(cands)[1] if cands else None


XLSX = current_xlsx()
RATE_CELL = "B5"
HEADER_ROW = 7          # the column-header row; data starts at 8
FIRST_DATA_ROW = 8

# The engagement is identified by its Excel TABLE name, not the sheet title:
# the sheet gets renamed between months (July renamed "Timesheet" ->
# "Expense Reconciliation") while the table name is stable because the KPI
# structured refs depend on it. TABS is bound to the live sheet titles by
# bind_tabs() right after the workbook loads.
_TAB_SPECS = {
    "LeadGenLog": {
        "csv_base": "hours-lead-generation",
        "aliases": {"lead", "leadgen", "lead-gen", "lg", "p2", "lead generation"},
    },
    "HoursLog": {
        "csv_base": "hours-timesheet",
        "aliases": {"time", "timesheet", "recon", "expense", "expense-recon", "p1",
                    "expense reconciliation"},
    },
}

TABS: dict = {}          # live sheet title -> {"table", "csv_base", "aliases"}


def bind_tabs(wb) -> None:
    """Map each engagement table to whatever sheet currently holds it."""
    TABS.clear()
    for ws in wb.worksheets:
        for tname in ws.tables:
            if tname in _TAB_SPECS:
                TABS[ws.title] = {**_TAB_SPECS[tname], "table": tname}
    missing = {t for t in _TAB_SPECS} - {m["table"] for m in TABS.values()}
    if missing:
        raise SystemExit(f"{XLSX.name}: no sheet carries table(s) {sorted(missing)}")


def csv_path(sheet: str) -> Path:
    """Per-month CSV mirror path, derived from the resolved workbook name:
    hours-tracker-2026-07-july.xlsx -> hours-lead-generation-2026-07-july.csv"""
    token = XLSX.stem[len("hours-tracker"):]   # e.g. "-2026-07-july"
    return XLSX.parent / f'{TABS[sheet]["csv_base"]}{token}.csv'


def resolve_tab(name: str) -> str:
    n = (name or "").strip().lower()
    for sheet, meta in TABS.items():
        if n == sheet.lower() or n in meta["aliases"]:
            return sheet
    raise SystemExit(f"unknown tab {name!r}; use one of {list(TABS)} or an alias")


def hours_formula(r: int) -> str:
    return (f'=IF(AND(ISNUMBER(C{r}),ISNUMBER(D{r})),'
            f'IF(D{r}>=C{r},(D{r}-C{r})*24,(D{r}-C{r}+1)*24),"")')


def earnings_formula(r: int) -> str:
    return f'=IF(AND(F{r}="Yes",ISNUMBER(E{r})),E{r}*${RATE_CELL[0]}${RATE_CELL[1:]},0)'


def period_formula(table: str) -> str:
    """Live B4 period stamp off the table, so it never drifts as rows are added."""
    return (f'=TEXT(MIN({table}[Date]),"yyyy-mm-dd")&" to "&'
            f'TEXT(MAX({table}[Date]),"yyyy-mm-dd")')


def last_data_row(ws) -> int:
    """Last row whose Date (col A) is populated; HEADER_ROW if the table is empty."""
    r = FIRST_DATA_ROW
    last = HEADER_ROW
    while True:
        v = ws.cell(row=r, column=1).value
        if v in (None, ""):
            # tolerate one stray gap, but stop on two blanks in a row
            if ws.cell(row=r + 1, column=1).value in (None, ""):
                break
        else:
            last = r
        r += 1
        if r > 5000:
            break
    return last


def parse_date(s) -> dt.datetime:
    if isinstance(s, dt.datetime):
        return s
    return dt.datetime.strptime(str(s), "%Y-%m-%d")


def parse_time(s) -> dt.time:
    if isinstance(s, dt.time):
        return s
    return dt.datetime.strptime(str(s), "%H:%M").time()


def span_hours(start: dt.time, end: dt.time) -> float:
    s = start.hour + start.minute / 60
    e = end.hour + end.minute / 60
    if e < s:
        e += 24
    return round(e - s, 2)


def read_rows(ws):
    """Return [(row_idx, date, task, start, end, billable, notes)] for data rows."""
    out = []
    for r in range(FIRST_DATA_ROW, last_data_row(ws) + 1):
        date = ws.cell(row=r, column=1).value
        if date in (None, ""):
            continue
        out.append((
            r, date,
            ws.cell(row=r, column=2).value,
            ws.cell(row=r, column=3).value,
            ws.cell(row=r, column=4).value,
            ws.cell(row=r, column=6).value,
            ws.cell(row=r, column=8).value,
        ))
    return out


def existing_keys(ws) -> set:
    keys = set()
    for _r, date, task, start, _end, _bill, _notes in read_rows(ws):
        keys.add((str(date)[:10], str(start), (task or "").strip()))
    return keys


def fmt_hours(h: float) -> str:
    s = f"{h:.2f}"
    return s[:-1] if s.endswith("0") else s   # 2.00->2.0, 2.50->2.5, 1.75->1.75


# ----------------------------------------------------------------------------- status

def cmd_status(wb) -> int:
    rate = wb[next(iter(TABS))][RATE_CELL].value
    print(f"{XLSX.name}  (rate {RATE_CELL} = EUR {rate}/hr)\n")
    for sheet in TABS:
        ws = wb[sheet]
        rows = read_rows(ws)
        tot = bill = 0.0
        for _r, _date, _task, start, end, billable, _notes in rows:
            if isinstance(start, dt.time) and isinstance(end, dt.time):
                h = span_hours(start, end)
                tot += h
                if str(billable).strip().lower() == "yes":
                    bill += h
        if rows:
            r, date, task, start, end, *_ = rows[-1]
            last = (f"{str(date)[:10]} {start.strftime('%H:%M') if isinstance(start, dt.time) else start}"
                    f"-{end.strftime('%H:%M') if isinstance(end, dt.time) else end}  "
                    f"\"{task}\"  (row {r})")
        else:
            last = "(no entries)"
        print(f"  [{sheet}]  {len(rows)} rows | total {fmt_hours(tot)}h | "
              f"billable {fmt_hours(bill)}h = EUR {bill * (rate or 0):.2f}")
        print(f"      last logged: {last}\n")
    return 0


# ----------------------------------------------------------------------------- add

def cmd_add(wb, specs: list, dry_run: bool) -> int:
    # group by resolved tab
    by_tab: dict[str, list] = {}
    for s in specs:
        by_tab.setdefault(resolve_tab(s["tab"]), []).append(s)

    wrote = 0
    for sheet, items in by_tab.items():
        ws = wb[sheet]
        table = ws.tables[TABS[sheet]["table"]]
        keys = existing_keys(ws)
        src = last_data_row(ws)              # style donor (an existing data row)
        if src < FIRST_DATA_ROW:
            # Fresh/empty table (a new month's tracker): no data row to donate
            # style/position. Use the first table-body row (kept as a formatted
            # blank buffer) as the style donor and start writing there.
            donor_row = FIRST_DATA_ROW
            cursor = FIRST_DATA_ROW - 1
        else:
            donor_row = src
            cursor = src

        for s in items:
            date = parse_date(s["date"])
            start = parse_time(s["start"])
            end = parse_time(s["end"])
            task = s["task"].strip()
            billable = s.get("billable", "Yes")
            key = (date.strftime("%Y-%m-%d"), str(start), task)
            if key in keys:
                print(f"[{sheet}] skip (already logged): {key[0]} {task}")
                continue
            cursor += 1
            r = cursor
            vals = {1: date, 2: task, 3: start, 4: end,
                    5: hours_formula(r), 6: billable, 7: earnings_formula(r), 8: None}
            h = span_hours(start, end)
            print(f"[{sheet}] row {r}: {key[0]} {start.strftime('%H:%M')}-"
                  f"{end.strftime('%H:%M')} ({fmt_hours(h)}h) {task}")
            if dry_run:
                continue
            for c in range(1, 9):
                donor = ws.cell(row=donor_row, column=c)
                cell = ws.cell(row=r, column=c)
                cell._style = copy(donor._style)
                cell.value = vals[c]
            keys.add(key)
            wrote += 1

        if not dry_run and cursor > src:
            table.ref = f"A{HEADER_ROW}:H{cursor}"
            # keep the B4 period stamp live off the table (was a static string
            # that silently drifted; make it a formula so it self-updates)
            ws["B4"] = period_formula(TABS[sheet]["table"])
            # keep the Billable dropdown covering the new rows (cosmetic, best-effort).
            # Billable is column F (E=Hours, F=Billable, G=Earnings); never G.
            try:
                for dv in ws.data_validations.dataValidation:
                    if dv.type == "list" and dv.formula1 and "Yes" in dv.formula1:
                        dv.sqref = f"F{FIRST_DATA_ROW}:F{cursor}"
            except Exception:
                pass

    if dry_run:
        print("\n(dry-run; nothing written)")
        return 0
    if not wrote:
        print("nothing new to write")
        return 0

    try:
        wb.save(XLSX)
    except PermissionError:
        print(f"LOCKED: {XLSX.name} is open in Excel. Close it and retry.",
              file=sys.stderr)
        return 2
    print(f"\nwrote {wrote} row(s) to {XLSX}")
    return 0


# ----------------------------------------------------------------------------- verify + csv

def verify(specs: list) -> bool:
    wb = load_workbook(XLSX)
    bind_tabs(wb)
    ok = True
    by_tab: dict[str, list] = {}
    for s in specs:
        by_tab.setdefault(resolve_tab(s["tab"]), []).append(s)
    for sheet, items in by_tab.items():
        ws = wb[sheet]
        present = existing_keys(ws)
        ref = ws.tables[TABS[sheet]["table"]].ref
        last = last_data_row(ws)
        if not ref.endswith(f"H{last}"):
            print(f"VERIFY: [{sheet}] table ref {ref} does not reach last data row {last}",
                  file=sys.stderr)
            ok = False
        for s in items:
            key = (parse_date(s["date"]).strftime("%Y-%m-%d"),
                   str(parse_time(s["start"])), s["task"].strip())
            if key not in present:
                print(f"VERIFY: [{sheet}] missing {key}", file=sys.stderr)
                ok = False
        # formula presence on the last rows
        for r in range(FIRST_DATA_ROW, last + 1):
            if ws.cell(row=r, column=1).value in (None, ""):
                continue
            if not str(ws.cell(row=r, column=5).value).startswith("=IF"):
                print(f"VERIFY: [{sheet}] row {r} Hours not a formula", file=sys.stderr)
                ok = False
    return ok


def export_csv(wb) -> None:
    for sheet in TABS:
        ws = wb[sheet]
        rate = ws[RATE_CELL].value or 0
        rows = read_rows(ws)
        tot_h = tot_e = 0.0
        out = [["Date", "Task", "Start", "End", "Hours", "Billable", "Earnings (EUR)"]]
        for _r, date, task, start, end, billable, _notes in rows:
            if not (isinstance(start, dt.time) and isinstance(end, dt.time)):
                continue
            h = span_hours(start, end)
            e = h * rate if str(billable).strip().lower() == "yes" else 0.0
            tot_h += h
            tot_e += e
            out.append([str(date)[:10], task,
                        start.strftime("%H:%M"), end.strftime("%H:%M"),
                        fmt_hours(h), billable, f"{e:.2f}"])
        out_csv = csv_path(sheet)
        with out_csv.open("w", newline="", encoding="utf-8") as fh:
            w = csv.writer(fh)
            w.writerows(out)
            w.writerow([])
            w.writerow(["", "TOTAL", "", "", fmt_hours(tot_h), "", f"{tot_e:.2f}"])
        print(f"refreshed {out_csv} ({len(rows)} rows, {fmt_hours(tot_h)}h)")


# ----------------------------------------------------------------------------- main

def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    g = ap.add_mutually_exclusive_group(required=True)
    g.add_argument("--status", action="store_true")
    g.add_argument("--add", metavar="ROWS_JSON", help="path to rows JSON, or - for stdin")
    g.add_argument("--export-csv", action="store_true")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    if XLSX is None or not XLSX.exists():
        print(f"no dated hours-tracker workbook found in {FOLDER}/", file=sys.stderr)
        return 1

    try:
        wb = load_workbook(XLSX)
    except PermissionError:
        print(f"LOCKED: {XLSX.name} is open in Excel. Close it and retry.",
              file=sys.stderr)
        return 2

    bind_tabs(wb)

    if args.status:
        return cmd_status(wb)

    if args.export_csv:
        export_csv(wb)
        return 0

    # --add
    raw = sys.stdin.read() if args.add == "-" else Path(args.add).read_text(encoding="utf-8")
    specs = json.loads(raw)
    if isinstance(specs, dict):
        specs = [specs]
    for s in specs:
        for field in ("tab", "date", "start", "end", "task"):
            if field not in s:
                print(f"row missing required field {field!r}: {s}", file=sys.stderr)
                return 1

    rc = cmd_add(wb, specs, args.dry_run)
    if rc != 0 or args.dry_run:
        return rc
    if not verify(specs):
        print("WROTE BUT VERIFY MISMATCH", file=sys.stderr)
        return 1
    # CSV mirrors are opt-in via --export-csv; --add no longer regenerates them
    # (owner 2026-07-08: the mirrors are unwanted, deleting them must stick).
    print("verified.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
