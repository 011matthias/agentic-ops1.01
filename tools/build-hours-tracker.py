# /// script
# requires-python = ">=3.11"
# dependencies = ["openpyxl>=3.1"]
# ///
"""Build a blank hours tracker at workspace/hours-tracker.xlsx."""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

OUT = Path("workspace/hours-tracker.xlsx")
ROWS = 500

HEADERS = ["Date", "Project", "Task", "Start", "End", "Hours", "Billable", "Notes"]
PROJECTS = ["agentic-ops", "brisken"]

wb = Workbook()
ws = wb.active
ws.title = "Log"

header_font = Font(name="Arial", bold=True, color="FFFFFF")
header_fill = PatternFill("solid", start_color="305496")
header_align = Alignment(horizontal="center", vertical="center")
thin = Side(border_style="thin", color="D9D9D9")
cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

for col, name in enumerate(HEADERS, start=1):
    c = ws.cell(row=1, column=col, value=name)
    c.font = header_font
    c.fill = header_fill
    c.alignment = header_align

widths = {"A": 12, "B": 16, "C": 42, "D": 10, "E": 10, "F": 10, "G": 11, "H": 50}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

date_fmt = "yyyy-mm-dd"
time_fmt = "hh:mm"
hours_fmt = "0.00"

for r in range(2, ROWS + 2):
    ws.cell(row=r, column=1).number_format = date_fmt
    ws.cell(row=r, column=4).number_format = time_fmt
    ws.cell(row=r, column=5).number_format = time_fmt
    f = (
        f'=IF(AND(ISNUMBER(D{r}),ISNUMBER(E{r})),'
        f'IF(E{r}>=D{r},(E{r}-D{r})*24,(E{r}-D{r}+1)*24),"")'
    )
    hc = ws.cell(row=r, column=6, value=f)
    hc.number_format = hours_fmt
    for col in range(1, 9):
        ws.cell(row=r, column=col).border = cell_border
        ws.cell(row=r, column=col).font = Font(name="Arial")

project_dv = DataValidation(
    type="list",
    formula1=f'"{",".join(PROJECTS)}"',
    allow_blank=True,
    showDropDown=False,
)
project_dv.error = "Project must be agentic-ops or brisken"
project_dv.errorTitle = "Invalid project"
project_dv.prompt = "Pick a project"
project_dv.promptTitle = "Project"
ws.add_data_validation(project_dv)
project_dv.add(f"B2:B{ROWS + 1}")

billable_dv = DataValidation(
    type="list",
    formula1='"Yes,No"',
    allow_blank=True,
    showDropDown=False,
)
ws.add_data_validation(billable_dv)
billable_dv.add(f"G2:G{ROWS + 1}")

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:H{ROWS + 1}"

summary_row = ROWS + 3
ws.cell(row=summary_row, column=5, value="Total hours").font = Font(name="Arial", bold=True)
ws.cell(row=summary_row, column=5).alignment = Alignment(horizontal="right")
tot = ws.cell(row=summary_row, column=6, value=f"=SUM(F2:F{ROWS + 1})")
tot.font = Font(name="Arial", bold=True)
tot.number_format = hours_fmt

ws.cell(row=summary_row + 1, column=5, value="agentic-ops").font = Font(name="Arial", bold=True)
ws.cell(row=summary_row + 1, column=5).alignment = Alignment(horizontal="right")
ao = ws.cell(
    row=summary_row + 1,
    column=6,
    value=f'=SUMIF(B2:B{ROWS + 1},"agentic-ops",F2:F{ROWS + 1})',
)
ao.font = Font(name="Arial", bold=True)
ao.number_format = hours_fmt

ws.cell(row=summary_row + 2, column=5, value="brisken").font = Font(name="Arial", bold=True)
ws.cell(row=summary_row + 2, column=5).alignment = Alignment(horizontal="right")
br = ws.cell(
    row=summary_row + 2,
    column=6,
    value=f'=SUMIF(B2:B{ROWS + 1},"brisken",F2:F{ROWS + 1})',
)
br.font = Font(name="Arial", bold=True)
br.number_format = hours_fmt

OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)
print(f"Wrote {OUT}")
