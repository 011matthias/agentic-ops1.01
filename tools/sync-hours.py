# /// script
# requires-python = ">=3.11"
# dependencies = ["openpyxl>=3.1"]
# ///
"""Sync workspace/hours-tracker.xlsx with git commit activity.

Idempotent. Tracks last-synced commit per repo in a hidden _meta sheet and only
appends new sessions on subsequent runs. Sessions cluster commits with gap
<= 30min and same project bucket. Each session row gets:

    Date  = local date of first commit
    Start = first commit time - 15min lead-in
    End   = last commit time + 5min wrap-up (min 20min total)
    Task  = pipe-joined commit subjects (truncated to 200 chars)
    Notes = "[auto] <short-hashes>"  -- this is the dedup key on re-run

Manual rows (without "[auto]" in Notes) are never touched.

Run:  uv run tools/sync-hours.py
Flags: --since "60 days ago"   # cap first-run history
       --dry-run                # print rows, don't write
"""

import argparse
import subprocess
from datetime import datetime, timedelta, time
from pathlib import Path

from openpyxl import load_workbook

REPOS = [
    ("agentic-ops", Path("c:/Users/neuma_p1qrsic/Repo/agentic-ops1"), "011matthias"),
]

SESSION_GAP_MIN = 30
LEAD_IN_MIN = 15
WRAP_UP_MIN = 5
MIN_SESSION_MIN = 20

XLSX = Path("workspace/hours-tracker.xlsx")
PROJECT_VALUES = ["agentic-ops", "brisken", "openclaw"]


def git_commits(repo: Path, author: str, since_hash: str | None, since_date: str | None):
    cmd = [
        "git", "-C", str(repo), "log",
        f"--author={author}",
        "--pretty=format:%H%x00%aI%x00%s",
        "--name-only",
        "--no-merges",
        "--reverse",
    ]
    if since_hash:
        cmd.append(f"{since_hash}..HEAD")
    elif since_date:
        cmd.append(f"--since={since_date}")
    out = subprocess.run(cmd, check=True, capture_output=True, text=True).stdout
    if not out.strip():
        return []
    commits = []
    for block in out.split("\n\n"):
        lines = [l for l in block.strip().split("\n") if l]
        if not lines:
            continue
        header = lines[0].split("\x00")
        if len(header) != 3:
            continue
        h, ts, subj = header
        commits.append({
            "hash": h,
            "ts": datetime.fromisoformat(ts).astimezone().replace(tzinfo=None),
            "subject": subj,
            "paths": lines[1:],
        })
    return commits


def bucket(repo_name: str, paths: list[str]) -> str:
    if repo_name == "openclaw":
        return "openclaw"
    if any(p.startswith("workspace/clients/brisken/") for p in paths):
        return "brisken"
    return "agentic-ops"


def cluster(commits: list[dict], repo_name: str) -> list[dict]:
    annotated = [{**c, "project": bucket(repo_name, c["paths"])} for c in commits]
    annotated.sort(key=lambda c: c["ts"])
    sessions: list[dict] = []
    for c in annotated:
        gap_ok = (
            sessions
            and sessions[-1]["project"] == c["project"]
            and (c["ts"] - sessions[-1]["end"]).total_seconds() / 60 <= SESSION_GAP_MIN
        )
        if gap_ok:
            sessions[-1]["end"] = c["ts"]
            sessions[-1]["commits"].append(c)
        else:
            sessions.append({
                "project": c["project"],
                "start": c["ts"],
                "end": c["ts"],
                "commits": [c],
            })
    return sessions


def to_row(s: dict) -> dict:
    start = s["start"] - timedelta(minutes=LEAD_IN_MIN)
    end = s["end"] + timedelta(minutes=WRAP_UP_MIN)
    if (end - start).total_seconds() / 60 < MIN_SESSION_MIN:
        end = start + timedelta(minutes=MIN_SESSION_MIN)
    subjects = " | ".join(c["subject"] for c in s["commits"])
    if len(subjects) > 200:
        subjects = subjects[:197] + "..."
    hashes = ", ".join(c["hash"][:7] for c in s["commits"])
    return {
        "date": start.date(),
        "project": s["project"],
        "task": subjects,
        "start": time(start.hour, start.minute),
        "end": time(end.hour, end.minute),
        "notes": f"[auto] {hashes}",
    }


def ensure_meta(wb):
    if "_meta" in wb.sheetnames:
        return wb["_meta"]
    meta = wb.create_sheet("_meta")
    meta["A1"] = "repo"
    meta["B1"] = "last_synced_commit"
    meta.sheet_state = "hidden"
    return meta


def read_meta(meta):
    out = {}
    for row in range(2, meta.max_row + 1):
        r = meta.cell(row=row, column=1).value
        h = meta.cell(row=row, column=2).value
        if r:
            out[r] = h
    return out


def write_meta(meta, meta_map):
    for row in range(2, meta.max_row + 1):
        for col in range(1, 3):
            meta.cell(row=row, column=col).value = None
    for i, (repo_name, h) in enumerate(meta_map.items(), start=2):
        meta.cell(row=i, column=1, value=repo_name)
        meta.cell(row=i, column=2, value=h)


def update_project_dv(ws):
    for dv in ws.data_validations.dataValidation:
        if dv.type == "list" and dv.formula1 and "agentic-ops" in dv.formula1:
            dv.formula1 = f'"{",".join(PROJECT_VALUES)}"'
            return


def find_next_row(ws) -> int:
    r = 2
    while ws.cell(row=r, column=1).value is not None:
        r += 1
    return r


def collect_existing_auto_hashes(ws) -> set[str]:
    """Hashes of commits already represented in [auto] rows (for safety dedup)."""
    seen = set()
    r = 2
    while ws.cell(row=r, column=1).value is not None:
        notes = ws.cell(row=r, column=8).value or ""
        if isinstance(notes, str) and notes.startswith("[auto]"):
            for h in notes.replace("[auto]", "").split(","):
                h = h.strip()
                if h:
                    seen.add(h)
        r += 1
    return seen


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--since", default="90 days ago",
                    help="git --since= for first-run history (default: 90 days ago)")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    wb = load_workbook(XLSX)
    ws = wb["Log"]
    meta = ensure_meta(wb)
    meta_map = read_meta(meta)
    existing_hashes = collect_existing_auto_hashes(ws)
    update_project_dv(ws)

    next_row = find_next_row(ws)
    appended = 0

    for repo_name, repo_path, author in REPOS:
        if not (repo_path / ".git").exists():
            print(f"skip {repo_name}: not a git repo at {repo_path}")
            continue
        since_hash = meta_map.get(repo_name)
        commits = git_commits(repo_path, author, since_hash, args.since if not since_hash else None)
        # Dedup against existing hashes (in case meta is missing/wrong)
        commits = [c for c in commits if c["hash"][:7] not in existing_hashes]
        if not commits:
            print(f"{repo_name}: nothing new")
            continue
        sessions = cluster(commits, repo_name)
        for s in sessions:
            row = to_row(s)
            if args.dry_run:
                print(f"  {row['date']} {row['project']:<12} {row['start']}-{row['end']} | {row['task'][:60]}")
                continue
            ws.cell(row=next_row, column=1, value=row["date"]).number_format = "yyyy-mm-dd"
            ws.cell(row=next_row, column=2, value=row["project"])
            ws.cell(row=next_row, column=3, value=row["task"])
            ws.cell(row=next_row, column=4, value=row["start"]).number_format = "hh:mm"
            ws.cell(row=next_row, column=5, value=row["end"]).number_format = "hh:mm"
            ws.cell(row=next_row, column=8, value=row["notes"])
            next_row += 1
            appended += 1
        meta_map[repo_name] = commits[-1]["hash"]
        print(f"{repo_name}: {len(commits)} commits -> {len(sessions)} sessions")

    if not args.dry_run:
        write_meta(meta, meta_map)
        wb.save(XLSX)
        print(f"Appended {appended} rows to {XLSX}")
        # Voice-pass the freshly-written cells. The hours tracker is
        # billing-visible; AI-tells in its cells (em-dashes, parenthetical
        # laundry-lists, fake-precision counts) are inappropriate (register
        # #160). post-write-gate.py never fires on this file (it is written by
        # this Bash-run script, not the Write/Edit tool), so check inline.
        # Advisory only; never blocks the sync.
        try:
            import json as _json
            import subprocess as _sp
            r = _sp.run(
                ["uv", "run", "tools/validate-output.py", str(XLSX), "--format", "json"],
                capture_output=True, text=True, timeout=60,
            )
            payload = _json.loads(r.stdout or "{}")
            if payload.get("total", 0):
                cats = ", ".join(f"{k}={v}" for k, v in payload.get("by_category", {}).items())
                print(f"[voice] {payload['total']} cell voice hit(s) in {XLSX}: {cats}")
                print("        Review for AI-tells; see feedback_human_voice_in_deliverables.")
        except Exception:
            pass
    else:
        print("(dry-run; no write)")


if __name__ == "__main__":
    main()
