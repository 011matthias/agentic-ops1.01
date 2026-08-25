# /// script
# requires-python = ">=3.11"
# dependencies = ["openpyxl>=3.1"]
# ///
"""Slice one ISO week out of the Brisken monthly hours tracker into its own workbook.

Backs the weekly time sheets Dirk approves (owner directive 2026-08-24: one sheet
per week, and NO work from before or after that week visible on ANY tab).

What that rules out, and this tool therefore guarantees:
  - no data row outside [Monday, Sunday] on any engagement tab;
  - no by-week overview row for any other week (the monthly book carries a
    five-Monday block; a weekly book carries exactly its own Monday);
  - no engagement tab at all when that engagement has no hours in the week
    (an empty tab still reads as "this engagement was in scope");
  - no empty filler rows inside the Excel table (the table ref is sized to the
    row count, so structured refs and the Billable dropdown match the data).

The month workbook stays the single source of truth; a weekly book is a derived
view and is safe to regenerate at any time.

Tabs are bound by TABLE name, never sheet title (titles get renamed between
months; the KPI structured refs depend on the table name).

Usage:
  --monday YYYY-MM-DD   week to build (default: the most recently completed week)
  --summary FILE.json   Week Summary content, see below
  --out FILE.xlsx       output path (default: weekly/hours-tracker-YYYY-MM-week-<mon><dd>-<dd>.xlsx)
  --dry-run             report what would be written, write nothing

Summary JSON:
  {"status": "Final",
   "what_was_done": {"<sheet title>": ["line", ...], ...},
   "focus": ["line", ...]}

Exit codes: 0 ok | 1 error (including hours that have nowhere to go).
"""

import argparse
import datetime as dt
import json
import re
import shutil
import sys
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parents[1]
TRACKER_DIR = ROOT / "workspace" / "hours-tracker"
WEEKLY_DIR = TRACKER_DIR / "weekly"
MONTH_RE = re.compile(r"hours-tracker-(\d{4})-(\d{2})-")

HEADER_ROW = 7
FIRST_DATA_ROW = 8
ACCENT = "FF1F4E79"
BAND = "FFD9E2F3"
EUR = '#,##0.00" €"'
RATE = 14


def month_books():
    return sorted(p for p in TRACKER_DIR.glob("hours-tracker-*.xlsx")
                  if MONTH_RE.search(p.name) and not p.name.startswith("~$"))


def last_completed_monday(today):
    """Monday of the last week that has fully ended (today's own week excluded)."""
    this_monday = today - dt.timedelta(days=today.weekday())
    return this_monday - dt.timedelta(days=7)


def tables_of(ws):
    return {t.name: t for t in ws.tables.values()}


def table_bounds(ref):
    m = re.match(r"[A-Z]+(\d+):[A-Z]+(\d+)", ref)
    return int(m.group(1)), int(m.group(2))


def read_rows(ws, table, lo, hi):
    """In-period data rows as plain values, sorted by (date, start)."""
    first, last = table_bounds(table.ref)
    out = []
    for r in range(first + 1, last + 1):
        date = ws.cell(row=r, column=1).value
        if not isinstance(date, (dt.datetime, dt.date)):
            continue
        d = date.date() if isinstance(date, dt.datetime) else date
        if not (lo <= d <= hi):
            continue
        out.append((d,
                    ws.cell(row=r, column=2).value,           # task
                    ws.cell(row=r, column=3).value,           # start
                    ws.cell(row=r, column=4).value,           # end
                    ws.cell(row=r, column=6).value or "Yes",  # billable
                    ws.cell(row=r, column=8).value))          # notes
    out.sort(key=lambda x: (x[0], x[2] or dt.time(0, 0)))
    return out


def hours_of(row):
    _, _, s, e, _billable, _notes = row
    if not (isinstance(s, dt.time) and isinstance(e, dt.time)):
        return 0.0
    a = s.hour * 60 + s.minute
    b = e.hour * 60 + e.minute
    if b <= a:
        b += 24 * 60
    return (b - a) / 60.0


def write_engagement(ws, table, rows, lo, hi):
    """Overwrite the tab with exactly `rows`; resize table, dropdown, by-week block."""
    tname = table.name
    _first, old_last = table_bounds(table.ref)

    for i, (date, task, start, end, billable, notes) in enumerate(rows):
        r = FIRST_DATA_ROW + i
        if r > old_last:  # only if a later month book donated rows
            for c in range(1, 9):
                ws.cell(row=r, column=c)._style = copy(
                    ws.cell(row=FIRST_DATA_ROW, column=c)._style)
        ws.cell(row=r, column=1, value=dt.datetime(date.year, date.month, date.day))
        ws.cell(row=r, column=2, value=task)
        ws.cell(row=r, column=3, value=start)
        ws.cell(row=r, column=4, value=end)
        ws.cell(row=r, column=5, value=(
            "=IF(AND(ISNUMBER(C{r}),ISNUMBER(D{r})),"
            "IF(D{r}>=C{r},(D{r}-C{r})*24,(D{r}-C{r}+1)*24),\"\")").format(r=r))
        ws.cell(row=r, column=6, value=billable)
        ws.cell(row=r, column=7, value=(
            "=IF(AND(F{r}=\"Yes\",ISNUMBER(E{r})),E{r}*$B$5,0)").format(r=r))
        ws.cell(row=r, column=8, value=notes)

    new_last = FIRST_DATA_ROW + len(rows) - 1
    for r in range(new_last + 1, max(old_last, ws.max_row) + 1):  # nothing lingers below
        for c in range(1, 9):
            ws.cell(row=r, column=c).value = None

    table.ref = "A{}:H{}".format(HEADER_ROW, new_last)
    ws.data_validations.dataValidation = []
    dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add("F{}:F{}".format(FIRST_DATA_ROW, new_last))

    ws["B4"] = "{:%Y-%m-%d} to {:%Y-%m-%d}".format(lo, hi)

    # by-week block: this week only, plus a control check that has to foot
    ws["J9"] = dt.datetime(lo.year, lo.month, lo.day)
    ws["K9"] = ("=SUMPRODUCT(({t}[Date]-WEEKDAY({t}[Date],2)+1=$J9)*{t}[Hours])"
                .format(t=tname))
    ws["L9"] = ("=SUMPRODUCT(({t}[Date]-WEEKDAY({t}[Date],2)+1=$J9)*{t}[Earnings])"
                .format(t=tname))
    for r in range(10, 14):
        for c in (10, 11, 12):
            ws.cell(row=r, column=c).value = None
    ws["J11"] = "Control check"
    ws["J11"].font = Font(bold=True)
    ws["K11"] = '=IF(ROUND(K9-K3,2)=0,"ties to table","CHECK MISMATCH")'
    ws.freeze_panes = "A{}".format(FIRST_DATA_ROW)


def build_summary(wb, week_label, engagements, summary):
    ws = wb.create_sheet("Week Summary", 0)
    for col, width in (("A", 30), ("B", 12), ("C", 14), ("D", 70)):
        ws.column_dimensions[col].width = width
    ws.row_dimensions[1].height = 18
    band = PatternFill("solid", fgColor=BAND)

    ws["A1"] = "Brisken, Weekly Time Sheet"
    ws["A1"].font = Font(bold=True, size=14, color=ACCENT)
    for r, (label, value) in enumerate(
            (("Week", week_label),
             ("Prepared by", "Matthias Silva"),
             ("Hourly rate", RATE),
             ("Status", summary.get("status", "Final"))), start=3):
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws.cell(row=r, column=2, value=value)
    ws["B5"].number_format = EUR

    head = 8
    for c, text in enumerate(("Engagement", "Hours", "Earnings"), start=1):
        cell = ws.cell(row=head, column=c, value=text)
        cell.font, cell.fill = Font(bold=True), band
    last = head
    for last, (title, _tname) in enumerate(engagements, start=head + 1):
        ws.cell(row=last, column=1, value=title)
        ws.cell(row=last, column=2,
                value="='{}'!K3".format(title)).number_format = "0.00"
        ws.cell(row=last, column=3,
                value="='{}'!L3".format(title)).number_format = EUR
    total = last + 1
    ws.cell(row=total, column=1, value="Total").font = Font(bold=True)
    cell = ws.cell(row=total, column=2, value="=SUM(B{}:B{})".format(head + 1, last))
    cell.font, cell.number_format = Font(bold=True), "0.00"
    cell = ws.cell(row=total, column=3, value="=SUM(C{}:C{})".format(head + 1, last))
    cell.font, cell.number_format = Font(bold=True), EUR

    row = total + 2
    done = summary.get("what_was_done") or {}
    if done:
        cell = ws.cell(row=row, column=1, value="What was done")
        cell.font, cell.fill = Font(bold=True), band
        row += 1
        for title, _tname in engagements:
            lines = done.get(title) or []
            if not lines:
                continue
            ws.cell(row=row, column=1, value="{}:".format(title)).font = Font(bold=True)
            row += 1
            for line in lines:
                ws.cell(row=row, column=1, value="   - {}".format(line))
                row += 1
        row += 1
    focus = summary.get("focus") or []
    if focus:
        cell = ws.cell(row=row, column=1, value="Focus for the coming week")
        cell.font, cell.fill = Font(bold=True), band
        row += 1
        for line in focus:
            ws.cell(row=row, column=1, value="   - {}".format(line))
            row += 1
        row += 1
    cell = ws.cell(row=row, column=1, value="Detail per entry is on the engagement tabs.")
    cell.font = Font(size=10, color="FF666666")
    cell.alignment = Alignment(horizontal="left")


def default_out(lo, hi):
    stem = "hours-tracker-{:%Y-%m}-week-{}{:%d}-{:%d}".format(
        lo, lo.strftime("%b").lower(), lo, hi)
    return WEEKLY_DIR / (stem + ".xlsx")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--monday")
    ap.add_argument("--summary")
    ap.add_argument("--out")
    ap.add_argument("--dry-run", action="store_true")
    a = ap.parse_args()

    lo = dt.date.fromisoformat(a.monday) if a.monday else last_completed_monday(
        dt.date.today())
    if lo.weekday() != 0:
        print("error: {} is not a Monday".format(lo), file=sys.stderr)
        return 1
    hi = lo + dt.timedelta(days=6)
    week_label = "{:%Y-%m-%d} to {:%Y-%m-%d} (week {})".format(
        lo, hi, lo.isocalendar().week)

    books = month_books()
    if not books:
        print("error: no monthly workbook found", file=sys.stderr)
        return 1

    per_table = {}
    base, base_count = None, -1
    for book in books:
        wb = load_workbook(book)
        found = {}
        for ws in wb.worksheets:
            for name, table in tables_of(ws).items():
                rows = read_rows(ws, table, lo, hi)
                if rows:
                    found[name] = rows
        total = sum(len(v) for v in found.values())
        if total > base_count:
            base, base_count = book, total
        for name, rows in found.items():
            per_table.setdefault(name, []).extend(rows)
        wb.close()

    if not per_table:
        print("no hours logged in {} to {}; nothing to build".format(lo, hi))
        return 0
    for name in per_table:
        per_table[name].sort(key=lambda x: (x[0], x[2] or dt.time(0, 0)))

    out = Path(a.out) if a.out else default_out(lo, hi)
    summary = json.loads(Path(a.summary).read_text(encoding="utf-8")) if a.summary else {}

    grand = 0.0
    print("week {}   base: {}".format(week_label, base.name))
    for name, rows in per_table.items():
        h = sum(hours_of(r) for r in rows)
        grand += h
        print("  {:18} {:2} row(s)  {:6.2f} h".format(name, len(rows), h))
        for d, task, s, e, _b, _n in rows:
            print("      {:%a %d}  {:%H:%M}-{:%H:%M}  {}".format(d, s, e, task))
    print("  {:18}            {:6.2f} h = EUR {:,.2f}".format("TOTAL", grand, grand * RATE))
    if a.dry_run:
        print("\n(dry-run; nothing written)")
        return 0

    WEEKLY_DIR.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(base, out)
    wb = load_workbook(out)

    engagements = []
    for ws in list(wb.worksheets):
        names = tables_of(ws)
        if not names:
            continue
        name, table = next(iter(names.items()))
        rows = per_table.get(name) or []
        if not rows:
            wb.remove(ws)  # an engagement with no hours gets no tab
            continue
        write_engagement(ws, table, rows, lo, hi)
        engagements.append((ws.title, name))

    orphan = set(per_table) - {n for _t, n in engagements}
    if orphan:
        print("error: hours in {} have no tab in {}".format(sorted(orphan), base.name),
              file=sys.stderr)
        return 1

    build_summary(wb, week_label, engagements, summary)
    wb.save(out)
    print("\nwrote {}".format(out))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
