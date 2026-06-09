"""validate-output.py cell scan: AI-tells in xlsx/csv/tsv cells (register #160).

CSV cases use the stdlib reader (always run in CI). The xlsx cases importorskip
openpyxl so the suite is robust whether or not the CI env carries the dep.
"""
import importlib.util
import json
import subprocess
import sys
from pathlib import Path

import pytest

from hooklib import TOOLS


def _load():
    path = TOOLS / "validate-output.py"
    spec = importlib.util.spec_from_file_location("validate_output", path)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


vo = _load()

# The exact register #160 hours-tracker string (em-dash via escape so no literal
# em-dash byte lands in this source file).
REGISTER_160 = (
    "wrote the openai client wrapper — hooked up tests (14), live OpenAI "
    "smoke + cost verification; BLUEPRINT/README/ANNEALING updates "
    "(Protocol + OpenAIClient + Mock)"
)


def test_is_cell_file():
    assert vo.is_cell_file(Path("x.xlsx"))
    assert vo.is_cell_file(Path("x.csv"))
    assert vo.is_cell_file(Path("x.tsv"))
    assert not vo.is_cell_file(Path("x.md"))
    assert not vo.is_cell_file(Path("infrastructure.yaml"))  # yaml excluded


def test_csv_positive(tmp_path):
    p = tmp_path / "d.csv"
    p.write_text(
        'task,notes\nbuild,"cleaned the data (rows, cols, dupes) — done"\n',
        encoding="utf-8",
    )
    cats = {h["category"] for h in vo.check_cells(p)}
    assert "em-dash" in cats
    assert "parenthetical-laundry-list" in cats


def test_csv_negative_numeric(tmp_path):
    p = tmp_path / "data.csv"
    p.write_text("id,amount\n1,100\n2,200\n", encoding="utf-8")
    assert vo.check_cells(p) == []


def test_register_160_categories_via_csv(tmp_path):
    p = tmp_path / "h.csv"
    p.write_text('task\n"' + REGISTER_160 + '"\n', encoding="utf-8")
    cats = {h["category"] for h in vo.check_cells(p)}
    assert {"em-dash", "fake-precision-count", "slash-joined-run",
            "formal-classname-run"} <= cats


def test_common_acronyms_not_flagged(tmp_path):
    # CI/CD, TCP/IP are 2-segment -> below the 3-segment slash-run threshold.
    p = tmp_path / "ok.csv"
    p.write_text("note\nset up CI/CD and TCP/IP, all good\n", encoding="utf-8")
    cats = {h["category"] for h in vo.check_cells(p)}
    assert "slash-joined-run" not in cats


def test_xlsx_positive_with_exclusions(tmp_path):
    openpyxl = pytest.importorskip("openpyxl")
    p = tmp_path / "hours.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = REGISTER_160                       # human cell -> hits
    ws["A2"] = "=SUM(B1:B3)"                       # formula -> skipped
    ws["A3"] = "[auto] hash1, hash2 — note"  # machine note -> skipped
    ws["A4"] = "clean human note, nothing odd"    # clean -> no hit
    wb.save(p)
    cells = {h["cell"] for h in vo.check_cells(p)}
    assert any(c.endswith("A1") for c in cells)
    assert not any(c.endswith(("A2", "A3", "A4")) for c in cells)


def test_xlsx_json_subprocess(tmp_path):
    pytest.importorskip("openpyxl")
    import openpyxl
    p = tmp_path / "h.xlsx"
    wb = openpyxl.Workbook()
    wb.active["A1"] = REGISTER_160
    wb.save(p)
    r = subprocess.run(
        [sys.executable, str(TOOLS / "validate-output.py"), str(p), "--format", "json"],
        capture_output=True, text=True,
    )
    payload = json.loads(r.stdout)
    assert payload["total"] >= 3
