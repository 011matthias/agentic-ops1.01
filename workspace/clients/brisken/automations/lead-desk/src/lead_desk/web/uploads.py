"""Campaign contact upload: CSV/xlsx -> global contact upsert + enrollment.

An existing contact (same email / natural key) is ADOPTED, never duplicated:
the upsert refreshes provided fields, the enrollment adds campaign
membership, and the full event history rides along because contact_id is
stable. Enrollment before approval is inert - nothing can send - so uploads
commit directly with no staging table; degrees stay editable and rows
removable until the approval freezes the list.
"""
from __future__ import annotations

import csv
import io

from ..identity import contact_id_for, natural_key
from .store import CONTACT_COLUMNS, ContactStore
from . import cadence
from .service import now_iso

# Header aliases, all matched lower/stripped (spaces and dashes -> underscore).
HEADER_ALIASES = {
    "e_mail": "email", "mail": "email", "email_address": "email",
    "firstname": "first_name", "first": "first_name", "vorname": "first_name",
    "lastname": "last_name", "last": "last_name", "surname": "last_name",
    "nachname": "last_name", "name": "full_name", "full_name": "full_name",
    "organisation": "company", "organization": "company", "account": "company",
    "firma": "company", "company_name": "company",
    "title": "job_title", "position": "job_title", "role": "job_title",
    "jobtitle": "job_title",
    "linkedin": "linkedin_url", "linkedin_profile": "linkedin_url",
    "li_url": "linkedin_url",
    "telephone": "phone", "tel": "phone", "mobile": "phone",
    "land": "country",
    "warmness": "degree", "warmth": "degree",
}

# Columns the upload may set directly (everything else is ignored + reported).
UPLOADABLE = tuple(c for c in CONTACT_COLUMNS if c not in (
    "contact_id", "natural_key", "campaign",
    "suppressed", "suppress_reason", "suppressed_at", "suppressed_by",
))


def _norm_header(h: str) -> str:
    key = (h or "").strip().lower().replace(" ", "_").replace("-", "_")
    return HEADER_ALIASES.get(key, key)


def _rows_from_csv(data: bytes) -> tuple[list[str], list[dict]]:
    text = data.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    headers = [_norm_header(h) for h in (reader.fieldnames or [])]
    rows = []
    for raw in reader:
        rows.append({_norm_header(k): (v or "").strip() for k, v in raw.items() if k})
    return headers, rows


def _rows_from_xlsx(data: bytes) -> tuple[list[str], list[dict]]:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    try:
        raw_header = list(next(it))
    except StopIteration:
        return [], []
    headers = [_norm_header(str(h) if h is not None else "") for h in raw_header]
    rows = []
    for r in it:
        if all(c is None for c in r):
            continue
        row = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            v = r[i] if i < len(r) else None
            row[h] = str(v).strip() if v is not None else ""
        rows.append(row)
    return headers, rows


def import_upload(store: ContactStore, campaign_id: str, filename: str,
                  data: bytes, user: str | None) -> dict:
    """Parse, upsert, enroll, classify. Returns the inline report dict."""
    if store.get_campaign(campaign_id) is None:
        return {"ok": False, "error": "unknown campaign"}
    name = (filename or "").lower()
    if name.endswith((".xlsx", ".xlsm")):
        headers, rows = _rows_from_xlsx(data)
    else:
        headers, rows = _rows_from_csv(data)
    if not rows:
        return {"ok": False, "error": "no data rows found"}

    known = set(UPLOADABLE) | {"full_name", "degree"}
    unmapped = sorted(h for h in headers if h and h not in known)
    now = now_iso()
    new_contacts = adopted = enrolled = already = degree_set = 0
    suppressed_hit = 0
    skipped: list[str] = []

    for ordinal, row in enumerate(rows):
        email = (row.get("email") or "").strip() or None
        first = (row.get("first_name") or "").strip() or None
        last = (row.get("last_name") or "").strip() or None
        if not first and not last and row.get("full_name"):
            parts = row["full_name"].split(None, 1)
            first = parts[0]
            last = parts[1] if len(parts) > 1 else None
        company = (row.get("company") or "").strip() or None
        if not email and not (first or last or company):
            skipped.append(f"row {ordinal + 2}: no email and no name/company")
            continue
        nk = natural_key(email, first, last, company)
        cid = contact_id_for(nk)
        existing = store.get_contact(cid)

        data_row: dict = {"contact_id": cid, "natural_key": nk,
                          "campaign": campaign_id,
                          "first_name": first, "last_name": last,
                          "company": company, "email": email}
        for col in UPLOADABLE:
            if col in data_row:
                continue
            v = (row.get(col) or "").strip()
            if v:
                data_row[col] = v
        if existing is not None:
            # Adopt: refresh only non-empty incoming fields; keep campaign tag.
            data_row.pop("campaign", None)
            data_row = {k: v for k, v in data_row.items()
                        if v not in (None, "") or k in ("contact_id", "natural_key")}
            data_row["campaign"] = existing["campaign"]
            adopted += 1
        else:
            new_contacts += 1
        store.upsert_contact(data_row, now)
        if (existing is not None and existing["suppressed"]) :
            suppressed_hit += 1

        if store.enroll(cid, campaign_id, user, now):
            enrolled += 1
        else:
            already += 1
        # An explicit degree column in the file is a manual assignment.
        want_degree = (row.get("degree") or "").strip().lower()
        if want_degree:
            enr = store.find_enrollment(cid, campaign_id)
            if enr is not None and not enr["approved_at"]:
                store.set_degree(enr["enrollment_id"], want_degree, "manual",
                                 f"upload column ({filename})")
                degree_set += 1

    classify = cadence.classify_enrollments(store, campaign_id, user, now)
    return {
        "ok": True, "rows": len(rows),
        "new_contacts": new_contacts, "adopted_existing": adopted,
        "enrolled": enrolled, "already_enrolled": already,
        "suppressed_enrolled": suppressed_hit,
        "manual_degrees_from_file": degree_set,
        "classified": classify["changed"],
        "unmapped_columns": unmapped, "skipped": skipped,
    }
