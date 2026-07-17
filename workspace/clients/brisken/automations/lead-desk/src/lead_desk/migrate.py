"""One-time, idempotent migration of the Rome master sheet into the Lead Desk.

Reads the live ``rome2026-post-event-master-contacts.xlsx`` ("Master contacts",
290 rows / 32 cols), maps it into ``contacts``, unifies the three "do not
contact" encodings into one ``suppressed`` flag, and parses the free-text
``outreach_log`` into structured ``outreach_events``. Optionally folds the
E1/E2/E3 send-log CSVs as precise ``sent``/``bounce`` events when present.

Re-runnable: contacts UPSERT by natural_key, events INSERT OR IGNORE by hash,
so a second run inserts nothing new.

    lead-desk-migrate --xlsx <path> --data ./lead-desk-data
"""
from __future__ import annotations

import argparse
import csv
import os
import re
from collections import Counter, defaultdict
from pathlib import Path

from .identity import contact_id_for, natural_key
from .web.service import now_iso
from .web.store import ContactStore

DEFAULT_XLSX = (
    "C:/Users/neuma_p1qrsic/Repo/agentic-ops1/workspace/clients/brisken/context/"
    "lead-generation/Rome-Event/event-admin/rome2026-post-event-master-contacts.xlsx"
)
DEFAULT_EXTRAS = (
    "C:/Users/neuma_p1qrsic/Repo/agentic-ops1/workspace/clients/brisken/context/"
    "lead-generation/Rome-Event"
)

_DATE_RE = re.compile(r"(\d{4}-\d{2}-\d{2})")

# Deterministic fallback timestamp (Rome event week) for events whose own date
# does not parse. Never use a wall-clock "now" for an event ts: it would change
# the event hash every run and break idempotency.
EVENT_WEEK_TS = "2026-06-24T00:00:00+00:00"

# The booth follow-up wave date, used when a post_event_outreach cell names no date.
POST_EVENT_DATE = "2026-07-08"

# Tier -> suppression reason for the exclusion tiers.
TIER_SUPPRESS = {
    "STOP": "stop", "DUPLICATE": "duplicate", "TEST": "test",
    "ORGANISER": "organiser", "OWN_TEAM": "own_team",
    "UNREACHABLE": "unreachable", "ANON": "anon",
}

# Fields the app owns once a contact exists: a scheduled sheet re-sync
# (preserve_app_fields=True) must not overwrite the operator's own pipeline
# work. Identity / classification / provenance / suppression stay
# sheet-authoritative; these do not.
APP_OWNED_ON_RESYNC = ("next_step", "next_step_due")

# --- Iteration 2 classifiers: real status the raw log misses -----------------

# Dirk personally reached the contact (relationship touch, not a campaign send).
# From dirk_notes markers, OR an if_we_know_them note that names Dirk with an
# engagement verb ("Met at TAC Brussels 2024 (Dirk personally engaged)").
_DIRK_NOTE_RE = re.compile(
    r"personal.*(?:outreach|dn)|individual outreach|dn[\s:]*personal|personally engaged",
    re.I,
)
_DIRK_IWK_RE = re.compile(r"dirk", re.I)
_DIRK_IWK_VERB_RE = re.compile(
    r"know|knew|met|meet|spoke|speak|talk|contact|reach|engag|relationship|introduc|connect",
    re.I,
)

# Deliberate, revisitable holds (an explicit next_step hold) -> off the active
# board but distinct from consent-suppressed. NOTE: tier='GA' is NOT a hold
# (owner directive 2026-07-17): the GA general-audience cohort stays in the
# active pipeline to receive its own GA outreach wave; once T3 + GA are out,
# per-tier non-responder follow-up starts.
_HELD_NEXT_RE = re.compile(r"on hold|do not send|excluded|covered by", re.I)
# Transient holds stay ACTIVE (surfaced via the next_step / dangling bucket).
_TRANSIENT_NEXT_RE = re.compile(
    r"ooo until|ooo auto|awaiting.*decision|scheduling in progress", re.I
)


def is_held(row: dict) -> bool:
    """A deliberate, revisitable hold: an explicit next_step hold. Transient
    holds (OOO, awaiting-decision, scheduling) stay active. GA is an active
    outreach tier, not a hold."""
    ns = str(row.get("next_step") or "").strip()
    if _HELD_NEXT_RE.search(ns) and not _TRANSIENT_NEXT_RE.search(ns):
        return True
    return False


def dirk_touch(row: dict) -> tuple[str, str] | None:
    """If Dirk personally reached this contact, return (channel, detail).
    Channel is inferred from the note (linkedin / meeting / email)."""
    dn = _s(row.get("dirk_notes"))
    iwk = _s(row.get("if_we_know_them"))
    note = None
    if dn and _DIRK_NOTE_RE.search(dn):
        note = dn
    elif iwk and _DIRK_IWK_RE.search(iwk) and _DIRK_IWK_VERB_RE.search(iwk):
        note = iwk
    if not note:
        return None
    low = note.lower()
    if "linkedin" in low:
        channel = "linkedin"
    elif any(w in low for w in ("met ", "meet", "meeting", "call", "spoke", "conversation")):
        channel = "meeting"
    else:
        channel = "email"
    return channel, note


def _s(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _yn(v) -> int:
    return 1 if str(v).strip().lower() in ("yes", "true", "1", "y") else 0


def _date(v) -> str | None:
    if v is None:
        return None
    m = _DATE_RE.search(str(v))
    return m.group(1) if m else None


def _ts(datestr: str | None) -> str | None:
    return f"{datestr}T00:00:00+00:00" if datestr else None


def suppression(row: dict) -> tuple[int, str | None]:
    """Collapse Tier=STOP, stop=X, and the two *_status 'do not contact' text
    values (plus the exclusion tiers) into one suppressed flag + reason."""
    stop = str(row.get("stop") or "").strip().upper() == "X"
    es = (row.get("email outreach_status") or "").strip().lower()
    ls = (row.get("linkedin_status") or "").strip().lower()
    tier = (row.get("Tier") or "").strip().upper()
    if "no consent" in es or "no consent" in ls:
        return 1, "no_consent"
    if stop:
        return 1, "stop"
    if "do not contact" in es or "do not contact" in ls:
        return 1, "do_not_contact"
    if tier in TIER_SUPPRESS:
        return 1, TIER_SUPPRESS[tier]
    # Held ranks below consent + the exclusion tiers (those are stronger and
    # permanent); a next_step hold is a revisitable off-board reason.
    if is_held(row):
        return 1, "held"
    return 0, None


def classify_log_line(line: str) -> tuple[str, str, str]:
    low = line.lower()
    ch = "linkedin" if "linkedin" in low else ("meeting" if ("meeting" in low or "call" in low) else "email")
    if "bounce" in low:
        return ("email", "inbound", "bounce")
    if any(w in low for w in ("response", "reply", "replied", "responded")):
        return (ch, "inbound", "reply")
    if "invite" in low:
        return (ch, "outbound", "invite")
    if any(w in low for w in ("sent", "e1", "e2", "e3", "follow", "nudge", "connect")):
        return (ch, "outbound", "sent")
    return (ch, "outbound", "note")


_EWAVE_RE = re.compile(r"\bE[123]\b", re.IGNORECASE)


def is_during_event(text: str) -> bool:
    """True for a during-event E-wave reference (E1/E2/E3). Graph is the
    authoritative source for these (see ground.py), so the sheet must not
    duplicate them into the timeline."""
    return bool(_EWAVE_RE.search(text or ""))


def import_workbook(store: ContactStore, xlsx: Path, campaign: str, report: dict,
                    preserve_app_fields: bool = False) -> dict:
    from openpyxl import load_workbook

    wb = load_workbook(str(xlsx), read_only=True, data_only=True)
    ws = wb["Master contacts"]
    it = ws.iter_rows(values_only=True)
    header = list(next(it))
    idx = {h: i for i, h in enumerate(header)}

    def col(r, name):
        i = idx.get(name)
        return r[i] if i is not None and i < len(r) else None

    now = now_iso()
    email_index: dict[str, str] = {}
    name_groups: dict[str, list[str]] = defaultdict(list)
    events_added = 0
    contacts = 0
    sheet_diffs: list[dict] = []   # app-owned fields where the sheet now differs

    for ordinal, r in enumerate(it):
        if all(c is None for c in r):
            continue
        first = _s(col(r, "first_name"))
        last = _s(col(r, "last_name"))
        company = _s(col(r, "company"))
        email = _s(col(r, "email"))
        email_l = email.lower() if email else None
        nk = natural_key(email, first, last, company, ordinal)
        cid = contact_id_for(nk)

        # A merged-away duplicate stays a tombstone: never resurrect it as a
        # fresh active contact on the next sync.
        _prior = store.get_contact_by_key(nk)
        if _prior is not None and _prior["merged_into"]:
            continue

        rowdict = {h: col(r, h) for h in header}
        supp, reason = suppression(rowdict)

        next_step = _s(col(r, "next_step"))
        data = {
            "contact_id": cid, "natural_key": nk, "campaign": campaign,
            "first_name": first, "last_name": last, "company": company,
            "job_title": _s(col(r, "job_title")), "email": email,
            "alt_email": _s(col(r, "alt_email")), "phone": _s(col(r, "phone")),
            "country": _s(col(r, "country")), "linkedin_url": _s(col(r, "linkedin_url")),
            "tier": _s(col(r, "Tier")), "tier_reason": _s(col(r, "Tier_reason")),
            "lead_type": _s(col(r, "lead_type")),
            # Sheet's human status, stored for DISPLAY only (owner 2026-07-15).
            # Sheet-authoritative on re-sync; deliberately NOT in
            # APP_OWNED_ON_RESYNC and NOT mapped into stage.
            "outreach_status": _s(col(r, "email outreach_status")),
            "suppressed": supp, "suppress_reason": reason,
            "suppressed_at": now if supp else None,
            "suppressed_by": "import" if supp else None,
            "crm_owner": _s(col(r, "crm_owner")),
            "next_step": next_step, "next_step_due": _date(next_step),
            "source": _s(col(r, "source")), "in_our_booth": _yn(col(r, "in_our_booth")),
            "scanned_at_booth": _yn(col(r, "scanned_at_booth")),
            "if_we_know_them": _s(col(r, "if_we_know_them")),
            "brisken_customer": _s(col(r, "brisken_customer")),
            "attendee_type": _s(col(r, "attendee_type")),
            "sponsor_opt_in": _yn(col(r, "sponsor_opt_in")), "no_show": _yn(col(r, "no_show")),
            "fob_encoded": _yn(col(r, "fob_encoded")),
            "booth_registered_at": _s(col(r, "booth_registered_at")),
            "crm_last_activity": _s(col(r, "crm_last_activity")),
            "dirk_notes": _s(col(r, "dirk_notes")),
        }
        # Sheet-follows-app re-sync: once a contact exists, keep the app's own
        # pipeline fields (next_step) rather than letting the sheet reset them.
        # But if the sheet now carries a DIFFERENT non-empty value for one of
        # those fields, record it so the operator sees "sheet differs" rather
        # than the edit vanishing silently.
        existing_row = store.get_contact_by_key(nk) if preserve_app_fields else None
        if existing_row is not None:
            for fld in APP_OWNED_ON_RESYNC:
                sheet_val = str(data.get(fld) or "").strip()
                board_val = str(existing_row[fld] or "").strip()
                if sheet_val and sheet_val != board_val:
                    sheet_diffs.append({"contact_id": cid, "field": fld,
                                        "sheet": sheet_val, "board": board_val})
            data = {k: v for k, v in data.items() if k not in APP_OWNED_ON_RESYNC}
        store.upsert_contact(data, now)
        contacts += 1
        if email_l:
            email_index[email_l] = cid
        alt = _s(col(r, "alt_email"))
        if alt:
            email_index[alt.lower()] = cid
        if first and last:
            name_groups[f"{first.lower()} {last.lower()}"].append(nk)

        # Parse outreach_log into events.
        last_out = _date(col(r, "last_outreach"))
        last_rep = _date(col(r, "last_reply"))
        log = col(r, "outreach_log")
        has_out = has_in = False
        if log:
            for raw in str(log).splitlines():
                line = raw.strip()
                if not line:
                    continue
                d = _date(line) or last_out or _date(col(r, "booth_registered_at"))
                ch, direction, typ = classify_log_line(line)
                if direction == "outbound":
                    has_out = True
                if direction == "inbound":
                    has_in = True
                # During-event (E1/E2/E3) is grounded from the mailboxes
                # (ground.py); do not duplicate it from the sheet. has_out/has_in
                # above still register the activity so the summary-date gap-fill
                # below stays quiet for these contacts.
                if is_during_event(line):
                    continue
                if store.add_event(
                    contact_id=cid, ts=_ts(d) or EVENT_WEEK_TS, channel=ch, direction=direction,
                    type=typ, detail=line, source="import", campaign=campaign, now=now,
                ):
                    events_added += 1
        # Fill stage-critical gaps from the summary dates.
        if last_out and not has_out:
            if store.add_event(contact_id=cid, ts=_ts(last_out), channel="email",
                               direction="outbound", type="sent",
                               detail="last_outreach", source="import",
                               campaign=campaign, now=now):
                events_added += 1
        if last_rep and not has_in:
            if store.add_event(contact_id=cid, ts=_ts(last_rep), channel="email",
                               direction="inbound", type="reply",
                               detail="last_reply", source="import",
                               campaign=campaign, now=now):
                events_added += 1

        # Dirk personal touch -> one outbound 'touch' event (reaches >= 'sent').
        # Stable ext_key so a re-run is a no-op even if the note text changes.
        touch = dirk_touch(rowdict)
        if touch:
            ch, detail = touch
            if store.add_event(contact_id=cid, ts=_ts(last_out) or EVENT_WEEK_TS,
                               channel=ch, direction="outbound", type="touch",
                               detail=detail, source="import",
                               ext_key=f"dirk-touch-{cid}", campaign=campaign, now=now):
                events_added += 1

        # Post-event follow-up phase: kept DISTINCT from during-event (E1/E2/E3),
        # which is grounded from the mailbox (see ground.py). The sheet's
        # post_event_outreach column is Dirk's tracking of the booth follow-up wave.
        pe = _s(col(r, "post_event_outreach"))
        if pe:
            pe_date = _date(pe) or POST_EVENT_DATE
            if store.add_event(contact_id=cid, ts=_ts(pe_date) or EVENT_WEEK_TS,
                               channel="email", direction="outbound", type="sent",
                               subject="Post-event follow-up", detail=f"Post-event: {pe}",
                               source="sheet-postevent", ext_key=f"pe-{cid}",
                               campaign=campaign, now=now):
                events_added += 1

    report["contacts"] = contacts
    report["events_from_sheet"] = events_added
    report["fuzzy_dups"] = {k: v for k, v in name_groups.items() if len(set(v)) > 1}
    report["sheet_diffs"] = sheet_diffs
    report["sheet_diff_count"] = len(sheet_diffs)
    return email_index


def fold_send_logs(store: ContactStore, email_index: dict, campaign: str,
                   extras: Path, report: dict) -> None:
    now = now_iso()
    added = 0
    for wave in ("E1", "E2", "E3"):
        path = extras / "email-campaign" / f"rome2026-send-log-{wave}.csv"
        if not path.exists():
            continue
        try:
            with path.open(newline="", encoding="utf-8-sig") as fh:
                for row in csv.DictReader(fh):
                    addr = (row.get("email") or "").strip().lower()
                    cid = email_index.get(addr)
                    if not cid:
                        continue
                    status = (row.get("status") or "").strip()
                    ts = _ts(_date(row.get("utc"))) or EVENT_WEEK_TS
                    is_bounce = "bounce" in status.lower()
                    if store.add_event(
                        contact_id=cid, ts=ts, channel="email",
                        direction="inbound" if is_bounce else "outbound",
                        type="bounce" if is_bounce else "sent",
                        detail=f"{wave} send-log: {status}", source="import",
                        # stable idempotency key so a re-run never re-inserts
                        ext_key=f"sendlog-{wave}-{addr}-{status}",
                        campaign=campaign, now=now,
                    ):
                        added += 1
        except Exception as exc:  # noqa: BLE001 - a malformed optional file must not abort the import
            report.setdefault("warnings", []).append(f"send-log {wave}: {exc}")
    report["events_from_send_logs"] = added


def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser(prog="lead-desk-migrate")
    p.add_argument("--xlsx", default=DEFAULT_XLSX)
    p.add_argument("--extras", default=DEFAULT_EXTRAS, help="Rome-Event dir for optional send logs.")
    p.add_argument("--data", default=os.environ.get("LEAD_DESK_DATA", "lead-desk-data"))
    p.add_argument("--campaign", default="rome-2026")
    p.add_argument("--no-send-logs", action="store_true")
    args = p.parse_args(argv)

    xlsx = Path(args.xlsx)
    if not xlsx.exists():
        print(f"ERROR: xlsx not found: {xlsx}")
        return 1
    db = Path(args.data).resolve() / "lead-desk.sqlite"
    report: dict = {}
    with ContactStore(db) as store:
        email_index = import_workbook(store, xlsx, args.campaign, report)
        if not args.no_send_logs:
            fold_send_logs(store, email_index, args.campaign, Path(args.extras), report)
        # Stage distribution after import.
        rows = store.board_rows(args.campaign)
        stages = Counter(r["stage"] for r in rows)
        supp = sum(1 for r in rows if r["suppressed"])
        supp_reasons = Counter(r["suppress_reason"] for r in rows if r["suppressed"])
        reached_dirk = sum(
            1 for r in rows if not r["suppressed"] and r["stage"] == "sent"
            and r["has_touch"] and not r["has_campaign_send"]
        )
        total_events = store.count_events()

    print(f"DB: {db}")
    print(f"contacts upserted: {report.get('contacts')}")
    print(f"suppressed: {supp}  ({dict(supp_reasons)})")
    print(f"events (sheet): {report.get('events_from_sheet')} | "
          f"(send-logs): {report.get('events_from_send_logs', 0)} | total in db: {total_events}")
    print(f"stage distribution: {dict(stages)}")
    print(f"reached (Dirk personal touch, no campaign send): {reached_dirk}")
    dups = report.get("fuzzy_dups") or {}
    if dups:
        print(f"POSSIBLE DUPLICATES (same name, different key), {len(dups)}:")
        for name, keys in dups.items():
            print(f"  {name}: {sorted(set(keys))}")
    for w in report.get("warnings", []):
        print(f"WARNING: {w}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
