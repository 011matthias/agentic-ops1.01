"""Appending statements to a living month (PR 2b-2b-2).

`POST /api/expense-batches/{id}/statement` is repeatable now: one card at a
time, a mid-month partial then the full cycle, the same file twice by
accident. PR 2b-2b-1 built the fold and said plainly that it was INERT on the
attach path, because a second upload was still refused and `existing` was
always empty. This round lifts that refusal, and these tests are what make
the fold load-bearing rather than decorative: every one of them fails if the
fold is unwired, and the file-scoping ones fail if the writeback is not
scoped per statement.

Three things are under test, and they are three because the round found two
hazards on top of the surface it set out to build:

1. **The append itself.** A second upload lands, a re-supplied charge does
   not double, and each upload is recorded in `statements[]`.
2. **A row belongs to a FILE, and a charge can be in several.** A month
   holding two workbooks that both have a row 14 will write Criss's resolved
   account into the wrong charge unless the writeback knows which file it is
   annotating; and a charge printed by both a partial and the closing cycle
   occupies a row in each, so the anchor is recorded per upload rather than
   on the charge.
3. **A contradiction is surfaced, never deduped.** Two uploads that disagree
   (a different account id for one card, a flipped sign) genuinely produce
   two rows. That is the deliberate call from 2a; what this round adds is
   saying so, because on screen it just looks like the month doubled.
"""
from __future__ import annotations

from decimal import Decimal
from pathlib import Path

import pytest

pytest.importorskip("fastapi")
pytest.importorskip("httpx")

from fastapi.testclient import TestClient  # noqa: E402
from openpyxl import Workbook, load_workbook  # noqa: E402

from expense_recon.llm.client import (  # noqa: E402
    ExtractedReceipt,
    FxJudgmentResult,
    MockLLMClient,
)
from expense_recon.web.app import create_app  # noqa: E402
from expense_recon.web.store import RunStore  # noqa: E402

EXAMPLES = Path(__file__).resolve().parent.parent / "examples"
JPG = b"\xff\xd8\xff\xe0fake-jpeg-bytes"


@pytest.fixture
def client(tmp_path, monkeypatch):
    monkeypatch.setenv("EXPENSE_RECON_RECEIPT_FIRST", "1")
    monkeypatch.delenv("OPENAI_API_KEY", raising=False)
    app = create_app(tmp_path)
    with TestClient(app) as c:
        c._data_root = tmp_path
        yield c


def _wire(monkeypatch, n=4):
    mock = MockLLMClient(
        extraction_responses=[
            ExtractedReceipt(
                date="2026-04-15", total="42.50", currency="USD",
                vendor="Staples", reference="", line_items=(),
                confidence=0.9, notes="", payment_hint=None,
            )
        ] * n,
        fx_responses=[
            FxJudgmentResult(
                is_match=True, same_purchase_confidence=0.9, implied_rate=1.0,
                converted_amount=Decimal("42.50"), reasoning="same purchase",
            )
        ] * 24,
    )
    monkeypatch.setattr(
        "expense_recon.cli._build_llm_client", lambda cfg: (mock, None)
    )
    return mock


def _batch(client, label="April 2026"):
    resp = client.post(
        "/api/expense-batches",
        files=[("files", ("a.jpg", JPG, "application/octet-stream"))],
        data={"legal_entity": "Corporate Services", "label": label},
    )
    assert resp.status_code == 200, resp.text
    assert client.get(f"/jobs/{resp.json()['job_id']}").json()["status"] == "done"
    return resp.json()["batch_id"]


# ── statement bodies ────────────────────────────────────────────────────
#
# Built here rather than taken from `examples/` because the point of most of
# these tests is that TWO files differ in a specific way (one row, one
# account id, one sign, one row ORDER), and a fixture pair cannot say which
# difference is the one under test.

_HEADER = "Date,Amount,Vendor\n"


def _csv(*rows: tuple[str, str, str]) -> bytes:
    body = "".join(f"{d},{a},{v}\n" for d, a, v in rows)
    return (_HEADER + body).encode()


def _xlsx(path: Path, *rows: tuple[str, str, str]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append(["Date", "Amount", "Vendor"])
    for row in rows:
        ws.append(list(row))
    wb.save(path)
    return path.read_bytes()


def _upload(client, batch_id, body: bytes, name="statement.csv", **data):
    form = {
        "account_id": "amex-9001",
        "account_legal_entities": '{"amex-9001": "Corporate Services"}',
        "account_card_currency": "USD",
        "map_transaction_date": "Date",
        "map_amount": "Amount",
        "map_vendor": "Vendor",
    }
    form.update({k: v for k, v in data.items() if v is not None})
    resp = client.post(
        f"/api/expense-batches/{batch_id}/statement",
        files={"statement": (name, body, "application/octet-stream")},
        data=form,
    )
    job = resp.json().get("job_id") if resp.status_code == 200 else None
    stage = None
    if job:
        done = client.get(f"/jobs/{job}").json()
        assert done["status"] == "done", done
        stage = done.get("stage")
    return resp, stage


def _run(client, batch_id):
    with RunStore(client._data_root / "recon-web.sqlite") as store:
        return store.get_run(batch_id)


def _charges(client, batch_id):
    return (_run(client, batch_id).snapshot or {}).get("transactions") or []


def _statements(client, batch_id):
    return (_run(client, batch_id).snapshot or {}).get("statements") or []


# ── 1. the append ───────────────────────────────────────────────────────


def test_the_full_cycle_after_a_partial_adds_only_what_is_new(
    client, monkeypatch
):
    """The shape a real month takes. Criss loads what the bank has printed so
    far, then the closing cycle, which re-prints everything plus the rest."""
    _wire(monkeypatch)
    batch_id = _batch(client)
    _upload(client, batch_id, _csv(
        ("2026-04-03", "42.50", "STAPLES"),
        ("2026-04-05", "12.00", "AWS"),
    ))
    _upload(client, batch_id, _csv(
        ("2026-04-03", "42.50", "STAPLES"),
        ("2026-04-05", "12.00", "AWS"),
        ("2026-04-22", "31.10", "UBER"),
    ), name="cycle.csv")

    charges = _charges(client, batch_id)
    assert len(charges) == 3
    assert [c["vendor_from_statement"] for c in charges] == [
        "STAPLES", "AWS", "UBER",
    ]
    assert len({c["transaction_id"] for c in charges}) == 3


def test_the_same_file_twice_changes_nothing_but_the_record(
    client, monkeypatch
):
    """The accident. Both uploads are recorded, because they both happened;
    the second contributes no charge, because it held none the month lacked."""
    _wire(monkeypatch)
    batch_id = _batch(client)
    body = _csv(("2026-04-03", "42.50", "STAPLES"))
    _upload(client, batch_id, body)
    _upload(client, batch_id, body)

    assert len(_charges(client, batch_id)) == 1
    entries = _statements(client, batch_id)
    assert [e["n_rows"] for e in entries] == [1, 1]
    assert [e["n_new"] for e in entries] == [1, 0]


def test_a_second_card_lands_beside_the_first(client, monkeypatch):
    """Per-card uploads are the point of the round. Two cards are two sets of
    charges in one month, and neither is replaced by the other."""
    _wire(monkeypatch)
    batch_id = _batch(client)
    _upload(client, batch_id, _csv(("2026-04-03", "42.50", "STAPLES")))
    _upload(
        client, batch_id, _csv(("2026-04-04", "80.00", "HERTZ")),
        name="chase.csv", account_id="chase-2838",
        account_legal_entities='{"chase-2838": "Corporate Services"}',
    )

    charges = _charges(client, batch_id)
    assert {c["account_id"] for c in charges} == {"amex-9001", "chase-2838"}
    assert len(charges) == 2


def test_each_upload_is_recorded_with_what_it_covered(client, monkeypatch):
    """`statements[]`: the month can say which files it has taken, over what
    period, and what each one contributed. `has_statement` only ever answered
    whether there was one at all."""
    _wire(monkeypatch)
    batch_id = _batch(client)
    _upload(client, batch_id, _csv(
        ("2026-04-03", "42.50", "STAPLES"),
        ("2026-04-22", "31.10", "UBER"),
    ), name="amex-april.csv", card_key="")

    entry = _statements(client, batch_id)[0]
    assert entry["file"] == "amex-april.csv"
    assert entry["upload_name"] == "amex-april.csv"
    assert entry["account_id"] == "amex-9001"
    assert entry["period_start"] == "2026-04-03"
    assert entry["period_end"] == "2026-04-22"
    assert (entry["n_rows"], entry["n_new"]) == (2, 2)
    assert entry["uploaded_at"]
    assert entry["writeback"] is False
    assert entry["advisory"] is None


def test_the_month_page_and_the_workbench_agree_on_the_uploads(
    client, monkeypatch
):
    """Both views carry `statements[]` and both read the same stored list.
    The month page is where the next upload is made; the workbench is where
    the reconciliation is read."""
    _wire(monkeypatch)
    batch_id = _batch(client)
    _upload(client, batch_id, _csv(("2026-04-03", "42.50", "STAPLES")))

    grid = client.get(f"/api/expense-batches/{batch_id}").json()
    workbench = client.get(f"/api/runs/{batch_id}").json()
    assert grid["statements"] == workbench["statements"]
    assert len(grid["statements"]) == 1


def test_the_month_re_matches_after_an_append(client, monkeypatch):
    """Allowed-but-inert is worse than refused (2b-2a). An appended charge
    that matches a receipt already in the pool has to move the match state,
    not sit there while the outcome describes the month as it was.

    Read from the STORED summary: the expense grid's summary is
    receipt-centric and carries no match counts."""
    _wire(monkeypatch)
    batch_id = _batch(client)
    _upload(client, batch_id, _csv(("2026-04-01", "9.99", "UNRELATED")))
    assert _run(client, batch_id).summary["n_matched"] == 0

    _upload(client, batch_id, _csv(
        ("2026-04-01", "9.99", "UNRELATED"),
        ("2026-04-15", "42.50", "STAPLES"),
    ), name="cycle.csv")

    assert _run(client, batch_id).summary["n_matched"] == 1
    assert _run(client, batch_id).summary["n_transactions"] == 2


# ── 2. a row's source travels with the row ──────────────────────────────


def test_two_uploads_of_the_same_filename_do_not_overwrite_each_other(
    client, monkeypatch
):
    """Criss's per-card exports carry the bank's own filename, so two cards
    genuinely arrive as one name. Overwriting is the dangerous outcome: the
    first upload's charges keep their `source_row` while the bytes beneath
    them become somebody else's rows."""
    _wire(monkeypatch)
    batch_id = _batch(client)
    _upload(client, batch_id, _csv(("2026-04-03", "42.50", "STAPLES")))
    _upload(
        client, batch_id, _csv(("2026-04-04", "80.00", "HERTZ")),
        account_id="chase-2838",
        account_legal_entities='{"chase-2838": "Corporate Services"}',
    )

    files = [e["file"] for e in _statements(client, batch_id)]
    assert files == ["statement.csv", "statement-2.csv"]
    assert all(e["upload_name"] == "statement.csv"
               for e in _statements(client, batch_id))
    work_dir = Path(_run(client, batch_id).work_dir)
    assert b"STAPLES" in (work_dir / "statement.csv").read_bytes()
    assert b"HERTZ" in (work_dir / "statement-2.csv").read_bytes()
    # And each file's anchors are its own, not the other's.
    anchors = (_run(client, batch_id).snapshot or {})["statement_anchors"]
    assert set(anchors) == {"statement.csv", "statement-2.csv"}
    assert not (set(anchors["statement.csv"]) & set(anchors["statement-2.csv"]))


def test_each_upload_records_the_rows_that_charge_occupies_in_it(
    client, monkeypatch
):
    """One charge sits in EVERY file that prints it, at a different row in
    each, so the row cannot live on the charge: a single field could only
    name one of the files. Here UBER is row 2 of the partial and row 4 of the
    cycle, and both have to be recorded."""
    _wire(monkeypatch)
    batch_id = _batch(client)
    _upload(client, batch_id, _csv(("2026-04-22", "31.10", "UBER")),
            name="partial.csv")
    _upload(client, batch_id, _csv(
        ("2026-04-03", "42.50", "STAPLES"),
        ("2026-04-05", "12.00", "AWS"),
        ("2026-04-22", "31.10", "UBER"),
    ), name="cycle.csv")

    uber = next(c for c in _charges(client, batch_id)
                if c["vendor_from_statement"] == "UBER")["transaction_id"]
    anchors = (_run(client, batch_id).snapshot or {})["statement_anchors"]
    assert anchors["partial.csv"][uber] == 2
    assert anchors["cycle.csv"][uber] == 4
    assert len(anchors["cycle.csv"]) == 3
    # The charge itself still points at where it was FIRST read, which is
    # what first-write-wins protects for decisions and stays untouched here.
    uber_row = next(c for c in _charges(client, batch_id)
                    if c["vendor_from_statement"] == "UBER")["source_row"]
    assert uber_row == 2


def test_the_closing_cycle_is_annotated_in_full_after_a_partial(
    client, monkeypatch, tmp_path
):
    """The canonical shape of the round, and the one the default download
    hits. Criss loads a mid-month partial, then the closing cycle; the cycle
    is the workbook she works from, so every one of its rows has to carry an
    account, not only the rows the cycle introduced.

    Anchoring on the charge's own first-read row would leave the repeats
    blank here, silently, and the blank cells would look like charges the
    tool could not resolve."""
    _wire(monkeypatch)
    batch_id = _batch(client)
    _upload(client, batch_id, _xlsx(
        tmp_path / "p.xlsx", ("2026-04-15", "42.50", "STAPLES"),
    ), name="partial.xlsx")
    _upload(client, batch_id, _xlsx(
        tmp_path / "c.xlsx",
        ("2026-04-15", "42.50", "STAPLES"),
        ("2026-04-20", "12.00", "AWS"),
        ("2026-04-22", "31.10", "UBER"),
    ), name="cycle.xlsx")

    cycle = _writeback_column_values(client, batch_id, "cycle.xlsx", tmp_path)
    assert set(cycle) == {2, 3, 4}, cycle
    partial = _writeback_column_values(client, batch_id, "partial.xlsx", tmp_path)
    assert set(partial) == {2}, partial


def _writeback_column_values(client, batch_id, name, tmp_path) -> dict[int, str]:
    """{row: written value} for one statement's writeback workbook.

    Reads the APPENDED column, never the workbook's own columns: Criss's
    vendor and amount cells are hers and are never touched, so asserting on
    them asserts a constant and would stay green with the scoping removed.
    What a wrong-file write changes is which rows carry a value and what that
    value says.
    """
    resp = client.get(
        f"/runs/{batch_id}/statement-categorized.xlsx", params={"file": name}
    )
    assert resp.status_code == 200, resp.text
    out = tmp_path / f"out-{name}"
    out.write_bytes(resp.content)
    ws = load_workbook(out).active
    col = next(
        c.column for c in ws[1] if c.value == "Zoho Account (tool)"
    )
    return {
        r: ws.cell(row=r, column=col).value
        for r in range(2, ws.max_row + 1)
        if ws.cell(row=r, column=col).value is not None
    }


def test_the_writeback_annotates_only_its_own_workbook(client, monkeypatch, tmp_path):
    """The hazard the fold surfaced, and the one that reaches Criss. Two
    workbooks, each with a charge at row 2 and one at row 3. Writing the
    month's charges into whichever workbook is current would put the Amex
    account beside the Chase charges, silently, because every row number is
    plausible in both files."""
    _wire(monkeypatch)
    batch_id = _batch(client)
    _upload(client, batch_id, _xlsx(
        tmp_path / "a.xlsx",
        # matches the one receipt in the pool -> writes an account
        ("2026-04-15", "42.50", "STAPLES"),
        ("2026-04-04", "11.00", "AMEXONLY"),
    ), name="amex.xlsx")
    _upload(client, batch_id, _xlsx(
        tmp_path / "b.xlsx",
        ("2026-04-05", "80.00", "HERTZ"),
        ("2026-04-06", "22.00", "CHASEONLY"),
        ("2026-04-07", "33.00", "CHASETHIRD"),
    ), name="chase.xlsx", account_id="chase-2838",
        account_legal_entities='{"chase-2838": "Corporate Services"}')

    amex = _writeback_column_values(client, batch_id, "amex.xlsx", tmp_path)
    chase = _writeback_column_values(client, batch_id, "chase.xlsx", tmp_path)

    # Chase has a third charge; Amex's workbook has no third data row. An
    # unscoped write puts that charge's account into a row of Criss's Amex
    # sheet that has no charge beside it at all.
    assert set(amex) == {2, 3}, amex
    assert set(chase) == {2, 3, 4}, chase
    # Amex row 2 is the one charge that matched a receipt, so it carries an
    # account; Chase row 2 matched nothing. An unscoped write lets the last
    # charge written to row 2 win, and both files end up saying the same
    # thing about a row that is not the same charge.
    assert amex[2] != chase[2], (amex, chase)


def test_a_workbook_that_held_no_rows_is_annotated_with_nothing(
    client, monkeypatch, tmp_path
):
    """"Recorded and empty" is not "not recorded". A workbook that parsed no
    charges has a real, empty anchor map, and reading that as "no map" drops
    the writeback back to placing every charge in the month by its own row
    number, which would put the Amex accounts into a workbook that has no
    charges at all."""
    _wire(monkeypatch)
    batch_id = _batch(client)
    _upload(client, batch_id, _xlsx(
        tmp_path / "real.xlsx",
        ("2026-04-15", "42.50", "STAPLES"),
        ("2026-04-16", "10.00", "SECOND"),
    ), name="real.xlsx")
    _upload(client, batch_id, _xlsx(tmp_path / "empty.xlsx"), name="empty.xlsx")

    anchors = (_run(client, batch_id).snapshot or {})["statement_anchors"]
    assert anchors["empty.xlsx"] == {}
    assert _writeback_column_values(client, batch_id, "empty.xlsx", tmp_path) == {}
    assert set(
        _writeback_column_values(client, batch_id, "real.xlsx", tmp_path)
    ) == {2, 3}


def test_a_workbook_the_month_never_loaded_is_not_addressable(
    client, monkeypatch, tmp_path
):
    """`?file=` is resolved against the run's own `statements[]`, never
    against the work dir. A name that is merely sanitized would still let a
    query string address a file the month never took."""
    _wire(monkeypatch)
    batch_id = _batch(client)
    _upload(client, batch_id, _xlsx(
        tmp_path / "real.xlsx", ("2026-04-03", "42.50", "STAPLES"),
    ), name="real.xlsx")
    work_dir = Path(_run(client, batch_id).work_dir)
    (work_dir / "elsewhere.xlsx").write_bytes(
        _xlsx(tmp_path / "elsewhere.xlsx", ("2026-04-09", "5.00", "OTHER"))
    )

    assert client.get(
        f"/runs/{batch_id}/statement-categorized.xlsx",
        params={"file": "elsewhere.xlsx"},
    ).status_code == 404
    assert client.get(
        f"/runs/{batch_id}/statement-categorized.xlsx",
        params={"file": "real.xlsx"},
    ).status_code == 200


# ── 3. a contradiction is surfaced, never deduped ───────────────────────


def test_one_card_typed_against_two_accounts_says_so(client, monkeypatch):
    """`account_id` is part of transaction identity, so the two uploads
    dedupe against nothing and the month doubles. Honest at the fold layer,
    since the rows really do claim to be different accounts, and not
    something the route may silently correct: the operator typed it. So it
    is said out loud, and nothing is dropped."""
    _wire(monkeypatch)
    batch_id = _batch(client)
    body = _csv(("2026-04-03", "42.50", "STAPLES"))
    _upload(client, batch_id, body, card_key="corp-1672")
    _, stage = _upload(
        client, batch_id, body, name="again.csv", card_key="corp-1672",
        account_id="amex-9001-v2",
        account_legal_entities='{"amex-9001-v2": "Corporate Services"}',
    )

    assert len(_charges(client, batch_id)) == 2
    advisory = _statements(client, batch_id)[1]["advisory"]
    assert advisory and "two account ids" in advisory
    assert stage and stage.startswith("warning:")
    assert "two account ids" in stage


def test_two_uploads_that_disagree_about_a_period_say_so(client, monkeypatch):
    """The sign-contradiction shape, and the reason 2a's deliberate call and
    this one are ONE call. A flipped sign gives the same printed row two
    content ids, so an upload over a period the month already covers can land
    100% new. Both readings stay; the month says that it now holds both."""
    _wire(monkeypatch)
    batch_id = _batch(client)
    _upload(client, batch_id, _csv(("2026-04-03", "42.50", "STAPLES")))
    _, stage = _upload(
        client, batch_id, _csv(("2026-04-03", "-42.50", "STAPLES")),
        name="resigned.csv",
    )

    assert len(_charges(client, batch_id)) == 2
    advisory = _statements(client, batch_id)[1]["advisory"]
    assert advisory and "already covers" in advisory
    assert stage and "already covers" in stage


def test_a_commit_never_drops_a_charge_the_month_gained_meanwhile(
    client, monkeypatch
):
    """The race the round opened, and the invariant that closes it.

    Both the attach path and every re-match read their transaction set
    minutes before they commit it, and a second upload can genuinely land in
    between now. Committing the older set would erase the appended charges
    with no error and no trace, which is the one outcome worse than refusing.

    Driven at the service layer with a real stored month: `rematch_month` is
    handed the charges as they were BEFORE an append, exactly as a slow
    concurrent job would hold them, and must refuse rather than write.
    """
    from expense_recon.web import service
    from expense_recon.web.serialize import snapshot_from_dict

    _wire(monkeypatch)
    batch_id = _batch(client)
    _upload(client, batch_id, _csv(("2026-04-03", "42.50", "STAPLES")))

    with RunStore(client._data_root / "recon-web.sqlite") as store:
        run = store.get_run(batch_id)
        stale, _, _, _ = snapshot_from_dict(run.snapshot)

    _upload(client, batch_id, _csv(
        ("2026-04-03", "42.50", "STAPLES"),
        ("2026-04-22", "31.10", "UBER"),
    ), name="cycle.csv")
    assert len(_charges(client, batch_id)) == 2

    with RunStore(client._data_root / "recon-web.sqlite") as store:
        with pytest.raises(service.RunInputError) as caught:
            service.rematch_month(
                store, run, transactions=stale,
                cfg=run.config or {}, entity="Corporate Services",
                now_iso="2026-04-30T00:00:00",
            )
    assert "while it reconciled" in str(caught.value)
    assert len(_charges(client, batch_id)) == 2


def test_a_clean_append_says_nothing(client, monkeypatch):
    """The advisory has to be quiet on the ordinary case, or it is noise the
    reviewer learns to skip. A second card over the same month is the normal
    shape of a per-card upload, not a contradiction."""
    _wire(monkeypatch)
    batch_id = _batch(client)
    _upload(client, batch_id, _csv(("2026-04-03", "42.50", "STAPLES")),
            card_key="corp-1672")
    _, stage = _upload(
        client, batch_id, _csv(("2026-04-04", "80.00", "HERTZ")),
        name="chase.csv", card_key="chase-2838", account_id="chase-2838",
        account_legal_entities='{"chase-2838": "Corporate Services"}',
    )

    assert [e["advisory"] for e in _statements(client, batch_id)] == [None, None]
    assert not (stage or "").startswith("warning:")
