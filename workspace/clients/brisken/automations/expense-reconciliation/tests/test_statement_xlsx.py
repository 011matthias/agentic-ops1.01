"""Tests for the Excel statement parser.

Mirrors `test_statement_csv.py` where behavior is shared (column-map
errors, accounting negatives, line-number errors, blank rows,
optional posting_date, integration with the matcher), plus
Excel-specific tests for native datetime cells, native float cells
(Decimal-from-float binary-noise guard), and string-typed cells
that fall back to the shared parsers.

Fixtures are generated in-memory per test via openpyxl. No binary
file is committed; openpyxl is a runtime dep so any test runner
already has it.
"""
from datetime import date, datetime
from decimal import Decimal

import pytest
from openpyxl import Workbook

from expense_recon.ingest._common import StatementParseError
from expense_recon.ingest.statement_xlsx import parse_statement_xlsx
from expense_recon.matching.deterministic import match_month
from expense_recon.matching.types import MatchType, Receipt


DEFAULT_MAP = {
    "transaction_date": "Date",
    "amount": "Amount",
    "vendor": "Description",
}


def _write_xlsx(path, rows, headers=("Date", "Description", "Amount")):
    """Write a list of row tuples to an .xlsx file under the given headers."""
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    wb.save(path)


@pytest.fixture
def amex_xlsx(tmp_path):
    """Mirror of the CSV Amex fixture. Uses native datetime + float
    cells (the common Excel shape) — exercises both the date-coercion
    and Decimal-from-float paths."""
    path = tmp_path / "amex.xlsx"
    rows = [
        (datetime(2026, 4, 1),  "COFFEE SHOP NYC",       5.75),
        (datetime(2026, 4, 3),  "DELANCEY TAVERN",       57.50),
        (datetime(2026, 4, 5),  "UBER * TRIP",           22.30),
        (datetime(2026, 4, 7),  "AMAZON.COM*RT3",        89.99),
        (datetime(2026, 4, 10), "   AMAZON RETURN   ", -15.00),
        (datetime(2026, 4, 12), "HOTEL PARIS FR",        112.30),
        (datetime(2026, 4, 15), "STAPLES NYC",           42.50),
    ]
    _write_xlsx(path, rows)
    return path


def test_parses_well_formed_xlsx(amex_xlsx):
    """Happy path: 7 data rows -> 7 Transactions; types correct."""
    txs = parse_statement_xlsx(
        amex_xlsx,
        column_map=DEFAULT_MAP,
        account_id="brisken-amex-usd",
        legal_entity_id="brisken-us",
        account_card_currency="USD",
    )
    assert len(txs) == 7
    first = txs[0]
    assert first.transaction_date == date(2026, 4, 1)
    assert isinstance(first.amount, Decimal)
    assert first.amount == Decimal("5.75")
    assert first.account_card_currency == "USD"
    assert first.transaction_currency == "USD"
    assert first.legal_entity_id == "brisken-us"
    assert first.posting_date is None
    assert first.vendor_from_statement == "COFFEE SHOP NYC"


def test_column_map_missing_required_key_raises(amex_xlsx):
    """Vendor is a required key; omitting it must fail early."""
    bad_map = {"transaction_date": "Date", "amount": "Amount"}  # no vendor
    with pytest.raises(StatementParseError) as exc:
        parse_statement_xlsx(
            amex_xlsx,
            column_map=bad_map,
            account_id="x",
            legal_entity_id="x",
            account_card_currency="USD",
        )
    assert "vendor" in str(exc.value)


def test_xlsx_missing_mapped_column_raises(amex_xlsx):
    """If the .xlsx has no column the map points at, fail with the
    column name visible to the caller."""
    bad_map = {**DEFAULT_MAP, "amount": "TotallyNotAColumn"}
    with pytest.raises(StatementParseError) as exc:
        parse_statement_xlsx(
            amex_xlsx,
            column_map=bad_map,
            account_id="x",
            legal_entity_id="x",
            account_card_currency="USD",
        )
    assert "TotallyNotAColumn" in str(exc.value)


def test_negative_amount_and_whitespace_preserved_correctly(amex_xlsx):
    """Refund row has native float -15.00; description has leading +
    trailing whitespace that should be stripped."""
    txs = parse_statement_xlsx(
        amex_xlsx,
        column_map=DEFAULT_MAP,
        account_id="brisken-amex-usd",
        legal_entity_id="brisken-us",
        account_card_currency="USD",
    )
    refund = next(t for t in txs if t.amount < 0)
    assert refund.amount == Decimal("-15.00")
    assert refund.vendor_from_statement == "AMAZON RETURN"


def test_native_float_amounts_have_no_binary_noise(tmp_path):
    """`Decimal(str(5.75)) == Decimal('5.75')`; `Decimal(5.75)` does
    not. The parser must use `str(value)` to avoid the IEEE-754
    binary noise. Verify with a few values prone to it."""
    path = tmp_path / "floats.xlsx"
    _write_xlsx(path, [
        (datetime(2026, 4, 1), "A", 0.1),
        (datetime(2026, 4, 2), "B", 0.2),
        (datetime(2026, 4, 3), "C", 1.05),
    ])
    txs = parse_statement_xlsx(
        path,
        column_map=DEFAULT_MAP,
        account_id="x",
        legal_entity_id="x",
        account_card_currency="USD",
    )
    assert txs[0].amount == Decimal("0.1")
    assert txs[1].amount == Decimal("0.2")
    assert txs[2].amount == Decimal("1.05")


def test_date_stored_as_string_cell(tmp_path):
    """If a cell is typed as text 'MM/DD/YYYY' rather than a real
    Excel date, fall back to the shared string parser. Common when
    a bank export writes dates as text."""
    path = tmp_path / "string_dates.xlsx"
    _write_xlsx(path, [
        ("04/01/2026", "COFFEE", 5.75),
    ])
    txs = parse_statement_xlsx(
        path,
        column_map=DEFAULT_MAP,
        account_id="x",
        legal_entity_id="x",
        account_card_currency="USD",
    )
    assert txs[0].transaction_date == date(2026, 4, 1)


def test_amount_stored_as_string_cell(tmp_path):
    """String-typed amount cells route through the shared parser
    (handles `$`, `,`, `(50.00)` accounting negatives)."""
    path = tmp_path / "string_amounts.xlsx"
    _write_xlsx(path, [
        (datetime(2026, 4, 1), "COFFEE", "$5.75"),
        (datetime(2026, 4, 2), "REFUND", "(50.00)"),
    ])
    txs = parse_statement_xlsx(
        path,
        column_map=DEFAULT_MAP,
        account_id="x",
        legal_entity_id="x",
        account_card_currency="USD",
    )
    assert txs[0].amount == Decimal("5.75")
    assert txs[1].amount == Decimal("-50.00")


def test_malformed_amount_raises_with_line_number(tmp_path):
    """Reconciliation-guarantee posture (v2 spec §25.5): never
    silently drop a row. Surface the offending row number — same
    1-indexed convention as CSV (header is row 1)."""
    path = tmp_path / "bad.xlsx"
    _write_xlsx(path, [
        (datetime(2026, 4, 1), "COFFEE", 5.75),
        (datetime(2026, 4, 2), "RESTAURANT", "not-a-number"),
    ])
    with pytest.raises(StatementParseError) as exc:
        parse_statement_xlsx(
            path,
            column_map=DEFAULT_MAP,
            account_id="x",
            legal_entity_id="x",
            account_card_currency="USD",
        )
    assert exc.value.line_number == 3  # header row 1; second data row is 3


def test_posting_date_optional_when_mapped(tmp_path):
    """Caller can map an optional posting_date column; blank cells
    yield posting_date=None on that row."""
    path = tmp_path / "with_posting.xlsx"
    _write_xlsx(
        path,
        [
            (datetime(2026, 4, 1), datetime(2026, 4, 3), "COFFEE",     5.75),
            (datetime(2026, 4, 2), None,                 "RESTAURANT", 20.00),
        ],
        headers=("Date", "Post Date", "Description", "Amount"),
    )
    txs = parse_statement_xlsx(
        path,
        column_map={**DEFAULT_MAP, "posting_date": "Post Date"},
        account_id="x",
        legal_entity_id="x",
        account_card_currency="USD",
    )
    assert txs[0].posting_date == date(2026, 4, 3)
    assert txs[1].posting_date is None


def test_blank_rows_skipped(tmp_path):
    """Blank rows (common at EOF in Excel exports) are ignored, not
    surfaced as errors."""
    path = tmp_path / "blanks.xlsx"
    _write_xlsx(path, [
        (datetime(2026, 4, 1), "COFFEE", 5.75),
        (None, None, None),
        (datetime(2026, 4, 2), "RESTAURANT", 20.00),
    ])
    txs = parse_statement_xlsx(
        path,
        column_map=DEFAULT_MAP,
        account_id="x",
        legal_entity_id="x",
        account_card_currency="USD",
    )
    assert len(txs) == 2


def test_per_row_currency_column(tmp_path):
    """Optional transaction_currency column overrides the
    account-card-currency default per row (v2 spec §20 layer 1)."""
    path = tmp_path / "fx.xlsx"
    _write_xlsx(
        path,
        [
            (datetime(2026, 4, 12), "HOTEL PARIS FR", 100.00, "EUR"),
            (datetime(2026, 4, 13), "COFFEE NYC",       5.75, "USD"),
        ],
        headers=("Date", "Description", "Amount", "TxCurrency"),
    )
    txs = parse_statement_xlsx(
        path,
        column_map={**DEFAULT_MAP, "transaction_currency": "TxCurrency"},
        account_id="x",
        legal_entity_id="x",
        account_card_currency="USD",
    )
    assert txs[0].transaction_currency == "EUR"
    assert txs[0].account_card_currency == "USD"  # account currency unchanged
    assert txs[1].transaction_currency == "USD"


def test_named_sheet_selection(tmp_path):
    """If the file has multiple sheets, the caller can target one
    by name. The default (active sheet) covers single-sheet exports."""
    path = tmp_path / "multi_sheet.xlsx"
    wb = Workbook()
    ws_active = wb.active
    ws_active.title = "Summary"
    ws_active.append(("Junk", "Stuff"))
    ws_active.append(("ignored", 1))
    ws_data = wb.create_sheet("Transactions")
    ws_data.append(("Date", "Description", "Amount"))
    ws_data.append((datetime(2026, 4, 1), "COFFEE", 5.75))
    wb.save(path)

    txs = parse_statement_xlsx(
        path,
        column_map=DEFAULT_MAP,
        account_id="x",
        legal_entity_id="x",
        account_card_currency="USD",
        sheet_name="Transactions",
    )
    assert len(txs) == 1
    assert txs[0].vendor_from_statement == "COFFEE"


def test_bool_amount_rejected(tmp_path):
    """A stray TRUE/FALSE cell in the amount column must NOT silently
    become Decimal(1) / Decimal(0). `bool` is an `int` subclass in
    Python; the parser rejects it explicitly."""
    path = tmp_path / "bool.xlsx"
    _write_xlsx(path, [
        (datetime(2026, 4, 1), "ODDITY", True),
    ])
    with pytest.raises(StatementParseError) as exc:
        parse_statement_xlsx(
            path,
            column_map=DEFAULT_MAP,
            account_id="x",
            legal_entity_id="x",
            account_card_currency="USD",
        )
    assert exc.value.line_number == 2


def test_integration_parser_to_matcher_happy_path(amex_xlsx):
    """End-to-end: .xlsx -> Transaction list -> match_month with
    synthetic receipts. Both seeded receipts produce EXACT matches;
    the other 5 transactions remain unmatched (reconciliation
    guarantee preserved, v2 spec §25.5)."""
    txs = parse_statement_xlsx(
        amex_xlsx,
        column_map=DEFAULT_MAP,
        account_id="brisken-amex-usd",
        legal_entity_id="brisken-us",
        account_card_currency="USD",
    )
    receipts = [
        Receipt(
            document_id="r-coffee",
            legal_entity_id="brisken-us",
            detected_date=date(2026, 4, 1),
            detected_total=Decimal("5.75"),
            detected_currency="USD",
            detected_vendor="COFFEE SHOP",
        ),
        Receipt(
            document_id="r-staples",
            legal_entity_id="brisken-us",
            detected_date=date(2026, 4, 15),
            detected_total=Decimal("42.50"),
            detected_currency="USD",
            detected_vendor="STAPLES",
        ),
    ]
    outcome = match_month(txs, receipts)
    assert len(outcome.matches) == 2
    assert all(m.match_type == MatchType.EXACT for m in outcome.matches)
    assert len(outcome.unmatched_transactions) == 5
