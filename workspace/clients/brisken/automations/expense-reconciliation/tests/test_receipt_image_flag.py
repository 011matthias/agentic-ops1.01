"""L4 — missing receipt image as a first-class flag (2026-07-15 walkthrough:
"ela esta faltando comprovante" is a real per-expense state Chris tracks).

The flag only surfaces when the run's receipt source carries image info at
all: the slice-1 receipts CSV never populates receipt_url/receipt_name, so
those runs render no badge and no MISSING column values (noise guard)."""
from __future__ import annotations

from datetime import date
from decimal import Decimal

from expense_recon.matching.deterministic import match_month
from expense_recon.matching.types import Receipt, Transaction
from expense_recon.output.reconciled_csv import (
    RECONCILED_COLUMNS,
    build_reconciled_rows,
)


def _tx(i: int, amount: str) -> Transaction:
    return Transaction(
        transaction_id=f"card-1:{i}",
        legal_entity_id="le",
        account_id="card-1",
        transaction_date=date(2026, 6, 10 + i),
        posting_date=None,
        amount=Decimal(amount),
        transaction_currency="USD",
        account_card_currency="USD",
        vendor_from_statement=f"VENDOR {i}",
    )


def _receipt(doc: str, amount: str, day: int, **kw) -> Receipt:
    return Receipt(
        document_id=doc,
        legal_entity_id="le",
        detected_date=date(2026, 6, day),
        detected_total=Decimal(amount),
        detected_currency="USD",
        detected_vendor=kw.pop("vendor", "VENDOR"),
        **kw,
    )


def test_has_receipt_image_property():
    assert _receipt("r1", "10", 10, receipt_url="https://x/img.jpg").has_receipt_image
    assert _receipt("r2", "10", 10, receipt_name="img.jpg").has_receipt_image
    assert not _receipt("r3", "10", 10).has_receipt_image


def test_reconciled_csv_marks_yes_and_missing():
    txs = [_tx(0, "10.00"), _tx(1, "20.00")]
    receipts = [
        _receipt("r0", "10.00", 10, vendor="VENDOR 0", receipt_name="img0.jpg"),
        _receipt("r1", "20.00", 11, vendor="VENDOR 1"),  # no image reference
    ]
    outcome = match_month(txs, receipts)
    rows = build_reconciled_rows(outcome, txs, receipts)
    col = RECONCILED_COLUMNS.index("Receipt Image")
    by_tx = {row[0 + 0]: row for row in rows}  # Account col is same; key on desc
    values = {row[3]: row[col] for row in rows}  # Description -> Receipt Image
    assert values["VENDOR 0"] == "Yes"
    assert values["VENDOR 1"] == "MISSING"


def test_reconciled_csv_suppresses_flag_without_image_info():
    txs = [_tx(0, "10.00")]
    receipts = [_receipt("r0", "10.00", 10, vendor="VENDOR 0")]  # slice-1 shape
    outcome = match_month(txs, receipts)
    rows = build_reconciled_rows(outcome, txs, receipts)
    col = RECONCILED_COLUMNS.index("Receipt Image")
    assert all(row[col] == "" for row in rows)


def test_report_note_carries_missing_image():
    from openpyxl import load_workbook

    from expense_recon.output.report_xlsx import write_report

    txs = [_tx(0, "10.00"), _tx(1, "20.00")]
    receipts = [
        _receipt("r0", "10.00", 10, vendor="VENDOR 0", receipt_url="https://x/i.jpg"),
        _receipt("r1", "20.00", 11, vendor="VENDOR 1"),
    ]
    outcome = match_month(txs, receipts)
    import tempfile
    from pathlib import Path

    with tempfile.TemporaryDirectory() as td:
        out = Path(td) / "report.xlsx"
        write_report(outcome, txs, receipts, out)
        wb = load_workbook(out)
        text = "\n".join(
            str(c.value)
            for ws in wb.worksheets
            for row in ws.iter_rows()
            for c in row
            if c.value is not None
        )
    assert "missing receipt image" in text
