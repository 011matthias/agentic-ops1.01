"""Report-writer unit tests (E1) + the --explain sheet (A8)."""
from __future__ import annotations

from datetime import date
from decimal import Decimal

from openpyxl import load_workbook

from expense_recon.matching.types import (
    Categorization,
    ClassificationSource,
    LineItem,
    Match,
    MatchOutcome,
    MatchType,
    Receipt,
    Transaction,
)
from expense_recon.output.report_xlsx import write_report


def _categorized(desc, amount, category) -> LineItem:
    return LineItem(
        description=desc,
        line_total=Decimal(amount),
        categorization=Categorization(
            category=category, zoho_account=None, confidence=0.95,
            source=ClassificationSource.LINE, reasoning="test",
        ),
    )


def _matched_tx() -> Transaction:
    return Transaction(
        transaction_id="t1", legal_entity_id="le1", account_id="amex-usd",
        transaction_date=date(2026, 4, 7), posting_date=None,
        amount=Decimal("180"), transaction_currency="USD",
        account_card_currency="USD", vendor_from_statement="AMAZON",
    )


def _unmatched_tx() -> Transaction:
    return Transaction(
        transaction_id="t2", legal_entity_id="le1", account_id="amex-usd",
        transaction_date=date(2026, 4, 9), posting_date=None,
        amount=Decimal("42"), transaction_currency="USD",
        account_card_currency="USD", vendor_from_statement="MYSTERY LLC",
    )


def _receipt() -> Receipt:
    return Receipt(
        document_id="r1", legal_entity_id="le1",
        detected_date=date(2026, 4, 7), detected_total=Decimal("180"),
        detected_currency="USD", detected_vendor="Amazon",
        line_items=(
            _categorized("Herman Miller chair", "150", "Equipment & Hardware"),
            _categorized("Coffee beans 2kg", "30", "Office Supplies & Consumables"),
        ),
    )


def _outcome() -> MatchOutcome:
    return MatchOutcome(
        matches=[Match("t1", "r1", MatchType.EXACT, 0.99, "exact", False)],
        unmatched_transactions=["t2"],
    )


def test_report_has_canonical_sheets(tmp_path):
    out = write_report(
        _outcome(), [_matched_tx(), _unmatched_tx()], [_receipt()],
        tmp_path / "report.xlsx",
        parse_errors=[("amex.csv", 5, "bad row")],
    )
    wb = load_workbook(out)
    assert "Summary" in wb.sheetnames
    assert "amex-usd" in wb.sheetnames          # per-card tab
    assert "Needs Review" in wb.sheetnames
    assert "Unmatched" in wb.sheetnames
    assert "Errors" in wb.sheetnames
    assert "Explain" not in wb.sheetnames        # not requested


def test_matched_lines_expand_to_one_row_each(tmp_path):
    out = write_report(_outcome(), [_matched_tx(), _unmatched_tx()], [_receipt()], tmp_path / "r.xlsx")
    ws = load_workbook(out)["amex-usd"]
    descriptions = [row[2] for row in ws.iter_rows(min_row=2, values_only=True) if row[2]]
    assert "Herman Miller chair" in descriptions
    assert "Coffee beans 2kg" in descriptions


def test_parse_errors_land_in_errors_sheet(tmp_path):
    out = write_report(
        _outcome(), [_matched_tx()], [_receipt()], tmp_path / "r.xlsx",
        parse_errors=[("amex.csv", 5, "unparseable amount")],
    )
    ws = load_workbook(out)["Errors"]
    cells = [c for row in ws.iter_rows(values_only=True) for c in row]
    assert "unparseable amount" in cells


def test_explain_sheet_present_only_when_requested(tmp_path):
    out = write_report(
        _outcome(), [_matched_tx(), _unmatched_tx()], [_receipt()],
        tmp_path / "r.xlsx", explain=True,
    )
    wb = load_workbook(out)
    assert "Explain" in wb.sheetnames
    ws = wb["Explain"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    by_tx = {r[2]: r[5] for r in rows}  # vendor -> Outcome
    assert by_tx["AMAZON"] == "MATCHED"
    assert by_tx["MYSTERY LLC"] == "UNMATCHED"


def _summary_labels(wb):
    ws = wb["Summary"]
    return {
        row[0]: row[1]
        for row in ws.iter_rows(values_only=True)
        if row and isinstance(row[0], str) and len(row) > 1
    }


def _card_spend(wb) -> float:
    """Sum the 'By card' Spend column."""
    ws = wb["Summary"]
    total = 0.0
    in_card = False
    for row in ws.iter_rows(values_only=True):
        if row[0] == "Card" and row[1] == "Spend":
            in_card = True
            continue
        if in_card:
            if row[0] is None:
                break
            if isinstance(row[1], (int, float)):
                total += row[1]
    return total


def test_summary_counts_transactions_not_pair_rows(tmp_path):
    """A9: a transaction with multiple FX candidate receipts must count
    as ONE in every Summary figure. Before the fix, judgment_required's
    per-candidate entries inflated the invariant, the review count, and
    Spend (a real $8.8K month rendered as $1.26M)."""
    tx_fx = Transaction(
        transaction_id="t-fx", legal_entity_id="le1", account_id="chase-2838",
        transaction_date=date(2026, 4, 15), posting_date=None,
        amount=Decimal("100"), transaction_currency="USD",
        account_card_currency="USD", vendor_from_statement="MEGA CENTER",
    )
    # Two candidate receipts for the SAME transaction -> two
    # judgment_required entries (the pre-bipartite reality).
    rec_a = Receipt(
        document_id="ra", legal_entity_id="le1", detected_date=date(2026, 4, 15),
        detected_total=Decimal("525"), detected_currency="BRL", detected_vendor="A",
    )
    rec_b = Receipt(
        document_id="rb", legal_entity_id="le1", detected_date=date(2026, 4, 15),
        detected_total=Decimal("530"), detected_currency="BRL", detected_vendor="B",
    )
    outcome = MatchOutcome(
        judgment_required=[
            Match("t-fx", "ra", MatchType.FX_JUDGMENT, 0.5, "fx a", True),
            Match("t-fx", "rb", MatchType.FX_JUDGMENT, 0.5, "fx b", True),
        ],
    )
    out = write_report(outcome, [tx_fx], [rec_a, rec_b], tmp_path / "r.xlsx")
    wb = load_workbook(out)
    labels = _summary_labels(wb)

    assert labels["Transactions"] == 1
    assert labels["Needs Review (FX / ambiguous)"] == 1  # not 2
    assert labels["Reconciliation invariant"] == "OK"    # not BROKEN
    # Spend = the single card charge, once — not the summed BRL receipts.
    assert _card_spend(wb) == 100.0


# ── Slice 10: charge categorization on the unmatched rows ────────────


def _charge_cat(source=ClassificationSource.LEARNED):
    return Categorization(
        category="Software & Subscriptions",
        zoho_account="Other Infra and IT Costs for Cloud Business",
        confidence=1.0, source=source,
        reasoning="from your Zoho Books posting history",
    )


def test_unmatched_sheet_carries_charge_category_columns(tmp_path):
    out = write_report(
        _outcome(), [_matched_tx(), _unmatched_tx()], [_receipt()],
        tmp_path / "r.xlsx",
        charge_categorizations={"t2": _charge_cat()},
    )
    ws = load_workbook(out)["Unmatched"]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[1]
    assert header[5:8] == ("Category", "Zoho A/C", "Source")
    tx_row = next(r for r in rows if r[2] == "MYSTERY LLC")
    assert tx_row[5] == "Software & Subscriptions"
    assert tx_row[6] == "Other Infra and IT Costs for Cloud Business"
    assert tx_row[7] == "LEARNED"


def test_unmatched_card_tab_row_carries_category_and_source(tmp_path):
    out = write_report(
        _outcome(), [_matched_tx(), _unmatched_tx()], [_receipt()],
        tmp_path / "r.xlsx",
        charge_categorizations={"t2": _charge_cat()},
    )
    ws = load_workbook(out)["amex-usd"]
    row = next(
        r for r in ws.iter_rows(min_row=2, values_only=True)
        if r[1] == "MYSTERY LLC"
    )
    # CARD_TAB_COLUMNS: Date, Vendor, Line item, Qty, Amount, Category,
    # Source, Zoho A/C, Note
    assert row[5] == "Software & Subscriptions"
    assert row[6] == "LEARNED"
    assert row[7] == "Other Infra and IT Costs for Cloud Business"
    assert "categorized from statement description" in row[8]


def test_without_side_map_unmatched_rows_unchanged(tmp_path):
    out = write_report(
        _outcome(), [_matched_tx(), _unmatched_tx()], [_receipt()],
        tmp_path / "r.xlsx",
    )
    ws = load_workbook(out)["amex-usd"]
    row = next(
        r for r in ws.iter_rows(min_row=2, values_only=True)
        if r[1] == "MYSTERY LLC"
    )
    assert row[5] is None or row[5] == ""   # no category invented
    assert row[6] == "REVIEW"


# ── §17 disposition column (CARD_TAB_COLUMNS index 9, appended last) ──


def test_card_tab_disposition_defaults_to_business(tmp_path):
    out = write_report(
        _outcome(), [_matched_tx(), _unmatched_tx()], [_receipt()],
        tmp_path / "r.xlsx",
    )
    ws = load_workbook(out)["amex-usd"]
    row = next(
        r for r in ws.iter_rows(min_row=2, values_only=True)
        if r[2] == "Herman Miller chair"
    )
    assert row[9] == "business"  # Disposition column, defaulted


def test_card_tab_disposition_reflects_map(tmp_path):
    out = write_report(
        _outcome(), [_matched_tx(), _unmatched_tx()], [_receipt()],
        tmp_path / "r.xlsx",
        dispositions={"t1": "reimbursable_personal"},
    )
    ws = load_workbook(out)["amex-usd"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Every line row of the matched tx t1 carries its disposition.
        if row[2] in ("Herman Miller chair", "Coffee beans 2kg"):
            assert row[9] == "reimbursable_personal"
