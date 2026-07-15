"""L1 + L6 + L5 — the statement_xlsx batch (2026-07-15 walkthrough).

L1: cell fill IS data (yellow row = already in Zoho, gray = subscription).
L6: a formula-derived mapped column gets a warning, never an abort.
L5: optional original_amount / original_currency / fx_rate columns carry
per-charge FX detail in tabular statements (xlsx AND csv symmetry).
"""
from __future__ import annotations

from datetime import date
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.styles import PatternFill

from expense_recon.ingest.statement_csv import parse_statement_csv_tolerant
from expense_recon.ingest.statement_xlsx import (
    _classify_rgb,
    parse_statement_xlsx,
    parse_statement_xlsx_tolerant,
)
from expense_recon.matching.deterministic import match_month
from expense_recon.matching.types import Receipt
from expense_recon.output.zoho_export import build_journal_rows
from expense_recon.web.serialize import snapshot_from_dict, snapshot_to_dict

COLUMN_MAP = {"transaction_date": "Date", "amount": "Amount", "vendor": "Description"}


def _fill(hex_rgb: str) -> PatternFill:
    return PatternFill(start_color=hex_rgb, end_color=hex_rgb, fill_type="solid")


def _write_colored_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(["Date", "Amount", "Description"])
    ws.append(["2026-06-01", "10.00", "PLAIN VENDOR"])          # no fill
    ws.append(["2026-06-02", "20.00", "POSTED VENDOR"])          # yellow row
    ws.append(["2026-06-03", "30.00", "SUBSCRIPTION VENDOR"])    # gray row
    for cell in ws[3]:
        cell.fill = _fill("FFFFEB9C")
    for cell in ws[4]:
        cell.fill = _fill("FFD9D9D9")
    wb.save(path)


@pytest.mark.parametrize(
    ("rgb", "expected"),
    [
        # yellows -> posted
        ((255, 255, 0), "posted"),      # FFFF00
        ((255, 255, 153), "posted"),    # FFFF99
        ((255, 230, 153), "posted"),    # FFE699
        ((255, 235, 156), "posted"),    # FFEB9C
        # grays -> subscription
        ((217, 217, 217), "subscription"),  # D9D9D9
        ((191, 191, 191), "subscription"),  # BFBFBF
        ((128, 128, 128), "subscription"),  # 808080
        # neither
        ((255, 255, 255), None),        # white
        ((0, 0, 0), None),              # black
        ((255, 192, 0), None),          # orange FFC000
        ((198, 239, 206), None),        # the report's own green
    ],
)
def test_classify_rgb_table(rgb, expected):
    assert _classify_rgb(*rgb) == expected


def test_fill_color_becomes_entry_status(tmp_path):
    path = tmp_path / "colored.xlsx"
    _write_colored_workbook(path)
    txs, issues = parse_statement_xlsx_tolerant(
        path, COLUMN_MAP, "card-1", "le", "USD"
    )
    assert [i for i in issues if i.severity == "error"] == []
    by_vendor = {t.vendor_from_statement: t.entry_status for t in txs}
    assert by_vendor == {
        "PLAIN VENDOR": None,
        "POSTED VENDOR": "posted",
        "SUBSCRIPTION VENDOR": "subscription",
    }


def test_entry_status_survives_snapshot_round_trip(tmp_path):
    path = tmp_path / "colored.xlsx"
    _write_colored_workbook(path)
    txs, _ = parse_statement_xlsx_tolerant(path, COLUMN_MAP, "card-1", "le", "USD")
    outcome = match_month(txs, [])
    snap = snapshot_to_dict(txs, [], outcome, [])
    txs2, _, _, _ = snapshot_from_dict(snap)
    assert [t.entry_status for t in txs2] == [t.entry_status for t in txs]
    # legacy snapshot (pre-L1, no key) still loads
    for t in snap["transactions"]:
        t.pop("entry_status")
    txs3, _, _, _ = snapshot_from_dict(snap)
    assert all(t.entry_status is None for t in txs3)


def test_formula_column_warns_but_never_aborts(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.append(["Date", "Amount", "Description", "Running"])
    ws.append(["2026-06-01", 10.0, "V1", "=B2"])
    ws.append(["2026-06-02", "=10+10", "V2", "=B3+D2"])  # formula in MAPPED Amount
    wb.save(tmp_path / "f.xlsx")

    txs, issues = parse_statement_xlsx_tolerant(
        tmp_path / "f.xlsx", COLUMN_MAP, "card-1", "le", "USD"
    )
    warnings = [i for i in issues if i.severity == "warning"]
    assert len(warnings) == 1
    assert "'Amount' is formula-derived" in warnings[0].message
    # the UNMAPPED Running column emits nothing
    assert not any("Running" in i.message for i in issues)
    # strict mode does NOT raise on the warning (a fresh programmatic
    # workbook has no cached formula values, so the formula cell itself
    # parses as an error-row in tolerant mode; strict must only raise for
    # severity=error rows -- assert the warning alone never raises)
    only_warnings_wb = Workbook()
    ws2 = only_warnings_wb.active
    ws2.append(["Date", "Amount", "Description"])
    ws2.append(["2026-06-01", 10.0, "=CONCAT(\"V\",1)"])  # formula in vendor
    only_warnings_wb.save(tmp_path / "w.xlsx")
    txs2 = parse_statement_xlsx(
        tmp_path / "w.xlsx", COLUMN_MAP, "card-1", "le", "USD"
    )  # must not raise: vendor formula has cached value None -> coerces to ""
    assert len(txs2) == 1


def test_fx_columns_xlsx(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.append(["Date", "Amount", "Description", "Original Amount", "Original Currency", "Exchange Rate"])
    ws.append(["2026-06-01", 58.10, "PADARIA SAO JOSE", 310.0, "brl", 0.187419])
    ws.append(["2026-06-02", 25.00, "US VENDOR", None, None, None])
    wb.save(tmp_path / "fx.xlsx")
    cmap = dict(
        COLUMN_MAP,
        original_amount="Original Amount",
        original_currency="Original Currency",
        fx_rate="Exchange Rate",
    )
    txs, issues = parse_statement_xlsx_tolerant(
        tmp_path / "fx.xlsx", cmap, "card-1", "le", "USD"
    )
    assert [i for i in issues if i.severity == "error"] == []
    brl, usd = txs
    assert brl.original_amount == Decimal("310.0")
    assert brl.original_currency == "BRL"
    assert brl.fx_rate == Decimal("0.187419")
    assert usd.original_amount is None and usd.original_currency is None


def test_fx_columns_csv(tmp_path):
    p = tmp_path / "fx.csv"
    p.write_text(
        "Date,Amount,Description,Original Amount,Original Currency,Exchange Rate\n"
        "2026-06-01,58.10,PADARIA SAO JOSE,310.00,brl,0.187419\n"
        "2026-06-02,25.00,US VENDOR,,,\n",
        encoding="utf-8",
    )
    cmap = dict(
        COLUMN_MAP,
        original_amount="Original Amount",
        original_currency="Original Currency",
        fx_rate="Exchange Rate",
    )
    txs, issues = parse_statement_csv_tolerant(p, cmap, "card-1", "le", "USD")
    assert issues == []
    assert txs[0].original_amount == Decimal("310.00")
    assert txs[0].original_currency == "BRL"
    assert txs[1].fx_rate is None


def test_fx_columns_feed_matching(tmp_path):
    """A BRL receipt matches on the statement's original BRL amount."""
    wb = Workbook()
    ws = wb.active
    ws.append(["Date", "Amount", "Description", "Original Amount", "Original Currency"])
    ws.append(["2026-06-01", 58.10, "PADARIA SAO JOSE", 310.0, "BRL"])
    wb.save(tmp_path / "fx.xlsx")
    cmap = dict(COLUMN_MAP, original_amount="Original Amount", original_currency="Original Currency")
    txs, _ = parse_statement_xlsx_tolerant(tmp_path / "fx.xlsx", cmap, "card-1", "le", "USD")
    receipt = Receipt(
        document_id="r1",
        legal_entity_id="le",
        detected_date=date(2026, 6, 1),
        detected_total=Decimal("310.00"),
        detected_currency="BRL",
        detected_vendor="Padaria Sao Jose",
    )
    outcome = match_month(txs, [receipt])
    all_pairs = list(outcome.matches) + list(outcome.judgment_required) + list(outcome.ambiguous)
    assert any(m.document_id == "r1" for m in all_pairs), "BRL receipt should pair"


def test_zoho_export_skips_posted_rows(tmp_path):
    path = tmp_path / "colored.xlsx"
    _write_colored_workbook(path)
    txs, _ = parse_statement_xlsx_tolerant(path, COLUMN_MAP, "card-1", "le", "USD")
    receipts = [
        Receipt(
            document_id=f"r{i}",
            legal_entity_id="le",
            detected_date=t.transaction_date,
            detected_total=t.amount,
            detected_currency="USD",
            detected_vendor=t.vendor_from_statement.title(),
        )
        for i, t in enumerate(txs)
    ]
    outcome = match_month(txs, receipts)
    matched_tx = {m.transaction_id for m in outcome.matches}
    posted_tx = {t.transaction_id for t in txs if t.entry_status == "posted"}
    assert posted_tx & matched_tx, "fixture should match the posted row"
    rows = build_journal_rows(
        outcome,
        {t.transaction_id: t for t in txs},
        {r.document_id: r for r in receipts},
    )
    refs = {row[3] for row in rows}  # Reference# column
    assert not (refs & posted_tx), "posted rows must never reach the journal"
    assert refs, "non-posted matched rows still export"
