"""Sheet writeback tests — L3 (2026-07-15 walkthrough).

Chris's own workbook comes back with ONE appended "Zoho Account (tool)"
column; her values, fills, and formulas survive byte-identical (the
load is data_only=False so formulas are never collapsed to cached
values).
"""
from __future__ import annotations

from datetime import date
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

from expense_recon.ingest.statement_xlsx import parse_statement_xlsx_tolerant
from expense_recon.matching.deterministic import match_month
from expense_recon.matching.types import (
    Categorization,
    ClassificationSource,
    LineItem,
    Match,
    MatchOutcome,
    MatchType,
    Receipt,
    Transaction,
)
from expense_recon.output.sheet_writeback import (
    WRITEBACK_HEADER,
    write_sheet_writeback,
)

COLUMN_MAP = {"transaction_date": "Date", "amount": "Amount", "vendor": "Description"}


def _write_workbook(path: Path, n_rows: int = 3) -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(["Date", "Amount", "Description"])
    for i in range(n_rows):
        ws.append([f"2026-06-0{i + 1}", 10.0 + i, f"VENDOR {i + 1}"])
    wb.save(path)


def _tx(row: int, entry_status: str | None = None, tid: str | None = None) -> Transaction:
    return Transaction(
        transaction_id=tid if tid is not None else f"card-1:{row}",
        legal_entity_id="le",
        account_id="card-1",
        transaction_date=date(2026, 6, 1),
        posting_date=None,
        amount=Decimal("10.00"),
        transaction_currency="USD",
        account_card_currency="USD",
        vendor_from_statement=f"VENDOR {row - 1}",
        entry_status=entry_status,
    )


def _line(desc: str, amount: str, account: str | None) -> LineItem:
    return LineItem(
        description=desc,
        line_total=Decimal(amount),
        categorization=Categorization(
            category="Equipment & Hardware", zoho_account=account,
            confidence=0.9, source=ClassificationSource.LINE, reasoning="t",
        ),
    )


def _receipt(doc_id: str = "r1", items=()) -> Receipt:
    return Receipt(
        document_id=doc_id, legal_entity_id="le",
        detected_date=date(2026, 6, 1), detected_total=Decimal("10.00"),
        detected_currency="USD", detected_vendor="Vendor",
        line_items=tuple(items),
    )


def _match(tid: str, doc_id: str = "r1") -> Match:
    return Match(tid, doc_id, MatchType.EXACT, 0.99, "x", False)


def _writeback_col(ws) -> int | None:
    hits = [c.column for c in ws[1] if c.value == WRITEBACK_HEADER]
    assert len(hits) <= 1, "writeback header must appear at most once"
    return hits[0] if hits else None


def test_column_appended_with_bold_header(tmp_path):
    src = tmp_path / "chris.xlsx"
    _write_workbook(src)
    outcome = MatchOutcome(unmatched_transactions=["card-1:2"])

    out = write_sheet_writeback(src, tmp_path / "out.xlsx", outcome, [_tx(2)], [])

    ws = load_workbook(out).active
    assert ws.max_column == 4  # her 3 columns + the appended one
    assert ws.cell(row=1, column=4).value == WRITEBACK_HEADER
    assert ws.cell(row=1, column=4).font.bold is True
    assert ws.cell(row=2, column=4).value == "(no receipt matched)"


def test_idempotent_rerun_reuses_column(tmp_path):
    src = tmp_path / "chris.xlsx"
    _write_workbook(src)
    first = write_sheet_writeback(
        src, tmp_path / "out1.xlsx",
        MatchOutcome(unmatched_transactions=["card-1:2"]), [_tx(2)], [],
    )
    # Second run takes the already-written-back file as its input and
    # lands a different verdict on the same row.
    out = write_sheet_writeback(
        first, tmp_path / "out2.xlsx",
        MatchOutcome(judgment_required=[_match("card-1:2")]), [_tx(2)], [],
    )

    ws = load_workbook(out).active
    assert ws.max_column == 4  # no second column appended
    assert _writeback_col(ws) == 4
    assert ws.cell(row=2, column=4).value == "(needs review)"  # overwritten


def test_existing_values_fill_and_formula_survive(tmp_path):
    src = tmp_path / "chris.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Date", "Amount", "Description", "Running"])
    ws.append(["2026-06-01", 10.0, "VENDOR 1", 10.0])
    ws.append(["2026-06-02", 20.0, "VENDOR 2", "=SUM(B2:B3)"])
    fill = PatternFill(start_color="FFFFEB9C", end_color="FFFFEB9C", fill_type="solid")
    ws["A2"].fill = fill
    wb.save(src)

    out = write_sheet_writeback(
        src, tmp_path / "out.xlsx",
        MatchOutcome(unmatched_transactions=["card-1:2", "card-1:3"]),
        [_tx(2), _tx(3)], [],
    )

    ws2 = load_workbook(out, data_only=False).active
    assert ws2["D3"].value == "=SUM(B2:B3)"  # formula string intact
    assert ws2["A2"].fill.start_color.rgb == "FFFFEB9C"  # her fill intact
    assert ws2["C2"].value == "VENDOR 1"
    assert ws2["B3"].value == 20.0
    assert ws2.cell(row=2, column=5).value == "(no receipt matched)"


def test_matched_transaction_gets_account(tmp_path):
    src = tmp_path / "chris.xlsx"
    _write_workbook(src, n_rows=1)
    txs, issues = parse_statement_xlsx_tolerant(src, COLUMN_MAP, "card-1", "le", "USD")
    assert [i for i in issues if i.severity == "error"] == []
    receipts = [
        Receipt(
            document_id="r1", legal_entity_id="le",
            detected_date=t.transaction_date, detected_total=t.amount,
            detected_currency="USD",
            detected_vendor=t.vendor_from_statement.title(),
            line_items=(_line("chair", "10.00", "6420 Office Equipment"),),
        )
        for t in txs
    ]
    outcome = match_month(txs, receipts)
    assert outcome.matches, "fixture should produce a match"

    out = write_sheet_writeback(src, tmp_path / "out.xlsx", outcome, txs, receipts)

    ws = load_workbook(out).active
    assert ws.cell(row=2, column=4).value == "6420 Office Equipment"


def test_multi_account_join(tmp_path):
    src = tmp_path / "chris.xlsx"
    _write_workbook(src)
    rec = _receipt(items=[
        _line("flight", "6.00", "6001 Travel"),
        _line("dinner", "3.00", "6002 Meals"),
        _line("taxi", "1.00", "6001 Travel"),  # duplicate account joins once
    ])
    outcome = MatchOutcome(matches=[_match("card-1:2")])

    out = write_sheet_writeback(src, tmp_path / "out.xlsx", outcome, [_tx(2)], [rec])

    ws = load_workbook(out).active
    assert ws.cell(row=2, column=4).value == "6001 Travel; 6002 Meals"


def test_already_posted_wins_over_match(tmp_path):
    src = tmp_path / "chris.xlsx"
    _write_workbook(src)
    rec = _receipt(items=[_line("chair", "10.00", "6420 Office Equipment")])
    outcome = MatchOutcome(matches=[_match("card-1:2")])

    out = write_sheet_writeback(
        src, tmp_path / "out.xlsx", outcome,
        [_tx(2, entry_status="posted")], [rec],
    )

    ws = load_workbook(out).active
    assert ws.cell(row=2, column=4).value == "(already in Zoho)"


def test_review_and_unmatched_placeholders(tmp_path):
    src = tmp_path / "chris.xlsx"
    _write_workbook(src)
    outcome = MatchOutcome(
        judgment_required=[_match("card-1:2")],
        ambiguous=[_match("card-1:3", "r2")],
        unmatched_transactions=["card-1:4"],
    )

    out = write_sheet_writeback(
        src, tmp_path / "out.xlsx", outcome, [_tx(2), _tx(3), _tx(4)], [],
    )

    ws = load_workbook(out).active
    assert ws.cell(row=2, column=4).value == "(needs review)"
    assert ws.cell(row=3, column=4).value == "(needs review)"  # ambiguous too
    assert ws.cell(row=4, column=4).value == "(no receipt matched)"


def test_unparseable_transaction_id_skipped(tmp_path):
    src = tmp_path / "chris.xlsx"
    _write_workbook(src)
    txs = [
        _tx(2, tid="pdf-tx-abc"),        # no ":row" tail at all
        _tx(3, tid="chase:notanint"),    # tail not an integer
        _tx(4),                          # sane anchor — still written
    ]
    outcome = MatchOutcome(
        unmatched_transactions=["pdf-tx-abc", "chase:notanint", "card-1:4"]
    )

    out = write_sheet_writeback(src, tmp_path / "out.xlsx", outcome, txs, [])

    ws = load_workbook(out).active
    assert ws.cell(row=4, column=4).value == "(no receipt matched)"
    # The unanchorable ids wrote nothing anywhere in the column.
    written = [
        ws.cell(row=r, column=4).value for r in range(2, ws.max_row + 1)
    ]
    assert written == [None, None, "(no receipt matched)"]
