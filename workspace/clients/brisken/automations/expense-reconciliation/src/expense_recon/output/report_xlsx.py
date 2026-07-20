"""Excel review report — BLUEPRINT LD-3 / LD-4.

Output structure (5 + N sheets, N = number of credit cards in the run):

    [ Summary ] [ {Card 1} ] [ {Card 2} ] ... [ Needs Review ] [ Unmatched ] [ Errors ]

Each matched transaction is expanded to one Excel row per line item on
its receipt (BLUEPRINT LD-2: one receipt → N journal entries). When a
receipt has no line items, the categorizer synthesizes a single
"(receipt total, no itemization)" line item and the row is marked
Tier 2 (VENDOR ⚠).

Row coloring per LD-4:

    LINE    → green   (Tier 1, trusted)
    LEARNED → blue    (Tier 1, recalled from a prior confirmed decision)
    VENDOR  → yellow  (Tier 2, confirm)
    REVIEW  → orange  (Tier 3, must touch)
    Unmatched tx / receipt → red-ish
"""
from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..matching.types import (
    Categorization,
    ClassificationSource,
    Match,
    MatchOutcome,
    MatchType,
    Receipt,
    Transaction,
)


FILL_LINE = PatternFill("solid", fgColor="FFC6EFCE")       # green
FILL_LEARNED = PatternFill("solid", fgColor="FFBDD7EE")    # blue (Tier 1, recalled)
FILL_VENDOR = PatternFill("solid", fgColor="FFFFEB9C")     # yellow
FILL_REVIEW = PatternFill("solid", fgColor="FFFCD5B4")     # orange
FILL_UNMATCHED = PatternFill("solid", fgColor="FFF8CBAD")  # red-ish
FILL_HEADER = PatternFill("solid", fgColor="FFD9E1F2")     # blue-grey
HEADER_FONT = Font(bold=True)


CARD_TAB_COLUMNS = (
    "Date",
    "Vendor",
    "Line item",
    "Qty",
    "Amount",
    "Category",
    "Source",
    "Zoho A/C",
    "Note",
)

NEEDS_REVIEW_COLUMNS = ("Card",) + CARD_TAB_COLUMNS


@dataclass
class _Row:
    """One Excel row before it's written. Carries the data + the tier
    so the writer can color uniformly across sheets.
    """

    card: str
    date: date | None
    vendor: str
    line_description: str
    quantity: Decimal | None
    amount: Decimal
    category: str | None
    source: ClassificationSource
    zoho_account: str | None
    note: str
    # A9: the originating transaction. A single transaction can expand
    # to many rows (one per receipt line item, and — pre-bipartite —
    # one set per FX candidate receipt). The Summary must aggregate by
    # transaction, never by row, so it carries the id here.
    transaction_id: str = ""

    @property
    def fill(self) -> PatternFill:
        return _fill_for_source(self.source)


def _fill_for_source(source: ClassificationSource) -> PatternFill:
    if source is ClassificationSource.LINE:
        return FILL_LINE
    if source is ClassificationSource.LEARNED:
        return FILL_LEARNED
    if source is ClassificationSource.VENDOR:
        return FILL_VENDOR
    return FILL_REVIEW


def write_report(
    outcome: MatchOutcome,
    transactions: list[Transaction],
    receipts: list[Receipt],
    out_path: str | Path,
    *,
    parse_errors: list[tuple[str, int, str]] | None = None,
    llm_cost: Decimal | None = None,
    explain: bool = False,
    charge_categorizations: dict[str, Categorization] | None = None,
) -> Path:
    """Write the LD-3 reconciliation report.

    `parse_errors` is the list of (file_name, line_number, message)
    tuples from ingest (slice 3a B1 — empty in slice 1).
    `llm_cost` is the run's aggregate Claude spend in USD
    (slice 2 D3 — None in slice 1).
    `explain` (A8) appends an "Explain" sheet: one row per transaction
    with its outcome bucket, confidence, and the scoring / judgment
    reason — the "why did (not) this match" debugging trail.
    `charge_categorizations` (Slice 10 side-map, transaction_id →
    Categorization) puts a category + source + account on the unmatched
    (receiptless) transaction rows; None => prior behaviour.
    """
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    tx_by_id = {tx.transaction_id: tx for tx in transactions}
    rec_by_id = {r.document_id: r for r in receipts}
    charge_cats = charge_categorizations or {}

    rows = _build_rows(outcome, tx_by_id, rec_by_id, charge_cats)
    rows_by_card = _group_by_card(rows)

    wb = Workbook()
    wb.remove(wb.active)

    _write_summary(wb, rows, outcome, transactions, receipts, llm_cost)
    for card in sorted(rows_by_card):
        _write_card_tab(wb, card, rows_by_card[card])
    _write_needs_review(wb, rows)
    _write_unmatched(wb, outcome, tx_by_id, rec_by_id, charge_cats)
    _write_errors(wb, parse_errors or [])
    if explain:
        _write_explain(wb, outcome, transactions)

    wb.save(out_path)
    return out_path


# ── row building ─────────────────────────────────────────────────────


def _build_rows(
    outcome: MatchOutcome,
    tx_by_id: dict[str, Transaction],
    rec_by_id: dict[str, Receipt],
    charge_cats: dict[str, Categorization] | None = None,
) -> list[_Row]:
    """Expand the matching outcome into per-line-item Excel rows.

    Every transaction surfaces in at least one row regardless of
    outcome. Matched txs expand by their receipt's line items;
    unmatched / FX / ambiguous txs become one or more REVIEW rows
    so the reconciliation guarantee (v2 spec §25.5) is visible.
    """
    rows: list[_Row] = []

    # L4: flag a matched expense with no receipt-image reference, but only
    # when this run's source carries image info at all (the slice-1
    # receipts CSV never does; flagging everything would be noise).
    has_image_info = any(r.has_receipt_image for r in rec_by_id.values())

    def _note(tx: Transaction, rec: Receipt, base: str) -> str:
        bits = [base] if base else []
        # L1: her sheet's fill-color annotation, carried into the note.
        if tx.entry_status == "posted":
            bits.append("already in Zoho (yellow row)")
        elif tx.entry_status == "subscription":
            bits.append("subscription (gray row)")
        if has_image_info and not rec.has_receipt_image:
            bits.append("missing receipt image")
        return " · ".join(bits)

    for match in outcome.matches:
        tx = tx_by_id.get(match.transaction_id)
        rec = rec_by_id.get(match.document_id)
        if tx is None or rec is None:
            continue
        rows.extend(_rows_from_match(tx, rec, match, extra_note=_note(tx, rec, "")))

    for match in outcome.judgment_required:
        tx = tx_by_id.get(match.transaction_id)
        rec = rec_by_id.get(match.document_id)
        if tx is None or rec is None:
            continue
        # FX cases ride on the matched receipt's line items but stay
        # REVIEW until the LLM judgment layer (slice 2) confirms.
        rows.extend(
            _rows_from_match(tx, rec, match, extra_note=_note(tx, rec, "FX — needs judgment"), force_review=True)
        )

    seen_ambiguous_tx: set[str] = set()
    for match in outcome.ambiguous:
        if match.transaction_id in seen_ambiguous_tx:
            continue  # only emit one row per ambiguous tx (top candidate)
        seen_ambiguous_tx.add(match.transaction_id)
        tx = tx_by_id.get(match.transaction_id)
        rec = rec_by_id.get(match.document_id)
        if tx is None or rec is None:
            continue
        rows.extend(
            _rows_from_match(tx, rec, match, extra_note=_note(tx, rec, "Ambiguous — pick one"), force_review=True)
        )

    for tx_id in outcome.unmatched_transactions:
        tx = tx_by_id.get(tx_id)
        if tx is None:
            continue
        rows.append(_unmatched_tx_row(tx, (charge_cats or {}).get(tx_id)))

    return rows


def _rows_from_match(
    tx: Transaction,
    receipt: Receipt,
    match: Match,
    *,
    extra_note: str,
    force_review: bool = False,
) -> list[_Row]:
    """Build the per-line rows for a single (tx, receipt) match.

    If the receipt has no line items (shouldn't happen after the
    categorizer synthesizes one, but defensive), the row falls back
    to the receipt total + REVIEW source.
    """
    note_bits: list[str] = []
    if extra_note:
        note_bits.append(extra_note)
    # Surface the scoring/judgment reasoning for the cases that carry an
    # informative one: PROBABLE tolerances, and the LLM verdicts from the
    # FX (D1b) and ambiguous (judge_ambiguous) judgment layers.
    if match.match_type in (
        MatchType.PROBABLE,
        MatchType.FX_JUDGMENT,
        MatchType.FX_REFERENCE,
        MatchType.AMBIGUOUS,
    ) and match.reason:
        note_bits.append(match.reason)
    note = " · ".join(note_bits)

    if not receipt.line_items:
        cat = Categorization(
            category=None, zoho_account=None,
            confidence=0.0,
            source=ClassificationSource.REVIEW,
            reasoning="receipt has no line items",
        )
        return [
            _Row(
                card=tx.account_id,
                date=tx.transaction_date,
                vendor=tx.vendor_from_statement,
                line_description="(no line items extracted)",
                quantity=None,
                amount=tx.amount,
                category=None,
                source=ClassificationSource.REVIEW,
                zoho_account=None,
                note=note,
                transaction_id=tx.transaction_id,
            )
        ]

    out: list[_Row] = []
    for item in receipt.line_items:
        cat = item.categorization or _default_categorization()
        source = cat.source if not force_review else ClassificationSource.REVIEW
        out.append(
            _Row(
                card=tx.account_id,
                date=tx.transaction_date,
                vendor=tx.vendor_from_statement,
                line_description=item.description,
                quantity=item.quantity,
                amount=item.line_total,
                category=cat.category if source is not ClassificationSource.REVIEW else None,
                source=source,
                zoho_account=cat.zoho_account if source is not ClassificationSource.REVIEW else None,
                note=note,
                transaction_id=tx.transaction_id,
            )
        )
    return out


def _unmatched_tx_row(
    tx: Transaction, charge_cat: Categorization | None = None
) -> _Row:
    """One row per receiptless charge. With a Slice-10 charge
    categorization, the row carries the category / account / source
    (LEARNED colors blue, VENDOR yellow) instead of a bare REVIEW row;
    the provenance rides in the note so Criss sees WHY."""
    if charge_cat is not None and charge_cat.category:
        note = "Unmatched transaction · categorized from statement description"
        if charge_cat.reasoning:
            note = f"{note} · {charge_cat.reasoning}"
        return _Row(
            card=tx.account_id,
            date=tx.transaction_date,
            vendor=tx.vendor_from_statement,
            line_description="(no receipt)",
            quantity=None,
            amount=tx.amount,
            category=charge_cat.category,
            source=charge_cat.source,
            zoho_account=charge_cat.zoho_account,
            note=note,
            transaction_id=tx.transaction_id,
        )
    return _Row(
        card=tx.account_id,
        date=tx.transaction_date,
        vendor=tx.vendor_from_statement,
        line_description="(no receipt)",
        quantity=None,
        amount=tx.amount,
        category=None,
        source=ClassificationSource.REVIEW,
        zoho_account=None,
        note="Unmatched transaction",
        transaction_id=tx.transaction_id,
    )


def _default_categorization() -> Categorization:
    return Categorization(
        category=None, zoho_account=None,
        confidence=0.0,
        source=ClassificationSource.REVIEW,
        reasoning="No categorization run",
    )


def _group_by_card(rows: list[_Row]) -> dict[str, list[_Row]]:
    out: dict[str, list[_Row]] = defaultdict(list)
    for r in rows:
        out[r.card].append(r)
    # chronological within each card
    for card in out:
        out[card].sort(key=lambda r: (r.date or date.min, r.vendor))
    return dict(out)


# ── sheet writers ────────────────────────────────────────────────────


def _write_summary(
    wb: Workbook,
    rows: list[_Row],
    outcome: MatchOutcome,
    transactions: list[Transaction],
    receipts: list[Receipt],
    llm_cost: Decimal | None,
) -> None:
    ws = wb.create_sheet("Summary")

    n_tx = len(transactions)
    n_rec = len(receipts)

    # A9: every Summary count is per DISTINCT transaction, never per
    # row. judgment_required / ambiguous hold one entry PER candidate
    # receipt, so a multi-candidate FX transaction would otherwise
    # inflate every count and money figure (a $8.8K month rendered as
    # $1.26M Spend with a false "invariant BROKEN" before this fix).
    matched_tx_ids = {m.transaction_id for m in outcome.matches}
    review_tx_ids = (
        {m.transaction_id for m in outcome.judgment_required}
        | {m.transaction_id for m in outcome.ambiguous}
    )
    unmatched_tx_ids = set(outcome.unmatched_transactions)
    refund_tx_ids = set(outcome.refunds)
    n_matched = len(matched_tx_ids)
    n_review = len(review_tx_ids)
    n_unmatched_tx = len(unmatched_tx_ids)
    n_refunds = len(refund_tx_ids)
    n_unmatched_rec = len(outcome.unmatched_receipts)
    # L1: fill-color annotations from the statement workbook.
    n_already_posted = sum(1 for t in transactions if t.entry_status == "posted")
    n_subscription = sum(1 for t in transactions if t.entry_status == "subscription")

    # Reconciliation invariant: every tx in exactly one bucket. The set
    # union catches BOTH silent drops and double-classification; the
    # count equality catches the same with a cheap arithmetic check.
    all_tx_ids = {tx.transaction_id for tx in transactions}
    invariant_ok = (
        (matched_tx_ids | review_tx_ids | unmatched_tx_ids | refund_tx_ids)
        == all_tx_ids
        and n_matched + n_review + n_unmatched_tx + n_refunds == n_tx
    )

    # Categorization tiers: count the postable journal rows only (rows
    # from matched transactions). FX / ambiguous / unmatched rows are
    # all REVIEW and not yet postable; counting them buries the real
    # categorization quality at "REVIEW 100%".
    matched_rows = [r for r in rows if r.transaction_id in matched_tx_ids]
    tier_counts: dict[ClassificationSource, int] = defaultdict(int)
    for r in matched_rows:
        tier_counts[r.source] += 1
    total_rows = sum(tier_counts.values()) or 1  # avoid div/0

    # Spend per card = sum of the actual card charges (one per
    # transaction, in the card currency), taken from the statement —
    # NOT from expanded rows.
    by_card_spend: dict[str, Decimal] = defaultdict(lambda: Decimal("0"))
    for tx in transactions:
        by_card_spend[tx.account_id] += tx.amount

    # "Needs attention" per card = distinct transactions on that card
    # that did not land a clean match (review + unmatched).
    needs_attention_tx_ids = review_tx_ids | unmatched_tx_ids
    by_card_attention: dict[str, int] = defaultdict(int)
    for tx in transactions:
        if tx.transaction_id in needs_attention_tx_ids:
            by_card_attention[tx.account_id] += 1

    # Category × card money: categorized journal money from matched
    # rows, plus one "(needs review)" bucket per card valued at the
    # transaction amount of each non-matched tx, so the cross-tab
    # totals reconcile back to Spend.
    cross_tab: dict[str, dict[str, Decimal]] = defaultdict(
        lambda: defaultdict(lambda: Decimal("0"))
    )
    for r in matched_rows:
        cat = r.category or "(unclassified)"
        cross_tab[cat][r.card] += r.amount
    for tx in transactions:
        if tx.transaction_id in needs_attention_tx_ids:
            cross_tab["(needs review)"][tx.account_id] += tx.amount
        elif tx.transaction_id in refund_tx_ids:
            # Refunds are their own bucket (3.10); keeping them in the
            # cross-tab makes its total reconcile back to Spend.
            cross_tab["(refunds / credits)"][tx.account_id] += tx.amount

    _append_section(ws, "RECONCILIATION SUMMARY")
    ws.append([])
    ws.append(["Transactions", n_tx])
    ws.append(["Receipts", n_rec])
    ws.append(["Matched", n_matched])
    ws.append(["Needs Review (FX / ambiguous)", n_review])
    ws.append(["Unmatched transactions", n_unmatched_tx])
    if n_refunds:
        ws.append(["Refunds / credits (own bucket, never receipt-matched)", n_refunds])
    ws.append(["Unmatched receipts", n_unmatched_rec])
    if n_already_posted:
        ws.append(["Already posted in Zoho (yellow rows)", n_already_posted])
    if n_subscription:
        ws.append(["Subscriptions (gray rows)", n_subscription])
    ws.append(["Reconciliation invariant", "OK" if invariant_ok else "BROKEN — investigate"])
    ws.append([])

    _append_section(ws, "Categorization source breakdown (postable journal rows)")
    ws.append([])
    for src in (ClassificationSource.LINE, ClassificationSource.VENDOR, ClassificationSource.REVIEW):
        count = tier_counts.get(src, 0)
        pct = f"{count / total_rows * 100:.0f}%"
        ws.append([f"  {src.value}", f"{count} ({pct})"])
    ws.append([])

    _append_section(ws, "By card")
    ws.append([])
    ws.append(["Card", "Spend", "Needs attention (tx)"])
    for card in sorted(by_card_spend):
        ws.append([
            card,
            float(by_card_spend[card]),
            by_card_attention.get(card, 0),
        ])
    ws.append([])

    _append_section(ws, "By category × by card")
    ws.append([])
    cards_sorted = sorted(by_card_spend)
    ws.append(["Category", *cards_sorted, "Total"])
    for cat in sorted(cross_tab):
        row = [cat]
        cat_total = Decimal("0")
        for card in cards_sorted:
            amt = cross_tab[cat].get(card, Decimal("0"))
            row.append(float(amt) if amt else "")
            cat_total += amt
        row.append(float(cat_total))
        ws.append(row)
    ws.append([])

    _append_section(ws, "LLM cost this run")
    ws.append([])
    ws.append([
        "  Estimated cost (USD)",
        float(llm_cost) if llm_cost is not None else "n/a (slice 1, no LLM)",
    ])

    ws.column_dimensions["A"].width = 56
    for col_idx in range(2, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14


def _append_section(ws, title: str) -> None:
    ws.append([title])
    ws.cell(row=ws.max_row, column=1).font = HEADER_FONT


def _write_card_tab(wb: Workbook, card: str, rows: list[_Row]) -> None:
    # Excel sheet names: max 31 chars, no `:\/?*[]`.
    ws = wb.create_sheet(_safe_sheet_name(card))
    _write_header_row(ws, CARD_TAB_COLUMNS)

    for r in rows:
        ws.append(_card_row_cells(r))
        _fill_last_row(ws, r.fill)

    # Subtotals at the bottom: per category + card total.
    if rows:
        ws.append([])
        ws.append(["Subtotals"])
        ws.cell(row=ws.max_row, column=1).font = HEADER_FONT

        by_cat: dict[str, Decimal] = defaultdict(lambda: Decimal("0"))
        for r in rows:
            cat = r.category or "(needs review)"
            by_cat[cat] += r.amount
        for cat in sorted(by_cat):
            ws.append([cat, "", "", "", float(by_cat[cat])])

        card_total = sum((r.amount for r in rows), start=Decimal("0"))
        ws.append([])
        ws.append(["Card total", "", "", "", float(card_total)])
        ws.cell(row=ws.max_row, column=1).font = HEADER_FONT
        ws.cell(row=ws.max_row, column=5).font = HEADER_FONT

    _autosize(ws)


def _write_needs_review(wb: Workbook, rows: list[_Row]) -> None:
    ws = wb.create_sheet("Needs Review")
    _write_header_row(ws, NEEDS_REVIEW_COLUMNS)

    # Sort: REVIEW first, then VENDOR ⚠, then by card + date.
    def sort_key(r: _Row):
        src_order = {
            ClassificationSource.REVIEW: 0,
            ClassificationSource.VENDOR: 1,
        }.get(r.source, 9)
        return (src_order, r.card, r.date or date.min)

    review_rows = [r for r in rows if r.source is not ClassificationSource.LINE]
    review_rows.sort(key=sort_key)

    for r in review_rows:
        ws.append([r.card, *_card_row_cells(r)])
        _fill_last_row(ws, r.fill)

    _autosize(ws)


def _write_unmatched(
    wb: Workbook,
    outcome: MatchOutcome,
    tx_by_id: dict[str, Transaction],
    rec_by_id: dict[str, Receipt],
    charge_cats: dict[str, Categorization] | None = None,
) -> None:
    ws = wb.create_sheet("Unmatched")
    charge_cats = charge_cats or {}

    ws.append(["Unmatched transactions"])
    ws.cell(row=ws.max_row, column=1).font = HEADER_FONT
    _write_header_row(
        ws,
        ("Card", "Date", "Vendor", "Amount", "Currency", "Category", "Zoho A/C", "Source"),
        start_row=2,
    )
    for tx_id in outcome.unmatched_transactions:
        tx = tx_by_id.get(tx_id)
        if tx is None:
            continue
        cat = charge_cats.get(tx_id)
        has_cat = cat is not None and bool(cat.category)
        ws.append([
            tx.account_id,
            tx.transaction_date.isoformat() if tx.transaction_date else "",
            tx.vendor_from_statement,
            float(tx.amount),
            tx.transaction_currency,
            cat.category if has_cat else "",
            (cat.zoho_account or "") if has_cat else "",
            cat.source.value if has_cat else "",
        ])
        # A categorized charge colors by its tier (LEARNED blue, VENDOR
        # yellow) so the postable ones read at a glance; bare unmatched
        # rows keep the red-ish fill.
        _fill_last_row(ws, _fill_for_source(cat.source) if has_cat else FILL_UNMATCHED)

    # Refunds / credits (3.10): their own section so a credit is never
    # read as a purchase awaiting a receipt. Reviewed, not receipt-matched.
    if outcome.refunds:
        blank_row = ws.max_row + 2
        ws.cell(
            row=blank_row, column=1,
            value="Refunds / credits (own bucket, never receipt-matched)",
        ).font = HEADER_FONT
        _write_header_row(
            ws, ("Card", "Date", "Vendor", "Amount", "Currency"),
            start_row=blank_row + 1,
        )
        for tx_id in outcome.refunds:
            tx = tx_by_id.get(tx_id)
            if tx is None:
                continue
            ws.append([
                tx.account_id,
                tx.transaction_date.isoformat() if tx.transaction_date else "",
                tx.vendor_from_statement,
                float(tx.amount),
                tx.transaction_currency,
            ])
            _fill_last_row(ws, FILL_REVIEW)

    blank_row = ws.max_row + 2
    ws.cell(row=blank_row, column=1, value="Unmatched receipts").font = HEADER_FONT
    _write_header_row(
        ws,
        ("Document ID", "Detected Date", "Detected Vendor", "Detected Total", "Detected Currency"),
        start_row=blank_row + 1,
    )
    for rec_id in outcome.unmatched_receipts:
        rec = rec_by_id.get(rec_id)
        if rec is None:
            continue
        ws.append([
            rec.document_id,
            rec.detected_date.isoformat() if rec.detected_date else "",
            rec.detected_vendor or "",
            float(rec.detected_total) if rec.detected_total is not None else "",
            rec.detected_currency or "",
        ])
        _fill_last_row(ws, FILL_UNMATCHED)

    _autosize(ws)


def _write_errors(wb: Workbook, parse_errors: list[tuple[str, int, str]]) -> None:
    ws = wb.create_sheet("Errors")
    _write_header_row(ws, ("File", "Line", "Error"))
    for file_name, line_no, msg in parse_errors:
        ws.append([file_name, line_no, msg])
        _fill_last_row(ws, FILL_UNMATCHED)
    _autosize(ws)


def _write_explain(
    wb: Workbook,
    outcome: MatchOutcome,
    transactions: list[Transaction],
) -> None:
    """A8: per-transaction outcome + scoring trail for debugging.

    One row per transaction, in input order: which bucket it landed in,
    the confidence, and the reason string the matcher / judgment layer
    attached. Makes "why did (not) this match" answerable at a glance.
    """
    ws = wb.create_sheet("Explain")
    _write_header_row(
        ws,
        ("Card", "Date", "Vendor", "Amount", "Currency", "Outcome", "Confidence", "Reason"),
    )

    # tx_id -> (label, Match | None). First write wins so a tx keeps its
    # primary bucket if it somehow appears in two.
    disposition: dict[str, tuple[str, Match | None]] = {}
    for m in outcome.matches:
        disposition.setdefault(m.transaction_id, ("MATCHED", m))
    for m in outcome.judgment_required:
        disposition.setdefault(m.transaction_id, ("FX_JUDGMENT", m))
    for m in outcome.ambiguous:
        disposition.setdefault(m.transaction_id, ("AMBIGUOUS", m))
    for tx_id in outcome.unmatched_transactions:
        disposition.setdefault(tx_id, ("UNMATCHED", None))
    for tx_id in outcome.refunds:
        disposition.setdefault(tx_id, ("REFUND", None))

    for tx in transactions:
        label, match = disposition.get(tx.transaction_id, ("UNKNOWN", None))
        if match:
            reason = match.reason
        elif label == "REFUND":
            reason = "Credit / refund; own bucket, never receipt-matched (LD-5 A5)."
        else:
            reason = "No candidate receipt."
        ws.append([
            tx.account_id,
            tx.transaction_date.isoformat() if tx.transaction_date else "",
            tx.vendor_from_statement,
            float(tx.amount),
            tx.transaction_currency,
            label,
            f"{match.confidence:.2f}" if match else "",
            reason,
        ])
        if label == "MATCHED":
            fill = FILL_LINE
        elif label in ("FX_JUDGMENT", "AMBIGUOUS", "REFUND"):
            fill = FILL_REVIEW
        else:
            fill = FILL_UNMATCHED
        _fill_last_row(ws, fill)

    _autosize(ws)


# ── cell / formatting helpers ────────────────────────────────────────


def _card_row_cells(r: _Row) -> list[object]:
    """Cells in CARD_TAB_COLUMNS order. Used by both card tabs and
    the Needs Review sheet (which prepends Card)."""
    source_label = r.source.value
    if r.source is ClassificationSource.VENDOR:
        source_label = "VENDOR (review)"
    return [
        r.date.isoformat() if r.date else "",
        r.vendor,
        r.line_description,
        float(r.quantity) if r.quantity is not None else "",
        float(r.amount),
        r.category or "",
        source_label,
        r.zoho_account or "",
        r.note,
    ]


def _write_header_row(ws, headers, start_row: int = 1) -> None:
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col, value=h)
        cell.fill = FILL_HEADER
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left")


def _fill_last_row(ws, fill: PatternFill) -> None:
    row = ws.max_row
    for col in range(1, ws.max_column + 1):
        ws.cell(row=row, column=col).fill = fill


def _autosize(ws) -> None:
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[letter]:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 50)


_SHEET_NAME_BAD = ":\\/?*[]"


def _safe_sheet_name(name: str) -> str:
    """Excel sheet names ≤ 31 chars, no `:\\/?*[]`."""
    cleaned = "".join(c if c not in _SHEET_NAME_BAD else "_" for c in name)
    return cleaned[:31]
