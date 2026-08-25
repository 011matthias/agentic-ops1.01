"""Sheet writeback — L3 (2026-07-15 walkthrough).

Chris uploads her own per-card .xlsx workbook as the statement. After
reconciliation this module writes HER OWN workbook back with ONE new
column appended — ``"Zoho Account (tool)"``, the resolved Zoho posting
account per statement row — and touches nothing else. Her existing
columns are never overwritten; her values, fills, and formulas survive
because the workbook is loaded with ``data_only=False``. Loading with
``data_only=True`` and saving would permanently replace every formula
with its cached value, so ``data_only=False`` is mandatory here.

openpyxl round-trip caveats (inherent to load/save, not fixable here):
charts, images, pivot tables, and some conditional-formatting
constructs do not survive the cycle. Cell values, formulas, fills,
fonts, number formats, and column widths do. Chris's per-card sheets
are plain tabular fills, so this is acceptable — and the original file
is never overwritten (output goes to ``out_path``).

Row anchor: tabular-statement transaction ids are
``"{account_id}:{row_index}"`` where ``row_index`` is the sheet row
(header = 1, data starts at 2) — see ``ingest/statement_xlsx.py``. Ids
of any other shape (PDF-parsed transactions) are skipped defensively:
they have no sheet row to anchor to.

Idempotent: re-running against an already written-back workbook finds
the existing ``"Zoho Account (tool)"`` header and reuses that column
instead of appending a second one; the rows it writes are overwritten
in place.

Scoped per file: a month can hold charges from several statement uploads
(the living month, PR 2b-2b-2), and a row number only addresses a cell in
the file it was read from. One charge legitimately sits in several of them
at different rows, because a mid-month partial and the closing cycle both
print it. So the caller passes ``anchors``, that WORKBOOK's own
transaction-id to row map, recorded when it was uploaded, and this module
writes exactly the charges that file contains, each at its own row.
"""
from __future__ import annotations

from pathlib import Path
from typing import TYPE_CHECKING

from openpyxl import load_workbook
from openpyxl.styles import Font

from ..matching.types import (
    Categorization,
    ClassificationSource,
    Match,
    MatchOutcome,
    Receipt,
    Transaction,
)
from .zoho_export import _resolve_account

if TYPE_CHECKING:
    from ..ingest.chart_of_accounts import ChartOfAccounts

WRITEBACK_HEADER = "Zoho Account (tool)"

# Placeholders per outcome bucket. `(uncategorized - assign)` mirrors
# the zoho_export flag so the two surfaces speak the same language.
_ALREADY_POSTED = "(already in Zoho)"
_UNCATEGORIZED = "(uncategorized - assign)"
_NEEDS_REVIEW = "(needs review)"
_NO_RECEIPT = "(no receipt matched)"


def _anchor_row(
    tx: "Transaction", anchors: "dict[str, int] | None" = None
) -> int | None:
    """The sheet row a transaction anchors to, or None when it has none
    (PDF-parsed transactions have no tabular row — skip, never crash).

    `anchors` is the workbook's OWN id-to-row map and wins outright when
    given, including its absences: a charge the file does not contain has no
    row in it and must not be written into it. This is what makes a month
    with several statements correct in both directions. Scoping on the
    charge's own `source_row` instead would annotate the closing cycle only
    for the charges the cycle introduced, leaving every row the mid-month
    partial had already printed blank in the file Criss actually works from.

    Without `anchors` (the CLI path, and runs that predate the map) the
    charge's own `source_row` is the answer, exactly as before.
    Transactions parsed before PR 2a do not: their ids were positional
    (``"{account_id}:{row_index}"``) and this function recovered the row
    by taking the id apart. Snapshots at rest still hold those ids, and a
    run created before PR 2a can still be re-exported, so the id-parsing
    path stays as the fallback. New parses never reach it.
    """
    if anchors is not None:
        row = anchors.get(tx.transaction_id)
        return row if row is not None and row >= 2 else None
    if tx.source_row is not None:
        return tx.source_row if tx.source_row >= 2 else None
    _, sep, tail = tx.transaction_id.rpartition(":")
    if not sep:
        return None
    try:
        row = int(tail)
    except ValueError:
        return None
    # Row 1 is the header; a parsed anchor below the data region would
    # clobber it. Defensive: data starts at 2.
    return row if row >= 2 else None


def _writeback_column(ws) -> int:
    """Find the writeback column in row 1, or create it at
    ``max_column + 1`` with a bold header."""
    for cell in ws[1]:
        if cell.value == WRITEBACK_HEADER:
            return cell.column
    col = ws.max_column + 1
    header = ws.cell(row=1, column=col, value=WRITEBACK_HEADER)
    header.font = Font(bold=True)
    return col


def _account_cell_value(
    rec: Receipt, coa: "ChartOfAccounts | None"
) -> str:
    """The matched receipt's posting account(s), for one cell.

    Each line item's categorization contributes ``zoho_account or
    category``; with a chart, the reference is resolved to the
    canonical Zoho account name (an unresolvable reference keeps the
    raw label — this is a review surface, so showing the pick beats
    hiding it). Distinct accounts join with "; " in first-seen order.
    No categorizations at all → flagged, never guessed.
    """
    values: list[str] = []
    for item in rec.line_items:
        cat = item.categorization
        if cat is None:
            continue
        ref = cat.zoho_account or cat.category
        if not ref:
            continue
        value = ref
        if coa is not None:
            value = _resolve_account(ref, coa) or ref
        if value not in values:
            values.append(value)
    if not values:
        return _UNCATEGORIZED
    return "; ".join(values)


def _cell_value(
    tx: Transaction,
    match_by_tx: dict[str, Match],
    rec_by_id: dict[str, Receipt],
    review_tx: set[str],
    unmatched_tx: set[str],
    coa: "ChartOfAccounts | None",
    charge_cats: "dict[str, Categorization] | None" = None,
) -> str | None:
    """The writeback cell value for one transaction, or None to leave
    the row untouched. Priority: posted > matched > review > unmatched
    (a yellow already-in-Zoho row wins even when it also matched).

    Slice 10: an unmatched charge with a side-map categorization writes
    its learned/guessed account instead of "(no receipt matched)" — a
    LEARNED account verbatim, a VENDOR guess with a "(confirm)" marker
    so the guess stays visible as a guess (LD-2)."""
    if tx.entry_status == "posted":
        return _ALREADY_POSTED
    match = match_by_tx.get(tx.transaction_id)
    if match is not None:
        rec = rec_by_id.get(match.document_id)
        if rec is None:
            return None  # defensive: match without its receipt
        return _account_cell_value(rec, coa)
    if tx.transaction_id in review_tx:
        return _NEEDS_REVIEW
    if tx.transaction_id in unmatched_tx:
        cat = (charge_cats or {}).get(tx.transaction_id)
        if cat is not None and cat.category:
            ref = cat.zoho_account or cat.category
            value = (_resolve_account(ref, coa) or ref) if coa is not None else ref
            if cat.source is ClassificationSource.LEARNED:
                return value
            return f"{value} (confirm)"
        return _NO_RECEIPT
    return None


def write_sheet_writeback(
    original_path: str | Path,
    out_path: str | Path,
    outcome: MatchOutcome,
    transactions: list[Transaction],
    receipts: list[Receipt],
    *,
    sheet_name: str | None = None,
    chart_of_accounts: "ChartOfAccounts | None" = None,
    charge_categorizations: "dict[str, Categorization] | None" = None,
    anchors: "dict[str, int] | None" = None,
) -> Path:
    """Write Chris's workbook back with the appended writeback column.

    Loads ``original_path`` (``data_only=False`` — mandatory, see module
    docstring; ``keep_vba`` for .xlsm), annotates ONLY the writeback
    column's cells + its header on the target sheet, and saves to
    ``out_path``. Returns ``out_path``.

    ``anchors`` is this workbook's own transaction-id to row map, so a month
    holding several statements writes each file exactly the charges it
    contains. Omit it and every charge is placed by its own ``source_row``,
    as before.
    """
    original_path = Path(original_path)
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    rec_by_id = {r.document_id: r for r in receipts}
    match_by_tx: dict[str, Match] = {}
    for m in outcome.matches:
        match_by_tx.setdefault(m.transaction_id, m)
    review_tx = {m.transaction_id for m in outcome.judgment_required} | {
        m.transaction_id for m in outcome.ambiguous
    }
    unmatched_tx = set(outcome.unmatched_transactions)

    wb = load_workbook(
        filename=original_path,
        data_only=False,
        keep_vba=(original_path.suffix.lower() == ".xlsm"),
    )
    try:
        ws = wb[sheet_name] if sheet_name is not None else wb.active
        col = _writeback_column(ws)
        for tx in transactions:
            row = _anchor_row(tx, anchors)
            if row is None:
                continue
            value = _cell_value(
                tx, match_by_tx, rec_by_id, review_tx, unmatched_tx,
                chart_of_accounts, charge_categorizations,
            )
            if value is None:
                continue
            ws.cell(row=row, column=col, value=value)
        wb.save(out_path)
    finally:
        wb.close()
    return out_path
