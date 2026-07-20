"""Pre-flight check — BLUEPRINT 5.14 / slice 5c.

`expense-recon doctor --config run.json` validates a run config WITHOUT
executing it: config JSON parses, files and folders exist, the statement
column_map covers the required logical fields and the mapped columns are
actually present in the file header, the receipt source resolves, the
`llm:` / `zoho:` blocks have their credentials in the environment, and
the output path is writable.

It is strictly read-only and makes NO network calls (it checks that the
OpenAI / Zoho env vars are *present*, never that they authenticate — that
would cost money and hit a live client system). The point is to catch the
boring failures (typo'd path, missing column, unset env var) on Chris's
machine before a real run, not to test connectivity.

Output is a banded report; exit code is 0 when there is no FAIL, 1 when
any check FAILs. WARN never changes the exit code.
"""
from __future__ import annotations

import argparse
import csv
import json
import os
import sys
from pathlib import Path

from .ingest._common import OPTIONAL_KEYS, REQUIRED_KEYS
from .ingest.receipts_csv import REQUIRED_RECEIPT_COLUMNS
from .ingest.receipts_folder import IMAGE_MIME

OK = "OK"
WARN = "WARN"
FAIL = "FAIL"

_VALID_STATEMENT_SUFFIXES = (".csv", ".xlsx", ".xlsm", ".pdf")

# 3.15 statement-source advisory thresholds (mirrored in
# web/service.py `_statement_source_advisory`): a receipt set is
# "foreign-heavy" when at least this many receipts, and at least this
# share of them, are in a currency other than the card's.
_FOREIGN_HEAVY_MIN = 3
_FOREIGN_HEAVY_SHARE = 0.3


class _Report:
    """Collects banded findings and renders them."""

    def __init__(self) -> None:
        self.findings: list[tuple[str, str, str]] = []

    def add(self, level: str, area: str, message: str) -> None:
        self.findings.append((level, area, message))

    def ok(self, area: str, message: str) -> None:
        self.add(OK, area, message)

    def warn(self, area: str, message: str) -> None:
        self.add(WARN, area, message)

    def fail(self, area: str, message: str) -> None:
        self.add(FAIL, area, message)

    @property
    def has_fail(self) -> bool:
        return any(level == FAIL for level, _, _ in self.findings)

    def render(self) -> str:
        lines = []
        for level, area, message in self.findings:
            lines.append(f"  [{level:4}] {area}: {message}")
        n_fail = sum(1 for level, _, _ in self.findings if level == FAIL)
        n_warn = sum(1 for level, _, _ in self.findings if level == WARN)
        lines.append("")
        if n_fail:
            lines.append(f"doctor: {n_fail} FAIL, {n_warn} WARN. Fix the FAILs before running.")
        elif n_warn:
            lines.append(f"doctor: 0 FAIL, {n_warn} WARN. Runnable; review the warnings.")
        else:
            lines.append("doctor: all checks passed.")
        return "\n".join(lines)


def _read_header(path: Path, sheet_name: str | None) -> list[str]:
    """Return the header row of a CSV or xlsx statement, lower-cased
    columns preserved as-is (we compare against the configured source
    names exactly). Raises on unreadable files."""
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as fh:
            reader = csv.reader(fh)
            for row in reader:
                return [c.strip() for c in row]
        return []
    # xlsx / xlsm
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name] if sheet_name else wb.active
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            return [str(c).strip() if c is not None else "" for c in row]
        return []
    finally:
        wb.close()


def _check_statement(report: _Report, cfg: dict, config_dir: Path) -> None:
    s = cfg.get("statement")
    if not isinstance(s, dict):
        report.fail("statement", "missing or not an object")
        return

    # The Chase statement PDF needs only path + legal_entity_id: account
    # ids come from the per-card markers inside the statement and there
    # is no column map (mirrors cli._load_statement).
    base_required = ("path", "legal_entity_id")
    missing = [k for k in base_required if k not in s]
    if missing:
        report.fail("statement", f"missing keys: {', '.join(missing)}")
        return

    path = (config_dir / s["path"]).resolve()
    if not path.exists():
        report.fail("statement", f"file not found: {path}")
        return
    suffix = path.suffix.lower()
    if suffix not in _VALID_STATEMENT_SUFFIXES:
        report.fail(
            "statement",
            f"unsupported file type {suffix!r} (use .csv/.xlsx/.xlsm/.pdf)",
        )
        return
    if suffix == ".pdf":
        report.ok(
            "statement",
            f"Chase statement PDF present: {path.name} (account ids come "
            f"from the per-card markers; no column map needed; carries "
            f"per-charge original-currency FX detail)",
        )
        return
    report.ok("statement", f"file present: {path.name}")

    required = ("account_id", "account_card_currency", "column_map")
    missing = [k for k in required if k not in s]
    if missing:
        report.fail("statement", f"missing keys: {', '.join(missing)}")
        return

    column_map = s["column_map"]
    if not isinstance(column_map, dict):
        report.fail("statement", "column_map is not an object")
        return
    missing_keys = [k for k in REQUIRED_KEYS if k not in column_map]
    if missing_keys:
        report.fail("statement", f"column_map missing required keys: {', '.join(missing_keys)}")
    else:
        report.ok("statement", f"column_map covers {', '.join(REQUIRED_KEYS)}")

    # The mapped source columns must actually exist in the file header.
    try:
        header = _read_header(path, s.get("sheet_name"))
    except Exception as exc:  # noqa: BLE001 — surface as a finding, don't crash
        report.fail("statement", f"could not read header: {exc}")
        return
    header_set = set(header)
    for logical, source in column_map.items():
        if logical not in (*REQUIRED_KEYS, *OPTIONAL_KEYS):
            report.warn("statement", f"column_map has unknown logical key {logical!r} (ignored at run time)")
            continue
        if source not in header_set:
            report.fail(
                "statement",
                f"column {source!r} (mapped from {logical!r}) not in header: {header}",
            )
    if header and not report.has_fail:
        report.ok("statement", f"all mapped columns present in header ({len(header)} columns)")


def _check_receipts(report: _Report, cfg: dict, config_dir: Path) -> None:
    r = cfg.get("receipts")
    if not isinstance(r, dict):
        report.fail("receipts", "missing or not an object")
        return
    if "path" not in r:
        report.fail("receipts", "path is required")
        return

    path = (config_dir / r["path"]).resolve()
    if not path.exists():
        report.fail("receipts", f"path not found: {path}")
        return

    # Source inference mirrors cli._load_receipts: directory -> folder,
    # .pdf -> the consolidated Zoho Expense report PDF, else csv.
    source = r.get("source") or (
        "folder" if path.is_dir()
        else "expense_report_pdf" if path.suffix.lower() == ".pdf"
        else "csv"
    )
    has_llm = isinstance(cfg.get("llm"), dict)

    if source == "folder":
        if not path.is_dir():
            report.fail("receipts", f"source 'folder' but {path} is not a directory")
            return
        if not has_llm:
            report.fail(
                "receipts",
                "source 'folder' needs an `llm:` block; OCR has no stub fallback",
            )
        files = [p for p in path.iterdir() if p.is_file() and not p.name.startswith(".")]
        supported = [p for p in files if p.suffix.lower() in IMAGE_MIME or p.suffix.lower() == ".pdf"]
        unsupported = [p for p in files if p not in supported]
        if not supported:
            report.warn("receipts", f"folder {path.name} has no image/PDF receipts yet")
        else:
            report.ok("receipts", f"folder mode: {len(supported)} receipt file(s) ready")
        if unsupported:
            names = ", ".join(p.name for p in unsupported[:5])
            report.warn(
                "receipts",
                f"{len(unsupported)} non-receipt file(s) will be skipped (Errors sheet): {names}",
            )
    elif source == "csv":
        if path.is_dir():
            report.fail("receipts", f"source 'csv' but {path} is a directory")
            return
        try:
            header = _read_header(path, None)
        except Exception as exc:  # noqa: BLE001
            report.fail("receipts", f"could not read CSV header: {exc}")
            return
        missing_cols = [c for c in REQUIRED_RECEIPT_COLUMNS if c not in set(header)]
        if missing_cols:
            report.fail("receipts", f"CSV missing required columns: {', '.join(missing_cols)}")
        else:
            report.ok("receipts", f"csv mode: required columns present")
    elif source == "expense_csv":
        if path.is_dir():
            report.fail("receipts", f"source 'expense_csv' but {path} is a directory")
            return
        column_map = r.get("column_map")
        if not isinstance(column_map, dict):
            report.fail("receipts", "source 'expense_csv' requires receipts.column_map")
            return
        try:
            header = _read_header(path, None)
        except Exception as exc:  # noqa: BLE001
            report.fail("receipts", f"could not read CSV header: {exc}")
            return
        header_set = set(header)
        missing_cols = [
            src for src in column_map.values() if src not in header_set
        ]
        if missing_cols:
            report.fail(
                "receipts",
                f"expense CSV missing mapped columns: {', '.join(missing_cols)}",
            )
        else:
            report.ok("receipts", "expense_csv mode: mapped columns present")
    elif source == "expense_report_pdf":
        if path.is_dir():
            report.fail(
                "receipts",
                f"source 'expense_report_pdf' but {path} is a directory",
            )
            return
        if path.suffix.lower() != ".pdf":
            report.fail(
                "receipts",
                f"source 'expense_report_pdf' needs a .pdf file, got {path.suffix!r}",
            )
            return
        report.ok(
            "receipts",
            f"expense report PDF present: {path.name} (text-layer parse, no LLM needed)",
        )
    else:
        report.fail(
            "receipts",
            f"source {source!r} not supported "
            f"(use 'csv', 'expense_csv', 'expense_report_pdf' or 'folder')",
        )


def _check_statement_source(report: _Report, cfg: dict, config_dir: Path) -> None:
    """3.15 statement-source advisory: WARN when a foreign-heavy receipt
    set is paired with a tabular (non-PDF) statement. Only the Chase
    statement PDF carries each foreign charge's original amount +
    currency — the input the deterministic exact-FX match consumes — so
    on a foreign-heavy month the statement source decides whether most
    matching is deterministic or lands in judgment. Advisory only:
    read-only, never raises, never changes the exit code."""
    s = cfg.get("statement")
    r = cfg.get("receipts")
    if not isinstance(s, dict) or not isinstance(r, dict):
        return
    stmt_path = s.get("path")
    rcpt_path = r.get("path")
    if not stmt_path or not rcpt_path:
        return
    if Path(stmt_path).suffix.lower() == ".pdf":
        return  # already the preferred source

    path = (config_dir / rcpt_path).resolve()
    if not path.exists() or path.is_dir():
        return  # folder OCR needs the LLM to know currencies; skip

    card_ccy = str(s.get("account_card_currency") or "USD").upper()
    currencies: list[str] = []
    try:
        suffix = path.suffix.lower()
        if suffix == ".pdf":
            from .ingest.expense_report_pdf import (
                parse_expense_report_pdf_tolerant,
            )

            receipts, _ = parse_expense_report_pdf_tolerant(
                path=path,
                legal_entity_id="doctor-preflight",
                default_currency=r.get("default_currency"),
            )
            currencies = [
                x.detected_currency for x in receipts if x.detected_currency
            ]
        elif suffix == ".csv":
            source = r.get("source") or "csv"
            col = (
                (r.get("column_map") or {}).get("currency")
                if source == "expense_csv"
                else "detected_currency"
            )
            if not col:
                return
            with path.open("r", encoding="utf-8-sig", newline="") as fh:
                for row in csv.DictReader(fh):
                    raw = (row.get(col) or "").strip()
                    if raw:
                        currencies.append(raw.upper())
        else:
            return
    except Exception:  # noqa: BLE001 — advisory only, never crash doctor
        return

    if not currencies:
        return
    foreign = sum(1 for c in currencies if c.upper() != card_ccy)
    if (
        foreign >= _FOREIGN_HEAVY_MIN
        and foreign / len(currencies) >= _FOREIGN_HEAVY_SHARE
    ):
        report.warn(
            "statement",
            f"{foreign} of {len(currencies)} receipts are foreign-currency "
            f"but the statement is {Path(stmt_path).suffix} — only the Chase "
            f"statement PDF carries each charge's original foreign amount, "
            f"which lets these match deterministically (no LLM). Prefer the "
            f"statement PDF for this month.",
        )


def _check_llm(report: _Report, cfg: dict) -> None:
    llm = cfg.get("llm")
    if not isinstance(llm, dict):
        report.ok("llm", "no `llm:` block; keyword-stub categorization (CSV receipts only)")
        return
    provider = llm.get("provider", "openai")
    if provider != "openai":
        report.fail("llm", f"provider {provider!r} not supported (only 'openai')")
    api_key_env = llm.get("api_key_env", "OPENAI_API_KEY")
    if not os.environ.get(api_key_env):
        report.fail("llm", f"env var {api_key_env!r} is not set")
    else:
        report.ok("llm", f"{api_key_env} is set (not validated against the API)")
    model = llm.get("model", "gpt-4o-mini")
    vision = llm.get("vision_model") or model
    report.ok("llm", f"model {model!r}, vision {vision!r}")


def _check_zoho(report: _Report, cfg: dict, config_dir: Path) -> None:
    z = cfg.get("zoho")
    if not isinstance(z, dict) or not z.get("enabled", True):
        report.ok("zoho", "no `zoho:` block; no chart-of-accounts / journal export")
        return

    source = z.get("coa_source", "api")
    if source == "api":
        creds = ("ZOHO_CLIENT_ID", "ZOHO_CLIENT_SECRET", "ZOHO_REFRESH_TOKEN", "ZOHO_ORG_ID")
        missing = [c for c in creds if not os.environ.get(c)]
        if missing:
            report.fail("zoho", f"coa_source 'api' but env vars unset: {', '.join(missing)}")
        else:
            report.ok("zoho", "API creds present (not validated against Zoho)")
    elif source == "csv":
        if "coa_csv_path" not in z:
            report.fail("zoho", "coa_source 'csv' requires coa_csv_path")
        else:
            coa = (config_dir / z["coa_csv_path"]).resolve()
            if not coa.exists():
                report.fail("zoho", f"chart-of-accounts CSV not found: {coa}")
            else:
                report.ok("zoho", f"chart-of-accounts CSV present: {coa.name}")
    else:
        report.fail("zoho", f"coa_source {source!r} not supported (use 'api' or 'csv')")

    if z.get("export_path"):
        export_parent = (config_dir / z["export_path"]).resolve().parent
        if not export_parent.exists():
            report.fail("zoho", f"export_path directory does not exist: {export_parent}")
        else:
            report.ok("zoho", "journal export will be written")
        # The balancing credit needs the statement's account_id in card_accounts.
        stmt = cfg.get("statement")
        card_accounts = z.get("card_accounts") or {}
        if isinstance(stmt, dict) and stmt.get("account_id"):
            acct = stmt["account_id"]
            if acct not in card_accounts:
                report.warn(
                    "zoho",
                    f"statement account_id {acct!r} not in card_accounts; "
                    f"its balancing credit will be flagged unmapped",
                )


def _check_output(report: _Report, cfg: dict, config_dir: Path) -> None:
    out_cfg = cfg.get("output") or {}
    out_path = (config_dir / (out_cfg.get("path") or "report.xlsx")).resolve()
    if not out_path.parent.exists():
        report.fail("output", f"output directory does not exist: {out_path.parent}")
    else:
        report.ok("output", f"report will be written to {out_path.name}")


def run_doctor(config_path: Path) -> int:
    """Validate a run config. Returns 0 (no FAIL) or 1 (any FAIL)."""
    config_path = config_path.resolve()
    report = _Report()

    if not config_path.exists():
        print(f"doctor: config file not found: {config_path}", file=sys.stderr)
        return 1
    try:
        cfg = json.loads(config_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        print(f"doctor: config is not valid JSON: {exc}", file=sys.stderr)
        return 1
    if not isinstance(cfg, dict):
        print("doctor: config root must be a JSON object", file=sys.stderr)
        return 1

    config_dir = config_path.parent
    print(f"doctor: checking {config_path.name}\n")
    _check_statement(report, cfg, config_dir)
    _check_receipts(report, cfg, config_dir)
    _check_statement_source(report, cfg, config_dir)
    _check_llm(report, cfg)
    _check_zoho(report, cfg, config_dir)
    _check_output(report, cfg, config_dir)

    print(report.render())
    return 1 if report.has_fail else 0


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        prog="expense-recon doctor",
        description="Pre-flight validation of a run config (read-only, no network).",
    )
    parser.add_argument(
        "--config", required=True, type=Path,
        help="Path to the JSON run config to validate.",
    )
    args = parser.parse_args(argv)
    return run_doctor(args.config)


if __name__ == "__main__":
    raise SystemExit(main())
