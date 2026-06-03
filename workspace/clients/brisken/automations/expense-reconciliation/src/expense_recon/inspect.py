"""Column-map inspector — ANNEALING B2.

Reads the header row of a statement file (CSV or xlsx) and prints a
suggested `column_map` JSON block for the run config. Heuristics use
case-insensitive regex against common bank-export header conventions
(English banks first; extend per-region as Chris's actual banks land).

Usage::

    expense-recon-inspect path/to/statement.csv
    expense-recon-inspect path/to/statement.xlsx --sheet "Sheet1"

Output is the `"column_map": { ... }` block plus, when a required
logical field can't be guessed, a `// TBD` note pointing at the
available columns. The user copies the block into their `run.json`
and edits the TBDs.

Conservative posture: the mapper never invents a mapping it isn't
≥80% sure about. False negatives ("can't guess, please fill in") are
fine; false positives (wrong header silently mapped) waste Chris's
debug time later.
"""
from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from pathlib import Path


# Regex patterns for each logical field. Order matters within a list:
# more-specific patterns first so "posting date" matches posting_date
# before being grabbed by the generic "date" pattern for transaction_date.
_HEURISTICS: dict[str, list[re.Pattern[str]]] = {
    "posting_date": [
        re.compile(r"^post(ing|ed)?\s*date$", re.I),
        re.compile(r"^posted$", re.I),
        re.compile(r"^settlement\s*date$", re.I),
    ],
    "transaction_date": [
        re.compile(r"^transaction\s*date$", re.I),
        re.compile(r"^trans\.?\s*date$", re.I),
        re.compile(r"^purchase\s*date$", re.I),
        re.compile(r"^date$", re.I),
        re.compile(r"^datum$", re.I),     # DE
        re.compile(r"^buchungsdatum$", re.I),  # DE Amex
    ],
    "amount": [
        re.compile(r"^amount$", re.I),
        re.compile(r"^transaction\s*amount$", re.I),
        re.compile(r"^value$", re.I),
        re.compile(r"^charge$", re.I),
        re.compile(r"^sum$", re.I),
        re.compile(r"^betrag$", re.I),    # DE
        re.compile(r"^montant$", re.I),   # FR
    ],
    "vendor": [
        re.compile(r"^description$", re.I),
        re.compile(r"^vendor$", re.I),
        re.compile(r"^merchant$", re.I),
        re.compile(r"^payee$", re.I),
        re.compile(r"^narration$", re.I),
        re.compile(r"^details$", re.I),
        re.compile(r"^name$", re.I),       # some EU exports
        re.compile(r"^empfänger$", re.I),  # DE
        re.compile(r"^empfaenger$", re.I),
        re.compile(r"^beschreibung$", re.I),  # DE
    ],
    "transaction_currency": [
        re.compile(r"^currency$", re.I),
        re.compile(r"^ccy$", re.I),
        re.compile(r"^waehrung$", re.I),   # DE (ASCII)
        re.compile(r"^währung$", re.I),    # DE (Unicode)
    ],
}


def guess_column_map(headers: list[str]) -> tuple[dict[str, str], list[str]]:
    """Map logical field name → source header. Returns (mapping, missing_required).

    Each header is matched against each logical field's patterns; the
    first logical field that matches wins ownership (so "Posting Date"
    doesn't end up under transaction_date even though "date" matches
    both). The mapping only contains confident matches.
    """
    mapping: dict[str, str] = {}
    headers_clean = [h.strip() for h in headers if h and h.strip()]

    # Greedy assignment: for each logical field in order, try each
    # header against each of that field's patterns; first hit wins.
    # A header that's been claimed by an earlier logical field is
    # not available for later ones.
    claimed: set[str] = set()
    for logical in ("posting_date", "transaction_date", "amount", "vendor", "transaction_currency"):
        for header in headers_clean:
            if header in claimed:
                continue
            for pattern in _HEURISTICS[logical]:
                if pattern.match(header):
                    mapping[logical] = header
                    claimed.add(header)
                    break
            if logical in mapping:
                break

    required = ("transaction_date", "amount", "vendor")
    missing_required = [k for k in required if k not in mapping]
    return mapping, missing_required


def read_csv_headers(path: Path) -> list[str]:
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        reader = csv.reader(fh)
        try:
            return next(reader)
        except StopIteration:
            return []


def read_xlsx_headers(path: Path, sheet_name: str | None = None) -> list[str]:
    # Lazy import to keep CSV-only invocations from paying the openpyxl
    # import cost.
    from openpyxl import load_workbook

    wb = load_workbook(filename=path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name] if sheet_name is not None else wb.active
        rows = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows)
        except StopIteration:
            return []
        return [
            (cell.strip() if isinstance(cell, str) else str(cell) if cell is not None else "")
            for cell in header_row
        ]
    finally:
        wb.close()


def inspect(path: Path, sheet_name: str | None = None) -> tuple[dict[str, str], list[str], list[str]]:
    """Read headers from `path` and return (mapping, missing_required, all_headers)."""
    suffix = path.suffix.lower()
    if suffix == ".csv":
        headers = read_csv_headers(path)
    elif suffix in (".xlsx", ".xlsm"):
        headers = read_xlsx_headers(path, sheet_name)
    else:
        raise ValueError(
            f"unsupported file extension {suffix!r}; expected .csv / .xlsx / .xlsm"
        )

    mapping, missing = guess_column_map(headers)
    return mapping, missing, [h for h in headers if h]


def format_output(
    mapping: dict[str, str], missing: list[str], all_headers: list[str]
) -> str:
    """Format the inspect result as a copy-paste-able block for run.json."""
    # Ordered output: transaction_date / amount / vendor first, then optionals.
    ordered_keys = (
        "transaction_date",
        "amount",
        "vendor",
        "posting_date",
        "transaction_currency",
    )
    ordered: dict[str, str] = {}
    for key in ordered_keys:
        if key in mapping:
            ordered[key] = mapping[key]

    body = json.dumps({"column_map": ordered}, indent=2)

    lines = [body]
    if missing:
        lines.append("")
        lines.append("// MISSING required field(s):")
        for key in missing:
            lines.append(f"//   {key}: TBD")
        lines.append("// Available headers in your file:")
        for h in all_headers:
            lines.append(f"//   - {h!r}")
    return "\n".join(lines)


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        prog="expense-recon-inspect",
        description=(
            "Inspect a bank/card statement file and suggest a column_map "
            "for the expense-recon run config (ANNEALING B2)."
        ),
    )
    parser.add_argument("path", type=Path, help="Statement file (.csv / .xlsx / .xlsm).")
    parser.add_argument(
        "--sheet", default=None,
        help="Worksheet name for .xlsx files. Defaults to the active sheet.",
    )
    args = parser.parse_args(argv)

    if not args.path.exists():
        print(f"ERROR: file not found: {args.path}", file=sys.stderr)
        return 2

    try:
        mapping, missing, headers = inspect(args.path, sheet_name=args.sheet)
    except ValueError as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 2

    print(format_output(mapping, missing, headers))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
