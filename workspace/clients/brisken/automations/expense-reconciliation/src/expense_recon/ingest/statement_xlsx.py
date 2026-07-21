"""Excel statement parser — v2 spec §7.1 (sibling to `statement_csv.py`).

Reads an .xlsx statement Brisken downloads today and produces a list
of `Transaction` objects. Same interface, same `StatementParseError`
posture, same column-map shape as the CSV parser; the only behavioral
differences are the Excel-native cell type handling described below.

Cell-type handling
------------------
openpyxl returns native Python objects from cells:

* `datetime.datetime` / `datetime.date` for real date cells → used
  directly via `.date()` (no string parsing).
* `int` / `float` for numeric cells → routed through
  ``Decimal(str(value))`` so the IEEE-754 binary noise that bites
  ``Decimal(5.75)`` is avoided.
* `str` for text cells → routed through the shared `parse_date` /
  `parse_amount` helpers so the same date formats and
  ``$ / , / (50.00)`` tolerances apply as in CSV.
* `None` / empty string → empty.

Row-numbering convention matches CSV: header is row 1, first data row
is row 2. ``StatementParseError.line_number`` is consistent across
formats.

Fill-color annotation (L1, 2026-07-15 walkthrough)
--------------------------------------------------
Chris's per-card workbook uses cell fill as data: a YELLOW row is
already entered in Zoho, a GRAY row is a subscription. The parser reads
the fills of the mapped columns' cells and annotates each transaction
with ``entry_status`` ("posted" / "subscription" / None). This needs
real ``Cell`` objects, so the workbook is loaded in normal (not
read-only, not values-only) mode; her sheets are small, memory is a
non-issue. The classification is tolerant of rgb / indexed / theme+tint
color types and errs toward None (an untagged row is always safe).

Formula-column scan (L6)
------------------------
``data_only=True`` collapses a formula cell to its cached value, which
is indistinguishable from typed bank data downstream. Her sheet's
running-amount column IS a formula she copy-pastes. A light second pass
(read-only, ``data_only=False``) checks the MAPPED columns for formula
cells and emits one ``ParseIssue(severity="warning")`` per affected
column, so a formula-derived column mapped as ``amount`` is visibly
flagged instead of silently trusted. Warnings never abort a run.
"""
from __future__ import annotations

from collections.abc import Mapping
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from ..matching.types import Transaction
from ._common import (
    OPTIONAL_KEYS,
    REQUIRED_KEYS,
    ParseIssue,
    StatementParseError,
    parse_amount,
    parse_date,
    validate_required_map,
)

__all__ = [
    "OPTIONAL_KEYS",
    "REQUIRED_KEYS",
    "StatementParseError",
    "parse_statement_xlsx",
    "parse_statement_xlsx_tolerant",
]

# ── L1 fill classification ────────────────────────────────────────────
#
# Tunable thresholds, unit-tested against the common Excel palette. Her
# real sheet may use a shade these misses; widen here, never inline.
#
# Yellow ("posted"): strongly red+green, weak blue. Catches FFFF00,
# FFFF99, FFE699, FFEB9C; excludes white (blue too high) and orange
# FFC000 (green too far below red).
# Gray ("subscription"): low saturation, mid lightness. Catches D9D9D9,
# BFBFBF, A6A6A6, 808080; excludes white and black.
_YELLOW_MIN_RED = 180
_YELLOW_GREEN_RATIO = 0.8
_YELLOW_BLUE_GAP = 40
_GRAY_MAX_SPREAD = 24
_GRAY_MIN_MEAN = 110
_GRAY_MAX_MEAN = 224

# Default-Office theme colors (index -> RGB hex) for `theme`-typed fills.
# Only the slots that plausibly produce yellows/grays matter; an unknown
# index resolves to None (unclassified is always safe).
_THEME_RGB = {
    0: "FFFFFF", 1: "000000", 2: "E7E6E6", 3: "44546A", 4: "4472C4",
    5: "ED7D31", 6: "A5A5A5", 7: "FFC000", 8: "5B9BD5", 9: "70AD47",
}


def _classify_rgb(r: int, g: int, b: int) -> str | None:
    if (
        r >= _YELLOW_MIN_RED
        and g >= _YELLOW_GREEN_RATIO * r
        and b <= min(r, g) - _YELLOW_BLUE_GAP
    ):
        return "posted"
    mean = (r + g + b) / 3
    if (
        max(r, g, b) - min(r, g, b) <= _GRAY_MAX_SPREAD
        and _GRAY_MIN_MEAN <= mean <= _GRAY_MAX_MEAN
    ):
        return "subscription"
    return None


def _apply_tint(component: int, tint: float) -> int:
    # The standard OOXML tint formula.
    if tint > 0:
        return round(component + (255 - component) * tint)
    if tint < 0:
        return round(component * (1 + tint))
    return component


def _fill_rgb(fill) -> tuple[int, int, int] | None:
    """The solid fill's RGB, or None when there is no classifiable solid
    fill (no fill, hatch patterns, system colors, unknown theme slots)."""
    try:
        if fill is None or fill.patternType != "solid":
            return None
        color = fill.fgColor
        if color is None:
            return None
        hex_rgb: str | None = None
        if color.type == "rgb" and isinstance(color.rgb, str):
            raw = color.rgb
            if len(raw) == 8:
                if raw == "00000000":  # the no-fill placeholder
                    return None
                raw = raw[2:]
            if len(raw) != 6:
                return None
            hex_rgb = raw
        elif color.type == "indexed":
            from openpyxl.styles.colors import COLOR_INDEX

            idx = color.indexed
            if idx is None or idx >= len(COLOR_INDEX) or idx in (64, 65):
                return None
            raw = COLOR_INDEX[idx]
            hex_rgb = raw[2:] if len(raw) == 8 else raw
        elif color.type == "theme":
            base = _THEME_RGB.get(color.theme)
            if base is None:
                return None
            r = int(base[0:2], 16)
            g = int(base[2:4], 16)
            b = int(base[4:6], 16)
            tint = float(color.tint or 0.0)
            return (_apply_tint(r, tint), _apply_tint(g, tint), _apply_tint(b, tint))
        if hex_rgb is None:
            return None
        return (int(hex_rgb[0:2], 16), int(hex_rgb[2:4], 16), int(hex_rgb[4:6], 16))
    except Exception:  # noqa: BLE001 - a weird color object is never fatal
        return None


def _classify_fill(cell) -> str | None:
    rgb = _fill_rgb(getattr(cell, "fill", None))
    if rgb is None:
        return None
    return _classify_rgb(*rgb)


def _row_entry_status(row_cells, mapped_indices: list[int]) -> str | None:
    """Classify a row from the fills of its MAPPED columns' cells (a
    full-row fill covers them; unmapped decoration columns don't vote).
    Majority over non-None classifications; a tie goes to "posted" (the
    annotation is display-plus-export-skip, so the stronger flag is the
    safe tie-break: a skipped-but-visible row beats a double-post)."""
    votes: dict[str, int] = {}
    for idx in mapped_indices:
        if idx < len(row_cells):
            label = _classify_fill(row_cells[idx])
            if label is not None:
                votes[label] = votes.get(label, 0) + 1
    if not votes:
        return None
    top = max(votes.values())
    winners = sorted(k for k, v in votes.items() if v == top)
    return "posted" if "posted" in winners else winners[0]


def _is_cell_empty(value: object) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and not value.strip():
        return True
    return False


def _coerce_str(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def _coerce_date(value: object) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return parse_date(_coerce_str(value))


def _coerce_amount(value: object) -> Decimal:
    # bool is a subclass of int in Python — reject explicitly so a stray
    # TRUE/FALSE cell never becomes Decimal(1) / Decimal(0).
    if isinstance(value, bool):
        raise ValueError(f"Not a number: {value!r}")
    if isinstance(value, (int, float)):
        # Decimal(str(5.75)) == Decimal("5.75"); Decimal(5.75) doesn't.
        return Decimal(str(value))
    return parse_amount(_coerce_str(value))


def parse_statement_xlsx(
    path: str | Path,
    column_map: Mapping[str, str],
    account_id: str,
    legal_entity_id: str,
    account_card_currency: str,
    sheet_name: str | None = None,
) -> list[Transaction]:
    """Strict mode: raise on the first row-level ERROR. Warnings (e.g.
    the L6 formula-column advisory) never raise.

    See `_common.py` module docstring for the header-vs-row error
    distinction. Backward-compatible with slice-1 callers.
    """
    transactions, issues = parse_statement_xlsx_tolerant(
        path, column_map, account_id, legal_entity_id, account_card_currency, sheet_name
    )
    hard = [i for i in issues if i.severity == "error"]
    if hard:
        raise hard[0].to_error()
    return transactions


def parse_statement_xlsx_tolerant(
    path: str | Path,
    column_map: Mapping[str, str],
    account_id: str,
    legal_entity_id: str,
    account_card_currency: str,
    sheet_name: str | None = None,
) -> tuple[list[Transaction], list[ParseIssue]]:
    """Tolerant mode (ANNEALING B1): collect row-level errors, continue."""
    validate_required_map(column_map)

    path = Path(path)
    file_name = path.name
    transactions: list[Transaction] = []
    issues: list[ParseIssue] = []
    # Normal (not read-only) load: L1 needs real Cell objects with full
    # style access for the fill classification. data_only=True gives the
    # cached VALUE of a formula cell (the L6 scan below flags those).
    wb = load_workbook(filename=path, data_only=True)
    try:
        ws = wb[sheet_name] if sheet_name is not None else wb.active

        rows = ws.iter_rows()
        try:
            header_cells = next(rows)
        except StopIteration:
            raise StatementParseError("Excel sheet has no rows") from None

        headers: list[str] = []
        for cell in header_cells:
            value = cell.value
            if value is None:
                headers.append("")
            elif isinstance(value, str):
                headers.append(value.strip())
            else:
                headers.append(str(value))

        if not any(headers):
            raise StatementParseError("Excel sheet has no header row")

        header_index: dict[str, int] = {}
        for i, h in enumerate(headers):
            if h:
                # First occurrence wins on duplicate headers; same as csv.DictReader.
                header_index.setdefault(h, i)

        for logical_key, source_col in column_map.items():
            if source_col not in header_index:
                raise StatementParseError(
                    f"Excel missing column {source_col!r} "
                    f"(mapped from logical key {logical_key!r}). "
                    f"Available columns: {[h for h in headers if h]}"
                )

        mapped_indices = sorted(
            {header_index[source_col] for source_col in column_map.values()}
        )

        for row_index, row_cells in enumerate(rows, start=2):
            row_values = tuple(c.value for c in row_cells)
            mapped: dict[str, object] = {}
            for logical, source_col in column_map.items():
                idx = header_index[source_col]
                mapped[logical] = row_values[idx] if idx < len(row_values) else None

            if all(_is_cell_empty(v) for v in mapped.values()):
                continue  # blank row, common at EOF

            try:
                txdate = _coerce_date(mapped["transaction_date"])
                amount = _coerce_amount(mapped["amount"])
                vendor = _coerce_str(mapped["vendor"])

                posting: date | None = None
                if "posting_date" in column_map and not _is_cell_empty(
                    mapped.get("posting_date")
                ):
                    posting = _coerce_date(mapped["posting_date"])

                tx_currency = account_card_currency.upper()
                if "transaction_currency" in column_map:
                    raw_cur = _coerce_str(mapped.get("transaction_currency"))
                    if raw_cur:
                        tx_currency = raw_cur.upper()

                # L5: optional per-charge FX detail (BRL on the USD card).
                original_amount: Decimal | None = None
                if "original_amount" in column_map and not _is_cell_empty(
                    mapped.get("original_amount")
                ):
                    original_amount = _coerce_amount(mapped["original_amount"])
                original_currency: str | None = None
                if "original_currency" in column_map:
                    raw_orig = _coerce_str(mapped.get("original_currency"))
                    if raw_orig:
                        original_currency = raw_orig.upper()
                fx_rate: Decimal | None = None
                if "fx_rate" in column_map and not _is_cell_empty(
                    mapped.get("fx_rate")
                ):
                    fx_rate = _coerce_amount(mapped["fx_rate"])

                # WS3: the per-row card, when the workbook prints one.
                card_last4: str | None = None
                if "card" in column_map:
                    card_last4 = _coerce_str(mapped.get("card")) or None
            except (KeyError, ValueError) as exc:
                issues.append(
                    ParseIssue(
                        file_name=file_name,
                        line_number=row_index,
                        message=str(exc),
                    )
                )
                continue

            raw_row = dict(
                zip(
                    [h for h in headers],
                    list(row_values) + [None] * max(0, len(headers) - len(row_values)),
                )
            )
            transactions.append(
                Transaction(
                    transaction_id=f"{account_id}:{row_index}",
                    legal_entity_id=legal_entity_id,
                    account_id=account_id,
                    transaction_date=txdate,
                    posting_date=posting,
                    amount=amount,
                    transaction_currency=tx_currency,
                    account_card_currency=account_card_currency.upper(),
                    vendor_from_statement=vendor,
                    raw_text=str(raw_row),
                    original_amount=original_amount,
                    original_currency=original_currency,
                    fx_rate=fx_rate,
                    entry_status=_row_entry_status(row_cells, mapped_indices),
                    card_last4=card_last4,
                )
            )
    finally:
        wb.close()

    issues.extend(
        _scan_formula_columns(path, sheet_name, column_map, file_name)
    )
    return transactions, issues


def _scan_formula_columns(
    path: Path,
    sheet_name: str | None,
    column_map: Mapping[str, str],
    file_name: str,
) -> list[ParseIssue]:
    """L6 — one warning per MAPPED column whose cells are formula-derived.

    The primary load uses data_only=True, which silently collapses a
    formula to its cached value; this streaming second pass (read-only,
    data_only=False) sees ``cell.data_type == 'f'`` and flags the column.
    Advisory only: never raises, never blocks the run."""
    issues: list[ParseIssue] = []
    try:
        wb = load_workbook(filename=path, read_only=True, data_only=False)
    except Exception:  # noqa: BLE001 - the primary load already succeeded
        return issues
    try:
        ws = wb[sheet_name] if sheet_name is not None else wb.active
        rows = ws.iter_rows()
        try:
            header_cells = next(rows)
        except StopIteration:
            return issues
        header_index: dict[str, int] = {}
        for i, cell in enumerate(header_cells):
            value = cell.value
            h = value.strip() if isinstance(value, str) else (str(value) if value is not None else "")
            if h:
                header_index.setdefault(h, i)
        watch: dict[int, str] = {}
        for logical, source_col in column_map.items():
            idx = header_index.get(source_col)
            if idx is not None:
                watch[idx] = source_col
        flagged: set[int] = set()
        for row_index, row_cells in enumerate(rows, start=2):
            for idx, source_col in watch.items():
                if idx in flagged or idx >= len(row_cells):
                    continue
                if row_cells[idx].data_type == "f":
                    flagged.add(idx)
                    issues.append(
                        ParseIssue(
                            file_name=file_name,
                            line_number=row_index,
                            message=(
                                f"column {source_col!r} is formula-derived: "
                                "cell values are cached formula results, not "
                                "bank statement data"
                            ),
                            severity="warning",
                        )
                    )
            if len(flagged) == len(watch):
                break
    except Exception:  # noqa: BLE001 - advisory scan must never break a parse
        return issues
    finally:
        wb.close()
    return issues
