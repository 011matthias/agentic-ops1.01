# /// script
# requires-python = ">=3.10"
# dependencies = ["openpyxl"]
# ///
"""
Rebuild the Rome-2026 post-event master contact sheet with a grounded, mutually
exclusive lead classification (H5 / T1 / T2 / T3) plus the supporting designation
columns.

Reads  : workspace/clients/brisken/context/lead-generation/Rome-Event/event-admin/
           rome2026-post-event-master-contacts.xlsx   (never modified)
         ...            /brisken-token-registrations.csv
Writes : output/leadgen-task-4/rome2026-post-event-master-contacts-v2.xlsx
         output/leadgen-task-4/lead-classification.csv

Grounding for the four lead classes (owner, 2026-07-09):
  H5 = the bespoke hottest-5 pack handed to Dirk to send himself.
       Roster is the literal To:/Cc: lines of
       deliverables/lead-generation/rome-2026/dirk-send-pack/README.md
  T1 = the 19 we sent from Dirk's Outlook on 2026-07-08
       (post_event_outreach == "Booth follow-up sent 2026-07-08")
  T2 = warm engaged: carries a real personal note from Dirk, reachable
  T3 = cold: reachable, no personal note; branches attended vs no-show

Every row lands in exactly one class. Non-lead rows get an explicit non-lead
class rather than being forced into a tier. Asserted at the end.

Run: uv run output/leadgen-task-4/build-master-v2.py
"""
import csv
import re
import sys
import unicodedata
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

REPO = Path(r"c:\Users\neuma_p1qrsic\Repo\agentic-ops1")
RE_DIR = REPO / "workspace/clients/brisken/context/lead-generation/Rome-Event"
SRC = RE_DIR / "event-admin/rome2026-post-event-master-contacts.xlsx"
TOKEN = RE_DIR / "event-admin/brisken-token-registrations.csv"
OUT_DIR = Path(__file__).parent
SNAPSHOT = OUT_DIR / "master-snapshot-2026-07-09T1833-298rows.csv"
OUT_XLSX = OUT_DIR / "rome2026-post-event-master-contacts-v2.xlsx"
OUT_CSV = OUT_DIR / "lead-classification.csv"

# --------------------------------------------------------------------------
# H5 roster: the literal To:/Cc: addresses of the six notes in the Dirk send pack.
# Source: deliverables/lead-generation/rome-2026/dirk-send-pack/README.md
# NOT a company-name match. A company match would sweep in 13 people who were
# never in the pack (Ana Matos, both Katkorias, Christian Forst, Miguel Carvalho,
# Lukas Blauth, Kenneth Bogert, five anonymized rows, Domenic).
# --------------------------------------------------------------------------
H5_PACK = {
    "michael.zucknick@volkswagen.de": ("hot5_vw", "To: VW note 1"),
    "steinar.pall.landroe@volkswagen.de": ("hot5_vw", "To: VW note 1b"),
    "jean-baptiste.disdet@jti.com": ("hot5_jti", "To: JTI note 2"),
    "paola.cuello@jti.com": ("hot5_jti", "To: JTI note 2"),
    "alejandro_jorge.herrera_la_grotta@roche.com": ("hot5_roche", "To: Roche note 3"),
    "dogan.yesil1@hotmail.com": ("hot5_roche", "Cc: Roche note 3"),
    "carol.tse@adidas.com": ("hot5_adidas", "To: Adidas note 4"),
    "alessandro.bonizzoni@lseg.com": ("hot5_lseg", "To: LSEG note 5"),
    "marco.favalli@lseg.com": ("hot5_lseg", "To: LSEG note 5"),
    "silvester.hetesi@lseg.com": ("hot5_lseg", "Cc: LSEG note 5"),
    "wiktor.jaszczak@lseg.com": ("hot5_lseg", "Cc: LSEG note 5"),
}

# canonical_account -> hot5 overlay (other people at the same five accounts)
HOT5_ACCOUNTS = {
    "Volkswagen AG": "hot5_vw",
    "JTI": "hot5_jti",
    "F. Hoffmann-La Roche AG": "hot5_roche",
    "adidas AG": "hot5_adidas",
    "LSEG": "hot5_lseg",
}

CANON = {
    "adidas": "adidas AG", "adidas ag": "adidas AG",
    "volkswagen": "Volkswagen AG", "volkswagen ag": "Volkswagen AG",
    "jti international": "JTI", "jti switzerland": "JTI", "jti": "JTI",
    "roche": "F. Hoffmann-La Roche AG",
    "f. hoffmann-la roche ag": "F. Hoffmann-La Roche AG",
    "f. hoffmann-la roche, ag": "F. Hoffmann-La Roche AG",
    "lseg": "LSEG",
    "dsv": "DSV A/S", "dsv a/s": "DSV A/S",
    "equinor": "Equinor ASA", "equinor asa": "Equinor ASA",
    "hitachi ltd": "Hitachi, Ltd.", "hitachi, ltd.": "Hitachi, Ltd.", "hitachi ltd.": "Hitachi, Ltd.",
    "robert bosch gmbh": "Robert Bosch GmbH",
    "shell": "Shell plc", "shell international": "Shell plc",
    "shell international limited": "Shell plc",
    "vodafone": "Vodafone", "vodafone group services gmbh": "Vodafone",
    "wiener städtische versicherung ag": "Wiener Staedtische Versicherung AG",
    "sap": "SAP SE", "sap se": "SAP SE",
    "nagarro": "Nagarro", "nagarro es gmbh": "Nagarro",
    "deloitte": "Deloitte", "deloitte gmbh": "Deloitte", "deloitte uk": "Deloitte",
    "akquinet": "AKQUINET",
    "ciments de l'atlas": "Ciments de l'Atlas",
    "tradeweb": "Tradeweb", "tradeweb - icd portal": "Tradeweb",
    "norsk hydro asa": "Norsk Hydro ASA", "hydro": "Norsk Hydro ASA",
    "nyk": "NYK Line", "nyk line": "NYK Line",
    "zatopek consulting": "Zatopek Consulting a.s.", "zatopek consulting, a.s.": "Zatopek Consulting a.s.",
    "intensum": "INTENSUM", "intensum luxembourg": "INTENSUM",
    "convista": "Convista", "convista ag": "Convista",
    "brisken": "Brisken",
}


def canon_account(company: str) -> str:
    c = company.strip()
    if not c:
        return ""
    key = unicodedata.normalize("NFKD", c).encode("ascii", "ignore").decode().lower().strip()
    key = re.sub(r"\s+", " ", key)
    if key in CANON:
        return CANON[key]
    raw = c.lower().strip()
    return CANON.get(raw, c)


# ---- verified raw-cell fixes (each new value traced to a source file) ----
FIX_NAME = {                       # email -> (first, last)
    "christos.kiosses@nagarro.com": ("Christos", "Kiosses"),   # reversed in TAC export
    "asako.teruki@nykgroup.com": ("Asako", "Teruki"),          # was all-lowercase
}
FIX_SALUTATION = {                 # email -> single given name for {First} merge
    "bettina.k.joergensen@dsv.com": "Bettina",
    "bandar.alghannam@aramco.com": "Bandar",
    "hardik.katkoria@adidas.com": "Hardik",
    "asako.teruki@nykgroup.com": "Asako",
    "christos.kiosses@nagarro.com": "Christos",
    # Owner 2026-07-09: salute as Steinar, like everyone else. Pall is a middle name.
    "steinar.pall.landroe@volkswagen.de": "Steinar",
}
FIX_TITLE = {                      # email -> title. Booth self-entry (token.csv) beats TAC's export.
    "erik.snersrud@hydro.com": "Global Head of Payments",
    "Pavitra.jogessar@fresenius.com": "VP - Risk Governance and Transformation",
    "cgeorgiou@bstdb.org": "Director IT",
    "dmorrison5@slb.com": "Global Treasury Digital and Performance Manager",
    "nedhal.abdulaal@aramco.com": "IT App/Sys Specialist - Projects Manager",
}
# Every key above must match a live row, or it is silently doing nothing.
# Sultan Alqahtani's 'Gm Treasury' was listed here against the wrong address and
# was in fact repaired by ACRONYMS. Asserted below rather than trusted.
ACRONYMS = {
    "Gm ": "GM ", "It ": "IT ", "Uk&i": "UK&I", "Apac": "APAC", "Erp": "ERP",
    "Sap": "SAP", "Vp ": "VP ", "Bp ": "BP ", "Cpa": "CPA", "Ceo": "CEO",
    "Cfo": "CFO", "Sr ": "Senior ", "app/sys": "App/Sys",
}
DUPLICATE_SUPPRESS = {"hardik1987@gmail.com": "hardik.katkoria@adidas.com"}
ICD_NOT_CUSTOMER = {  # Dirk's own dn_edits copy says No (SQL); the CRM domain-inferred a false Yes
    "eleni.souli@icdportal.com", "roderick.mackenzie@icdportal.com",
    "dan.staniford@icdportal.com", "sebastian.ramos@icdportal.com",
}
BAD_LINKEDIN = {"https://www.linkedin.com/in/at.linkedin.com"}

# No verified address, and never registered at the booth, so there is no
# self-entered address to fall back on. Owner rule 2026-07-09: if they were not
# at the booth their contact info is not in the token CSV, and no email goes out.
# The address on the row belongs to a different BSTDB person (R. Tsompani); it is
# moved to alt_email so no merge field can reach it.
EXCLUDE_FROM_EMAIL = {
    "rtsompani@bstdb.org":
        "not at the booth (absent from brisken-token-registrations.csv) and the "
        "address on file belongs to R. Tsompani, not Victoria Boclinca",
}
FREE_MAIL = ("@hotmail.com", "@gmail.com", "@seznam.cz", "@yahoo.com", "@gmx.", "@outlook.com")

# Owner rulings, 2026-07-09.
ENGAGED = {  # email -> (tier, why). Beats the SAP-partner deferral.
    "k.ashok@accenture.com": ("T2", "already responded to the Rome outreach; live "
                                    "Accenture MDH referral (owner, 2026-07-09)"),
}
ASK_DIRK = {  # email -> what we are waiting on. Blocks the send.
    "leonid.opanasik@gmail.com": "asked Dirk for a DSV corporate address (2026-07-09)",
}
FORST_GMAIL = "christian.forst@gmail.com"
FORST_CORP = "christian.forst@adidas-group.com"   # verified present in token.csv (he tapped twice)
MORRISON_TS = "2026-06-24T09:16:25.47173+00:00"   # token.csv; the sheet held Galera's timestamp

# Removed from the master on 2026-07-09 by owner decision: none of these people
# attended Rome, they arrived through the event campaign or as referrals, so they
# do not belong in a post-event attendee list. Their details are preserved in the
# account dossiers. Listed here so the drift check can tell a deliberate removal
# from a silent regression.
INTENTIONALLY_REMOVED = {
    "William Askew", "Alex Kerr", "Frank Appelman", "Kei-Fai Liew",   # -> account-shell.md 1a
    "Akash Gupta",            # Maersk, replied to the campaign, never attended
    "Isabelle Badoux",        # Sanofi, Sales Nav research add
    "Adela Dolezalova", "Maria Moeller",   # Zalando, referred by Lokesh Doggala
}

SENIORITY_BANDS = [
    ("c_suite", r"\bCEO\b|\bCFO\b|\bCOO\b|\bCTO\b|\bCMO\b|\bCRO\b|Chief |President|Founder|Owner|Managing Director|Board Member|\bPartner\b(?!ship)"),
    ("svp_vp", r"\bEVP\b|\bSVP\b|\bVP\b|Vice President|Global Head|Head Of|Head of|\bGM\b|General Manager"),
    ("director", r"Director|Abteilungsleitung|Group Treasurer"),
    ("senior_manager", r"Senior Manager|Principal|\bLead\b|Leader"),
    ("manager", r"Manager"),
    ("specialist_ic", r"Specialist|Consultant|Analyst|Advisor|Adviser|Engineer|Offic|Associate|\bStaff\b|Accountant|Architect|Expert|Treasurer|\bBA\b|Support|Executive"),
]


def seniority(title: str) -> str:
    t = (title or "").strip()
    if not t or t.upper() == "TBD":
        return "unknown"
    for band, pat in SENIORITY_BANDS:
        if re.search(pat, t, re.I):
            return band
    return "unknown"


def clean_first(first: str) -> str:
    """The given name a {First} merge should use. Drops a parenthetical nickname
    ('Hardik(Hrisha Papa)') and any middle name, so no salutation reads
    'Hi Hardik(Hrisha,'."""
    s = re.sub(r"\(.*?\)?$", "", (first or "").strip()).strip()
    return s.split(" ")[0] if s else ""


def fix_acronyms(t: str) -> str:
    # A title typed entirely in lower case ('assistant manager') reads as a typo
    # in a merge field. Title-case it before the acronym pass.
    if t and t == t.lower() and any(c.isalpha() for c in t):
        t = t.title()
    for a, b in ACRONYMS.items():
        if t.startswith(a):
            t = b + t[len(a):]
        t = t.replace(" " + a, " " + b)
    # A pipe inside a title breaks every markdown table and CSV consumer downstream.
    return t.replace(" | ", ", ").replace("|", ",")


# --------------------------------------------------------------------------
def ga_note(n: str) -> bool:
    """Dirk's general-awareness hold. Catches the two rows where he wrote the
    intent in prose instead of the bare 'GA' token."""
    s = (n or "").strip()
    return s.upper() == "GA" or s.rstrip(". ").endswith("=> GA") or "General awareness (GA)" in s


def artifact_note(n: str) -> bool:
    """A dirk_notes value that is bookkeeping, not an instruction to send a warm
    1:1. Without this, 15 routing artifacts get promoted into T2."""
    s = (n or "").strip().lower()
    return (
        s.startswith("cc:")
        or s.startswith("cc on the")
        or "see comment" in s
        or s == "visionary mail"
    )


def personal_note(n: str) -> bool:
    s = (n or "").strip()
    return bool(s) and not ga_note(s) and not artifact_note(s)


def read_live() -> tuple[list[str], list[dict]]:
    ws = openpyxl.load_workbook(SRC, data_only=True).active
    hdr = [str(c.value).strip() for c in ws[1]]
    rows = [
        {h: ("" if v is None else str(v).strip()) for h, v in zip(hdr, r)}
        for r in ws.iter_rows(min_row=2, values_only=True)
        if any(v is not None and str(v).strip() for v in r)
    ]
    return hdr, rows


def report_drift(rows: list[dict]) -> None:
    """The live sheet is co-authored on SharePoint and shared with parallel task
    sessions, so it can move under a build. Diff it against the pinned snapshot
    and account for every difference. The 8 removals in INTENTIONALLY_REMOVED
    were deliberate; anything else that vanished is unexplained and fails."""
    if not SNAPSHOT.exists():
        print("no pinned snapshot; skipping drift check")
        return
    with open(SNAPSHOT, encoding="utf-8") as f:
        snap = [{k: (v or "").strip() for k, v in r.items()} for r in csv.DictReader(f)]

    def key(r):
        return (r["first_name"].strip().lower(), r["last_name"].strip().lower(),
                r["company"].strip().lower())

    def name(r):
        return f'{r["first_name"]} {r["last_name"]}'.strip()

    live_keys = {key(r) for r in rows}
    snap_keys = {key(r) for r in snap}
    gone = [r for r in snap if key(r) not in live_keys]
    added = [r for r in rows if key(r) not in snap_keys]
    unexplained = [r for r in gone if name(r) not in INTENTIONALLY_REMOVED]

    print(f"drift check: snapshot {len(snap)} rows -> live {len(rows)} rows")
    print(f"  removed: {len(gone)} ({len(gone) - len(unexplained)} known-intentional, "
          f"{len(unexplained)} unexplained)")
    for r in gone:
        tag = "intentional" if name(r) in INTENTIONALLY_REMOVED else "UNEXPLAINED"
        print(f'    [{tag}] {name(r)} | {r["company"]}')
    for r in added:
        print(f'    [new] {name(r)} | {r["company"]}')
    assert not unexplained, (
        "rows vanished from the live sheet that are not on the intentional-removal "
        "list: " + str([name(r) for r in unexplained]))


def main() -> int:
    # The live sheet is the source of truth for WHO is a lead. The pinned snapshot
    # exists only to prove that every row that left did so on purpose.
    hdr, rows = read_live()
    print(f"loaded {len(rows)} rows x {len(hdr)} cols from the live {SRC.name}")
    report_drift(rows)
    print()

    token_emails = {
        (d["email"] or "").strip().lower() for d in csv.DictReader(open(TOKEN, encoding="utf-8"))
    }
    assert FORST_CORP in token_emails, "Forst corporate email not in token.csv; refusing to invent it"

    # A fix keyed on an address that no row carries repairs nothing and reads as
    # coverage. Fail loudly instead.
    live = {r["email"].strip().lower() for r in rows if r["email"].strip()}
    for label, keys in (("FIX_NAME", FIX_NAME), ("FIX_SALUTATION", FIX_SALUTATION),
                        ("FIX_TITLE", FIX_TITLE), ("ICD_NOT_CUSTOMER", ICD_NOT_CUSTOMER),
                        ("DUPLICATE_SUPPRESS", DUPLICATE_SUPPRESS), ("H5_PACK", H5_PACK),
                        ("EXCLUDE_FROM_EMAIL", EXCLUDE_FROM_EMAIL), ("ENGAGED", ENGAGED),
                        ("ASK_DIRK", ASK_DIRK)):
        dead = [k for k in keys if k.lower() not in live]
        assert not dead, f"{label} has keys matching no row: {dead}"

    # ---------------- pass 1: raw-cell repairs ----------------
    fixlog: list[str] = []
    for r in rows:
        em = r["email"].strip()
        eml = em.lower()

        if eml in FIX_NAME:
            f, l = FIX_NAME[eml]
            if (r["first_name"], r["last_name"]) != (f, l):
                fixlog.append(f'name  {em}: {r["first_name"]} {r["last_name"]} -> {f} {l}')
                r["first_name"], r["last_name"] = f, l

        if em in FIX_TITLE or eml in FIX_TITLE:
            new = FIX_TITLE.get(em) or FIX_TITLE[eml]
            if r["job_title"] != new:
                fixlog.append(f'title {em}: {r["job_title"]!r} -> {new!r}')
                r["job_title"] = new
        elif r["job_title"]:
            t2 = fix_acronyms(r["job_title"])
            if t2 != r["job_title"]:
                fixlog.append(f'title {em or r["company"]}: {r["job_title"]!r} -> {t2!r}')
                r["job_title"] = t2

        if not r["job_title"] and (r["first_name"] or r["last_name"]):
            r["job_title"] = "TBD"

        if r["linkedin_url"] in BAD_LINKEDIN:
            fixlog.append(f'linkedin {em}: dropped {r["linkedin_url"]!r} (not a person profile)')
            r["linkedin_url"] = ""

        if eml in ICD_NOT_CUSTOMER and r["brisken_customer"] == "Yes":
            fixlog.append(f'customer {em}: Yes -> No (SQL)   [Dirk dn_edits is authoritative]')
            r["brisken_customer"] = "No (SQL)"

        if eml == FORST_GMAIL:
            fixlog.append(f"email {em}: promote corporate {FORST_CORP}, gmail -> alt_email")
            r["email"], r["alt_email"] = FORST_CORP, FORST_GMAIL
            r["source"], r["no_show"] = "Booth only", "No"

        if eml == "dmorrison5@slb.com":
            if r["booth_registered_at"] != MORRISON_TS:
                fixlog.append(f'booth_ts {em}: {r["booth_registered_at"]} -> {MORRISON_TS} (was Galera\'s)')
                r["booth_registered_at"] = MORRISON_TS
            if r["alt_email"].strip().lower() == eml:
                r["alt_email"] = ""
                fixlog.append(f"alt_email {em}: cleared self-duplicate")

        if eml == "egalera@slb.com":
            if r["alt_email"].strip().lower() == "dmorrison5@slb.com":
                fixlog.append(f"alt_email {em}: cleared cross-wired dmorrison5@slb.com")
                r["alt_email"] = ""
            if r["source"] == "Booth":
                r["source"] = "Booth only"

        if eml in EXCLUDE_FROM_EMAIL:
            fixlog.append(f"email {em}: cleared primary, moved to alt_email "
                          f"({EXCLUDE_FROM_EMAIL[eml]})")
            r["_excluded_email"] = EXCLUDE_FROM_EMAIL[eml]
            r["alt_email"], r["email"] = r["email"], ""

        if r["first_name"].strip() == "Domenic" and r["source"] == "TAC only":
            r["source"] = "Booth only"   # if_we_know_them: "he also was at our booth"
            fixlog.append("source Domenic/JTI: TAC only -> Booth only")

    # ---------------- pass 2: derived designation columns ----------------
    def has_channel(r) -> bool:
        return any(r[c].strip() for c in ("email", "alt_email", "phone", "linkedin_url"))

    def classify(r) -> tuple[str, str]:
        eml = r["email"].strip().lower()
        dom = eml.split("@")[-1] if "@" in eml else ""
        acct = r["_canon"]
        if r["brisken_customer"] == "test row" or "example@ex.com" in eml:
            return "TEST", "test fixture row"
        if dom in ("brisken.com", "unpauseai.com") or acct == "Brisken" or "unpause" in r["company"].lower():
            return "OWN_TEAM", "Brisken / UnpauseAI staff"
        if eml in DUPLICATE_SUPPRESS:
            return "DUPLICATE", f"same person as {DUPLICATE_SUPPRESS[eml]}"
        if r["attendee_type"] == "TAC":
            return "ORGANISER", "TA Cook event organiser, own relationship thread"
        if not r["first_name"].strip() and not r["last_name"].strip():
            return "ANON", "TA Cook sponsor_opt_in=No; PII withheld, org-only"
        if eml in H5_PACK:
            return "H5", f"in Dirk's bespoke send pack ({H5_PACK[eml][1]})"
        if r["post_event_outreach"].startswith("Booth follow-up sent"):
            return "T1", "emailed from Dirk's Outlook 2026-07-08"
        if r["stop"].upper() == "X":
            return "STOP", "stop list: competitor / SI / never-contact"
        if eml in ENGAGED:
            return ENGAGED[eml]
        if ga_note(r["dirk_notes"]):
            # General awareness means NOT a warm lead (owner, 2026-07-09).
            return "GA", "dirk_notes = general awareness, not a warm lead; hold"
        # A personal note from Dirk outranks the structural SAP deferral: his
        # own Tier-2 roster names Eprox, Nagarro, SAP and Zanders people, all
        # of whom carry attendee_type = SAP partner / SAP employee.
        if personal_note(r["dirk_notes"]) and has_channel(r):
            return "T2", "warm: carries a personal note from Dirk"
        if r["attendee_type"] in ("SAP partner", "SAP employee", "SAP analyst") and acct != "LSEG":
            return "DEFERRED", f"Dirk's rule: {r['attendee_type']} deferred (no personal note; LSEG carve-out does not apply)"
        if not has_channel(r):
            return "UNREACHABLE", "no email, phone or LinkedIn on file"
        return "T3", "cold: reachable, no personal note"

    for r in rows:
        r["_canon"] = canon_account(r["company"])
    # lead_type is orthogonal to lead_class. A person can be a warm T2 contact AND a
    # partner rather than a treasury prospect; Dirk wrote "personal outreach DN" on 13
    # Deloitte / KPMG / PwC / Nagarro / Zanders / SAP rows. They earn the warm touch,
    # they must never receive the treasury pitch.
    LEAD_TYPE = {
        "SAP customer": "prospect", "Other": "prospect", "Prospect (non-attendee)": "prospect",
        "SAP partner": "partner_si", "SAP employee": "sap_internal", "SAP analyst": "analyst",
        "TAC": "organiser", "BRISKEN": "internal", "EXAMPLE": "internal",
    }

    for r in rows:
        cls, why = classify(r)
        eml = r["email"].strip().lower()
        r["lead_class"] = cls
        r["lead_class_reason"] = why
        if cls in ("OWN_TEAM", "TEST"):
            r["lead_type"] = "internal"
        elif cls == "ORGANISER":
            r["lead_type"] = "organiser"
        else:
            r["lead_type"] = LEAD_TYPE.get(r["attendee_type"], "unknown")
        r["t3_branch"] = ("no_show" if r["no_show"] == "Yes" else "attended") if cls == "T3" else ""
        r["canonical_account"] = r["_canon"]
        r["priority_account"] = (
            H5_PACK[eml][0] if eml in H5_PACK else HOT5_ACCOUNTS.get(r["_canon"], "")
        )
        r["salutation_first"] = FIX_SALUTATION.get(eml, clean_first(r["first_name"]))
        r["seniority"] = seniority(r["job_title"]) if cls not in ("ANON",) else "unknown"
        r["is_customer"] = "TRUE" if r["brisken_customer"].strip() == "Yes" else "FALSE"
        r["duplicate_of"] = DUPLICATE_SUPPRESS.get(eml, "")

        if not r["first_name"].strip() and not r["last_name"].strip():
            r["contactability"] = "tac_optout_anon"
        elif r["fob_encoded"].lower() == "true":
            r["contactability"] = "direct_booth"
        elif r["sponsor_opt_in"] == "Yes":
            r["contactability"] = "tac_optin"
        elif "Referral" in r["source"]:
            r["contactability"] = "referral"
        elif "Sales Nav" in r["source"]:
            r["contactability"] = "salesnav_prospect"
        elif r["if_we_know_them"].strip() or r["crm_last_activity"].strip():
            r["contactability"] = "personal_relationship"
        else:
            r["contactability"] = "unverified"

        r["booth_network_send"] = "TRUE" if (
            r["fob_encoded"].lower() == "true"
            and r["stop"].upper() != "X"
            and cls not in ("ORGANISER", "TEST", "OWN_TEAM", "DUPLICATE")
            and r["email"].strip()
        ) else "FALSE"

        hold = "none"
        if r.get("_excluded_email"):
            hold = "excluded_no_verified_email"
        elif eml in ASK_DIRK:
            hold = "owner_decision"
        elif eml.endswith(FREE_MAIL) and eml not in token_emails and cls in ("H5", "T1", "T2", "T3"):
            # A free-mail address the person typed into the Brisken Token themselves
            # IS their chosen channel, and it carries the direct_booth lawful basis.
            # Only an unverified free-mail address, one they never gave us at the
            # booth, needs a corporate address before a send.
            hold = "needs_corporate_email"
        elif cls == "UNREACHABLE":
            hold = "needs_enrichment"
        r["send_hold"] = hold

        base = {
            "H5": "dirk_bespoke_pack", "T1": "tier1_booth_followup",
            "T2": "tier2_dirk_personal", "T3": "tier3_sequence",
            "ORGANISER": "dirk_organiser_thread",
        }.get(cls, "none")
        if eml in ENGAGED:
            base = "dirk_referral_thread"
        elif r["_canon"] == "Shell plc" and r["no_show"] == "Yes":
            base = "dirk_askew_thread"
        elif r["dirk_notes"].strip().lower().startswith("cc:"):
            base = "cc_on_georgiou"
        elif r["contactability"] == "referral":
            base = "dirk_referral_thread"
        if hold != "none":
            base = "none"
        r["email_owner"] = base

        if not r["linkedin_url"].strip() or cls not in ("H5", "T1", "T2", "T3", "ORGANISER"):
            r["linkedin_owner"] = "none"
        elif cls in ("H5", "T1", "T2", "ORGANISER"):
            r["linkedin_owner"] = "dirk"
        else:
            r["linkedin_owner"] = "matthias"

    # ---------------- assertions: the partition must hold ----------------
    from collections import Counter
    c = Counter(r["lead_class"] for r in rows)
    assert sum(c.values()) == len(rows), f"partition lost rows: {sum(c.values())} of {len(rows)}"
    assert c["H5"] == 11, f"H5 must be the 11 send-pack addresses, got {c['H5']}"
    assert c["T1"] == 19, f"T1 must be the 19 Outlook sends, got {c['T1']}"
    # 90 rows have both names blank (TA Cook withheld PII, sponsor_opt_in=No).
    # One of those 90 is Brisken's own company row, which OWN_TEAM claims first.
    blank_both = sum(1 for r in rows if not r["first_name"].strip() and not r["last_name"].strip())
    assert blank_both == 90, f"expected 90 PII-withheld rows, got {blank_both}"
    assert c["ANON"] == 89, f"ANON must be 89 (90 blank-name minus Brisken's own row), got {c['ANON']}"
    assert c["OWN_TEAM"] == 4, f"OWN_TEAM must be 4, got {c['OWN_TEAM']}"
    assert c["ORGANISER"] == 1 and c["TEST"] == 1
    assert sum(1 for r in rows if r["fob_encoded"].lower() == "true") == 91
    h5, t1 = {r["email"].lower() for r in rows if r["lead_class"] == "H5"}, {
        r["email"].lower() for r in rows if r["lead_class"] == "T1"}
    assert not (h5 & t1), "H5 and T1 must be disjoint"

    # ---------------- write ----------------
    NEW = ["lead_class", "lead_type", "lead_class_reason", "t3_branch", "priority_account",
           "canonical_account", "contactability", "seniority", "salutation_first",
           "booth_network_send", "email_owner", "linkedin_owner", "send_hold",
           "is_customer", "duplicate_of"]
    out_hdr = NEW[:4] + hdr + NEW[4:]
    # The sheet's user-facing name for the classification is "Tier".
    DISPLAY = {"lead_class": "Tier", "lead_class_reason": "Tier_reason"}
    disp = lambda h: DISPLAY.get(h, h)

    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    ws2.title = "Master contacts"
    ws2.append([disp(h) for h in out_hdr])
    for r in rows:
        ws2.append([r.get(h, "") for h in out_hdr])

    head_fill = PatternFill("solid", fgColor="1F3864")
    for i, h in enumerate(out_hdr, start=1):
        cell = ws2.cell(row=1, column=i)
        cell.value = disp(h)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws2.column_dimensions[get_column_letter(i)].width = min(
            42, max(12, len(h) + 2, *(len(str(r.get(h, ""))) for r in rows[:80]) if rows else 12)
        )
    CLASS_FILL = {
        "H5": "C00000", "T1": "ED7D31", "T2": "FFC000", "T3": "70AD47",
        "GA": "BFBFBF", "DEFERRED": "D9D9D9", "STOP": "808080", "ANON": "E7E6E6",
        "ORGANISER": "9DC3E6", "OWN_TEAM": "F2F2F2", "TEST": "F2F2F2",
        "DUPLICATE": "F2F2F2", "UNREACHABLE": "FFE699",
    }
    for i, r in enumerate(rows, start=2):
        cell = ws2.cell(row=i, column=1)
        cell.fill = PatternFill("solid", fgColor=CLASS_FILL[r["lead_class"]])
        cell.font = Font(bold=True, size=10,
                         color="FFFFFF" if r["lead_class"] in ("H5", "STOP") else "000000")
    ws2.freeze_panes = "D2"
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(out_hdr))}{len(rows)+1}"
    wb2.save(OUT_XLSX)

    with open(OUT_CSV, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["Tier", "lead_type", "t3_branch", "priority_account", "first_name",
                    "last_name", "company", "canonical_account", "job_title", "seniority", "email",
                    "linkedin_url", "email_owner", "linkedin_owner", "send_hold",
                    "booth_network_send", "Tier_reason"])
        order = {k: i for i, k in enumerate(
            ["H5", "T1", "T2", "T3", "UNREACHABLE", "GA", "DEFERRED", "ORGANISER",
             "STOP", "ANON", "DUPLICATE", "OWN_TEAM", "TEST"])}
        for r in sorted(rows, key=lambda r: (order[r["lead_class"]], r["canonical_account"].lower(),
                                             r["last_name"].lower())):
            w.writerow([r[k] for k in ("lead_class", "lead_type", "t3_branch", "priority_account",
                                       "first_name", "last_name", "company", "canonical_account",
                                       "job_title", "seniority", "email", "linkedin_url",
                                       "email_owner", "linkedin_owner", "send_hold",
                                       "booth_network_send", "lead_class_reason")])

    print(f"\n{len(fixlog)} raw-cell repairs applied:")
    for line in fixlog:
        print("  -", line)
    print("\nlead_class partition:")
    for k, n in sorted(c.items(), key=lambda kv: (order.get(kv[0], 99) if (order := {
            "H5": 0, "T1": 1, "T2": 2, "T3": 3, "UNREACHABLE": 4, "GA": 5, "DEFERRED": 6,
            "ORGANISER": 7, "STOP": 8, "ANON": 9, "DUPLICATE": 10, "OWN_TEAM": 11, "TEST": 12}) else 99, kv[0])):
        print(f"  {k:12s} {n:>3}")
    print(f"\n  contactable leads (H5+T1+T2+T3) = {c['H5']+c['T1']+c['T2']+c['T3']}")
    print(f"  booth_network_send TRUE          = {sum(1 for r in rows if r['booth_network_send']=='TRUE')}")
    print(f"  send_hold != none                = {sum(1 for r in rows if r['send_hold']!='none')}")
    print(f"\nwrote {OUT_XLSX}\nwrote {OUT_CSV}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
